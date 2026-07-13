from io import BytesIO
from unittest.mock import AsyncMock
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook

import app.services.analytics_service as analytics_module
from app.core.official_identity import MUNICIPALITY_CNPJ, MUNICIPALITY_NAME

from app.services.analytics_service import (
    ANALYTICS_EXPORT_NO_CACHE_HEADERS,
    AnalyticsService,
    calculate_consumption_l_100km,
    calculate_driver_risk_score,
    calculate_tco_per_km,
    calculate_variance_percentage,
    neutralize_analytics_spreadsheet_text,
)


def test_calculate_consumption_l_100km():
    assert calculate_consumption_l_100km(50, 500) == 10
    assert calculate_consumption_l_100km(20, 0) is None


def test_calculate_tco_per_km():
    assert calculate_tco_per_km(300, 200, 100, 500) == 1.2
    assert calculate_tco_per_km(100, 0, 0, 0) is None


def test_calculate_driver_risk_score():
    score = calculate_driver_risk_score(fines_count=4, claims_count=2, anomalies_count=3)
    assert score == pytest.approx(2.8)


def test_calculate_variance_percentage():
    assert calculate_variance_percentage(12, 10) == 20
    assert calculate_variance_percentage(8, 10) == -20
    assert calculate_variance_percentage(10, 0) is None


@pytest.mark.parametrize(
    "payload",
    [
        "=1+1",
        "+SUM(A1:A2)",
        "-2+3",
        "@SUM(A1:A2)",
        " \t=cmd|' /C calc'!A0",
    ],
)
def test_analytics_formula_injection_is_neutralized(payload):
    assert neutralize_analytics_spreadsheet_text(payload) == f"'{payload}"


def _analytics_insight(**overrides):
    insight = {
        "metric": "consumption_l_100km",
        "severity": "HIGH",
        "current_value": 14.75,
        "category_average": 10.25,
        "variance_percentage": 43.9,
        "message": "Consumo acima da média da categoria.",
        "recommended_action": "Agendar inspeção mecânica preventiva",
    }
    insight.update(overrides)
    return insight


@pytest.mark.asyncio
async def test_analytics_xlsx_export_is_real_official_and_formula_safe():
    formula_payloads = ["=1+1", "+SUM(A1:A2)", "-2+3", "@SUM(A1:A2)", " \t=cmd|' /C calc'!A0"]
    insights = [
        _analytics_insight(
            metric=payload,
            recommended_action="\t+cmd|' /C calc'!A0" if index == 0 else "Ação segura",
        )
        for index, payload in enumerate(formula_payloads)
    ]
    service = AnalyticsService(db=None)
    service.insights = AsyncMock(return_value=insights)

    response = await service.export(30, "xlsx", vehicle_type="SEDAN")

    assert response.status_code == 200
    assert response.body.startswith(b"PK")
    assert response.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert response.headers["cache-control"] == ANALYTICS_EXPORT_NO_CACHE_HEADERS["Cache-Control"]
    assert response.headers["pragma"] == "no-cache"
    assert response.headers["expires"] == "0"
    assert response.headers["x-content-type-options"] == "nosniff"
    assert response.headers["content-disposition"] == 'attachment; filename="analytics-30d.xlsx"'
    service.insights.assert_awaited_once_with(30, vehicle_type="SEDAN", organization_id=None)

    with ZipFile(BytesIO(response.body)) as archive:
        members = set(archive.namelist())
        assert "[Content_Types].xml" in members
        assert "xl/workbook.xml" in members
        assert "xl/worksheets/sheet1.xml" in members
        assert any(member.startswith("xl/media/") for member in members)

    workbook = load_workbook(BytesIO(response.body), data_only=False)
    worksheet = workbook["Análises"]
    assert worksheet["B1"].value == MUNICIPALITY_NAME.upper()
    assert worksheet["A4"].value == "RELATÓRIO DE ANÁLISES DA FROTA"
    assert "UTC-03:00" in worksheet["A6"].value
    assert [worksheet.cell(row=8, column=index).value for index in range(1, 7)] == [
        "metric",
        "severity",
        "current_value",
        "category_average",
        "variance_percentage",
        "recommended_action",
    ]
    for row_index, payload in enumerate(formula_payloads, start=9):
        cell = worksheet.cell(row=row_index, column=1)
        assert cell.value == f"'{payload}"
        assert cell.data_type != "f"
    action_cell = worksheet.cell(row=9, column=6)
    assert action_cell.value == "'\t+cmd|' /C calc'!A0"
    assert action_cell.data_type != "f"
    assert worksheet.cell(row=9, column=3).value == 14.75
    metadata = workbook["Informações"]
    assert metadata["B1"].value == MUNICIPALITY_NAME
    assert metadata["B3"].value == MUNICIPALITY_CNPJ
    assert "UTC-03:00" in metadata["B7"].value
    workbook.close()


@pytest.mark.asyncio
async def test_analytics_pdf_export_is_real_official_and_contains_dataset(monkeypatch):
    captured_paragraphs = []
    original_paragraph = analytics_module.Paragraph

    def recording_paragraph(text, *args, **kwargs):
        captured_paragraphs.append(str(text))
        return original_paragraph(text, *args, **kwargs)

    monkeypatch.setattr(analytics_module, "Paragraph", recording_paragraph)
    service = AnalyticsService(db=None)
    service.insights = AsyncMock(return_value=[_analytics_insight()])

    response = await service.export(90, "pdf")

    assert response.status_code == 200
    assert response.body.startswith(b"%PDF-")
    assert response.media_type == "application/pdf"
    assert b"/Subtype /Image" in response.body
    assert b"/FontFile2" in response.body
    assert response.headers["cache-control"] == ANALYTICS_EXPORT_NO_CACHE_HEADERS["Cache-Control"]
    assert response.headers["pragma"] == "no-cache"
    assert response.headers["expires"] == "0"
    assert response.headers["x-content-type-options"] == "nosniff"
    assert response.headers["content-disposition"] == 'attachment; filename="analytics-90d.pdf"'
    service.insights.assert_awaited_once_with(90, vehicle_type=None, organization_id=None)

    rendered_content = "\n".join(captured_paragraphs)
    assert MUNICIPALITY_NAME.upper() in rendered_content
    assert "RELATÓRIO DE ANÁLISES DA FROTA" in rendered_content
    assert "últimos 90 dias" in rendered_content
    assert "UTC-03:00" in rendered_content
    assert "Consumo de combustível (L/100 km)" in rendered_content
    assert "Consumo acima da média da categoria." in rendered_content
    assert "Agendar inspeção mecânica preventiva" in rendered_content
