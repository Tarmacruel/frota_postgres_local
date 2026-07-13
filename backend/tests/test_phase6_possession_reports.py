from __future__ import annotations

import importlib.util
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import AsyncMock, MagicMock
from uuid import uuid4

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook
from pydantic import ValidationError

from app.api.routes import possession as possession_routes
from app.models.user import UserRole
from app.repositories.possession_report_repository import PossessionReportRepository
from app.schemas.possession_report import (
    PossessionReportMode,
    PossessionReportFilters,
    PossessionReportPreferenceIn,
    PossessionReportPreset,
    PossessionReportRequest,
)
from app.services.possession_report_registry import (
    REPORT_COLUMN_BY_KEY,
    ReportDataClassification,
    ReportRowContext,
    preset_columns,
)
from app.services.possession_report_service import (
    REPORT_NO_CACHE_HEADERS,
    PossessionReportService,
    PreparedCell,
    PreparedReport,
    neutralize_spreadsheet_text,
)


def _user(role: UserRole) -> SimpleNamespace:
    return SimpleNamespace(
        id=uuid4(),
        name=f"Usuário {role.value}",
        email=f"{role.value.lower()}@example.test",
        role=role,
        organization_id=None,
    )


def test_phase6_migration_has_confirmed_parent_and_scoped_preference_payload():
    migration_path = Path(__file__).parents[1] / "alembic" / "versions" / "0040_add_user_report_preferences.py"
    spec = importlib.util.spec_from_file_location("phase6_migration", migration_path)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    assert module.revision == "0040_report_preferences"
    assert module.down_revision == "0039_possession_trips"


def test_summary_preset_is_minimized_and_registry_is_typed():
    columns = preset_columns(UserRole.ADMIN, PossessionReportMode.POSSESSION, PossessionReportPreset.SUMMARY)
    keys = [column.key for column in columns]
    assert keys == [
        "possession_number",
        "vehicle_plate",
        "driver_name",
        "possession_start",
        "possession_end",
        "possession_status",
        "trip_count",
        "total_trip_kilometers",
    ]
    assert not {"driver_document", "driver_contact", "capture_latitude", "return_request_id"} & set(keys)
    assert all(column.extractor and column.roles and column.modes for column in REPORT_COLUMN_BY_KEY.values())

    standard_keys = {
        column.key
        for column in preset_columns(
            UserRole.PADRAO,
            PossessionReportMode.POSSESSION,
            PossessionReportPreset.SUMMARY,
        )
    }
    assert "driver_name" not in standard_keys


def test_metadata_varies_by_role_and_never_sends_forbidden_columns():
    standard = PossessionReportService.metadata(_user(UserRole.PADRAO))
    production = PossessionReportService.metadata(_user(UserRole.PRODUCAO))
    admin = PossessionReportService.metadata(_user(UserRole.ADMIN))
    standard_mode = standard["modes"][0]
    production_mode = production["modes"][0]
    admin_mode = admin["modes"][0]
    assert "driver_document" not in {column["key"] for column in standard_mode["columns"]}
    assert "driver_name" not in {column["key"] for column in standard_mode["columns"]}
    assert "driver_document" in {column["key"] for column in production_mode["columns"]}
    assert "return_request_id" not in {column["key"] for column in production_mode["columns"]}
    assert "return_request_id" in {column["key"] for column in admin_mode["columns"]}
    assert {preset["key"] for preset in standard_mode["presets"]} == {"SUMMARY", "CUSTOM"}
    assert standard["can_export_xlsx"] is False
    assert production["can_export_xlsx"] is True


def test_backend_rejects_unknown_restricted_and_mode_incompatible_columns():
    service = PossessionReportService(AsyncMock())
    with pytest.raises(HTTPException) as restricted:
        service._resolve_columns(
            PossessionReportRequest(mode="POSSESSION", preset="CUSTOM", column_keys=["driver_document"]),
            UserRole.PADRAO,
        )
    assert restricted.value.status_code == 403
    assert restricted.value.detail["code"] == "REPORT_COLUMN_FORBIDDEN"
    with pytest.raises(HTTPException) as unknown:
        service._resolve_columns(
            PossessionReportRequest(mode="POSSESSION", preset="CUSTOM", column_keys=["sql_expression"]),
            UserRole.ADMIN,
        )
    assert unknown.value.detail["code"] == "REPORT_COLUMN_UNKNOWN"
    with pytest.raises(HTTPException) as mismatch:
        service._resolve_columns(
            PossessionReportRequest(mode="POSSESSION", preset="CUSTOM", column_keys=["trip_origin"]),
            UserRole.ADMIN,
        )
    assert mismatch.value.detail["code"] == "REPORT_COLUMN_MODE_MISMATCH"


def test_legacy_row_extractors_do_not_fabricate_return_confirmation():
    possession = SimpleNamespace(
        end_date=datetime(2026, 7, 13, tzinfo=timezone.utc),
        return_confirmations=[],
    )
    context = ReportRowContext(possession=possession)
    value = REPORT_COLUMN_BY_KEY["return_status"].extractor(context)
    assert value == "ENCERRADA_SEM_CONFIRMAÇÃO_VERSIONADA"


@pytest.mark.parametrize(
    ("raw", "display"),
    [
        ("ATIVA", "Ativa"),
        ("EM_ANDAMENTO", "Em andamento"),
        ("DEVOLUÇÃO_CONFIRMADA", "Devolução confirmada"),
        ("ENCERRADA_SEM_CONFIRMAÇÃO_VERSIONADA", "Encerrada sem confirmação eletrônica"),
    ],
)
def test_official_report_status_values_are_humanized(raw, display):
    column = REPORT_COLUMN_BY_KEY["return_status"]
    prepared = PossessionReportService._prepare_cell(
        column,
        ReportRowContext(
            possession=SimpleNamespace(
                end_date=None,
                return_confirmations=[],
            )
        ),
    )
    assert PossessionReportService._status_display(raw) == display
    assert "_" not in PossessionReportService._status_display(raw)
    assert "_" not in prepared.display


def test_pdf_and_xlsx_use_same_prepared_column_order_and_dataset():
    request = PossessionReportRequest(mode="POSSESSION", preset="CUSTOM", column_keys=["vehicle_plate", "driver_name"])
    prepared = PreparedReport(
        request=request,
        columns=(REPORT_COLUMN_BY_KEY["vehicle_plate"], REPORT_COLUMN_BY_KEY["driver_name"]),
        rows=((PreparedCell(raw="ABC1D23", display="ABC1D23"), PreparedCell(raw="Condutor teste", display="Condutor teste")),),
    )
    service = PossessionReportService(AsyncMock())
    pdf = service._build_pdf(prepared, _user(UserRole.ADMIN))
    xlsx = service._build_xlsx(prepared, _user(UserRole.ADMIN))
    workbook = load_workbook(filename=__import__("io").BytesIO(xlsx), read_only=True)
    worksheet = workbook["Posses"]
    assert pdf.startswith(b"%PDF-")
    assert b"/Subtype /Image" in pdf
    assert [worksheet.cell(1, index).value for index in (1, 2)] == ["Placa", "Condutor"]
    assert [worksheet.cell(2, index).value for index in (1, 2)] == ["ABC1D23", "Condutor teste"]
    metadata = workbook["Metadados"]
    assert metadata["B1"].value == "Prefeitura Municipal de Teixeira de Freitas"
    assert metadata["B3"].value == "13.650.403/0001-28"
    assert workbook.properties.creator == "Prefeitura Municipal de Teixeira de Freitas"


def test_document_filter_summary_never_repeats_search_text_or_internal_ids():
    vehicle_id = uuid4()
    driver_id = uuid4()
    request = PossessionReportRequest(
        filters=PossessionReportFilters(
            vehicle_id=vehicle_id,
            driver_id=driver_id,
            search="529.982.247-25",
            has_return=True,
        )
    )

    summary = PossessionReportService._plain_filter_summary(request)

    assert str(vehicle_id) not in summary
    assert str(driver_id) not in summary
    assert "529.982.247-25" not in summary
    assert "Veículo: selecionado" in summary
    assert "Busca textual: aplicada" in summary
    assert "Devolução registrada: Sim" in summary


@pytest.mark.parametrize("payload", ["=1+1", "+SUM(A1:A2)", "-2+3", "@IMPORT", "\t=cmd"])
def test_formula_injection_is_neutralized(payload):
    assert neutralize_spreadsheet_text(payload).startswith("'")


def test_formula_injection_is_neutralized_in_generated_workbook():
    request = PossessionReportRequest(mode="POSSESSION", preset="CUSTOM", column_keys=["driver_name"])
    prepared = PreparedReport(
        request=request,
        columns=(REPORT_COLUMN_BY_KEY["driver_name"],),
        rows=((PreparedCell(raw="=HYPERLINK(\"https://invalid.test\")", display="texto"),),),
    )
    xlsx = PossessionReportService(AsyncMock())._build_xlsx(prepared, _user(UserRole.ADMIN))
    workbook = load_workbook(filename=__import__("io").BytesIO(xlsx), read_only=True)
    assert workbook["Posses"]["A2"].value.startswith("'=")


def test_excessive_server_filter_interval_is_rejected():
    now = datetime.now(timezone.utc)
    with pytest.raises(ValidationError):
        PossessionReportFilters(date_from=now - timedelta(days=367), date_to=now)


@pytest.mark.asyncio
async def test_prepare_forwards_normalized_filters_to_bounded_server_query():
    service = PossessionReportService(AsyncMock())
    service.reports.load = AsyncMock(return_value=[])
    now = datetime.now(timezone.utc)
    request = PossessionReportRequest(
        filters=PossessionReportFilters(
            date_from=now - timedelta(days=7),
            date_to=now,
            search="ABC1D23",
            has_return_confirmation=True,
        )
    )
    prepared = await service.prepare(request, _user(UserRole.ADMIN), row_limit=42)
    assert prepared.row_count == 0
    kwargs = service.reports.load.await_args.kwargs
    assert kwargs["filters"] is request.filters
    assert kwargs["limit"] == 42
    assert kwargs["include_operational_search"] is True

    service.reports.load.reset_mock()
    await service.prepare(request, _user(UserRole.PADRAO), row_limit=42)
    assert service.reports.load.await_args.kwargs["include_operational_search"] is False


@pytest.mark.asyncio
async def test_standard_report_rejects_driver_filter_before_querying():
    service = PossessionReportService(AsyncMock())
    service.reports.load = AsyncMock()
    request = PossessionReportRequest(filters=PossessionReportFilters(driver_id=uuid4()))

    with pytest.raises(HTTPException) as exc:
        await service.prepare(request, _user(UserRole.PADRAO), row_limit=42)

    assert exc.value.status_code == 403
    assert exc.value.detail["code"] == "REPORT_FILTER_FORBIDDEN"
    service.reports.load.assert_not_awaited()


@pytest.mark.asyncio
async def test_standard_report_search_predicate_excludes_driver_and_free_text():
    db = AsyncMock()
    result = MagicMock()
    result.scalars.return_value.unique.return_value.all.return_value = []
    db.execute.return_value = result

    await PossessionReportRepository(db).load(
        mode=PossessionReportMode.POSSESSION,
        filters=PossessionReportFilters(search="Pessoa protegida"),
        organization_id=None,
        limit=10,
        include_operational_search=False,
    )

    where_clause = str(db.execute.await_args.args[0].whereclause)
    assert "plate" in where_clause
    assert "driver_name" not in where_clause
    assert "observation" not in where_clause


@pytest.mark.asyncio
async def test_standard_role_cannot_export_xlsx_before_loading_dataset():
    service = PossessionReportService(AsyncMock())
    service.reports.load = AsyncMock()
    service.audit.record = AsyncMock()
    with pytest.raises(HTTPException) as exc:
        await service.render_xlsx(PossessionReportRequest(), _user(UserRole.PADRAO))
    assert exc.value.status_code == 403
    assert exc.value.detail["code"] == "REPORT_OPERATIONAL_EXPORT_FORBIDDEN"
    service.reports.load.assert_not_awaited()
    assert service.audit.record.await_args.kwargs["details"]["outcome"] == "FAILURE"


@pytest.mark.asyncio
async def test_preview_audit_contains_metadata_without_report_rows(monkeypatch):
    service = PossessionReportService(AsyncMock())
    prepared = PreparedReport(
        request=PossessionReportRequest(),
        columns=tuple(preset_columns(UserRole.ADMIN, PossessionReportMode.POSSESSION, PossessionReportPreset.SUMMARY)),
        rows=((PreparedCell(raw="segredo-da-linha", display="segredo-da-linha"),) * 8,),
    )
    service.prepare = AsyncMock(return_value=prepared)
    service._build_pdf = lambda *_args: b"%PDF-test"
    service.audit.record = AsyncMock()
    await service.render_pdf(PossessionReportRequest(), _user(UserRole.ADMIN))
    details = service.audit.record.await_args.kwargs["details"]
    serialized = str(details)
    assert details["record_count"] == 1
    assert details["outcome"] == "SUCCESS"
    assert "segredo-da-linha" not in serialized
    assert "column_keys" in details


@pytest.mark.asyncio
async def test_preference_persists_only_mode_preset_and_authorized_keys():
    user = _user(UserRole.ADMIN)
    service = PossessionReportService(AsyncMock())
    service.preferences.upsert = AsyncMock(return_value=SimpleNamespace())
    service.audit.record = AsyncMock()
    result = await service.update_preference(
        PossessionReportPreferenceIn(
            mode="POSSESSION",
            preset="CUSTOM",
            column_keys=["vehicle_plate", "driver_name"],
        ),
        user,
    )
    stored = service.preferences.upsert.await_args.kwargs["config"]
    assert stored == {
        "mode": "POSSESSION",
        "preset": "CUSTOM",
        "column_keys": ["vehicle_plate", "driver_name"],
    }
    assert "filter" not in str(stored).lower()
    assert result["sanitized"] is False
    service.db.commit.assert_awaited_once()


@pytest.mark.asyncio
async def test_report_route_has_no_store_and_protected_content_disposition(monkeypatch):
    class FakeReportService:
        def __init__(self, _db):
            pass

        async def render_pdf(self, *_args):
            return b"%PDF-test", "relatorio-posses.pdf"

    monkeypatch.setattr(possession_routes, "PossessionReportService", FakeReportService)
    response = await possession_routes.preview_possession_report_pdf(
        PossessionReportRequest(),
        db=AsyncMock(),
        current_user=_user(UserRole.ADMIN),
    )
    assert response.headers["cache-control"].startswith("private, no-store")
    assert response.headers["content-disposition"] == 'inline; filename="relatorio-posses.pdf"'
    assert response.media_type == "application/pdf"
    assert REPORT_NO_CACHE_HEADERS["X-Content-Type-Options"] == "nosniff"


def test_security_metadata_columns_are_admin_manual_only():
    technical = REPORT_COLUMN_BY_KEY["return_request_id"]
    assert technical.classification == ReportDataClassification.SECURITY_METADATA
    assert technical.roles == frozenset({UserRole.ADMIN})
    assert technical.manual_only is True
    assert technical.presets == frozenset()
