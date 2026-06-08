from __future__ import annotations

import io
import json
import re
import unicodedata
from collections import defaultdict
from datetime import date, datetime, time, timedelta, timezone
from decimal import Decimal, InvalidOperation, ROUND_FLOOR
from pathlib import Path
from uuid import UUID, uuid4

from fastapi import HTTPException, UploadFile, status
from sqlalchemy import delete, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.organization_scope import is_production_user, production_scope_is_empty, scoped_organization_id
from app.models.master_data import Organization
from app.models.payment_process import (
    PaymentChecklistStatus,
    PaymentContract,
    PaymentContractAmendment,
    PaymentContractStatus,
    PaymentProcess,
    PaymentProcessChecklistItem,
    PaymentProcessKind,
    PaymentProcessReference,
    PaymentProcessReferenceType,
    PaymentProcessStage,
    PaymentProcessStageEvent,
    PaymentSupplier,
)
from app.models.user import User
from app.repositories.payment_process_repository import PaymentProcessRepository
from app.schemas.payment_process import (
    PaymentContractAmendmentCreate,
    PaymentContractCreate,
    PaymentContractUpdate,
    PaymentProcessChecklistUpdate,
    PaymentProcessCreate,
    PaymentProcessDelete,
    PaymentProcessStageUpdate,
    PaymentProcessUpdate,
    PaymentSupplierCreate,
    PaymentSupplierUpdate,
)
from app.schemas.common import PaginatedResponse, build_pagination
from app.services.audit_service import AuditService

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError:  # pragma: no cover - dependency is pinned in requirements.
    Workbook = None
    load_workbook = None
    InvalidFileException = Exception


MAX_IMPORT_ROWS = 4000
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
MISSING_TEXTS = {"", "-", "--", "N/A", "NA", "NULL", "NONE", "SELECIONE"}
TEMPLATE_HEADERS = [
    "Tipo Processo",
    "Nº do Processo",
    "Sistema",
    "Status",
    "Fatura",
    "NF",
    "Tipo Nota",
    "Unidade",
    "Emissão NF",
    "Tipo",
    "Valor",
    "Fornecedor",
    "Contrato",
    "Saldo Restante",
    "Localização",
    "Observações",
]


TEMPLATE_HEADERS = [
    "Tipo Processo",
    "Número do Processo",
    "Sistema",
    "Status",
    "Etapa",
    "Fatura",
    "NF",
    "Tipo Nota",
    "Unidade",
    "Emissão NF",
    "Competência",
    "Vencimento",
    "Tipo",
    "Valor",
    "Fornecedor",
    "CNPJ Fornecedor",
    "Contrato",
    "Tipo Contrato",
    "Objeto Contrato",
    "Inicio Vigencia",
    "Fim Vigencia",
    "Valor Inicial Contrato",
    "Valor Atualizado Contrato",
    "Saldo Restante",
    "Empenho",
    "Data Empenho",
    "Liquidação",
    "Data Liquidação",
    "Ordem Pagamento",
    "Data Ordem Pagamento",
    "Data Pagamento",
    "Responsável Etapa",
    "Localização",
    "Observações",
]

STAGE_LABELS = {
    PaymentProcessStage.ASSEMBLY: "Montagem",
    PaymentProcessStage.REVIEW: "Conferência",
    PaymentProcessStage.COMMITMENT: "Empenho",
    PaymentProcessStage.LIQUIDATION: "Liquidação",
    PaymentProcessStage.PAYMENT: "Pagamento",
    PaymentProcessStage.PAID: "Pago",
    PaymentProcessStage.ARCHIVED: "Arquivado",
    PaymentProcessStage.RETURNED: "Devolvido",
    PaymentProcessStage.CANCELLED: "Cancelado",
}
STAGE_ORDER = [
    PaymentProcessStage.ASSEMBLY,
    PaymentProcessStage.REVIEW,
    PaymentProcessStage.COMMITMENT,
    PaymentProcessStage.LIQUIDATION,
    PaymentProcessStage.PAYMENT,
    PaymentProcessStage.PAID,
    PaymentProcessStage.ARCHIVED,
]
TERMINAL_STAGES = {PaymentProcessStage.CANCELLED, PaymentProcessStage.RETURNED}
CHECKLIST_DEFAULTS = {
    PaymentProcessStage.ASSEMBLY: ["Fatura/NF identificada", "Contrato vinculado", "Competência informada"],
    PaymentProcessStage.REVIEW: ["Dados conferidos", "Atesto registrado", "Pendências justificadas"],
    PaymentProcessStage.COMMITMENT: ["Empenho informado"],
    PaymentProcessStage.LIQUIDATION: ["Liquidação informada"],
    PaymentProcessStage.PAYMENT: ["Ordem de pagamento informada", "Vencimento acompanhado"],
    PaymentProcessStage.PAID: ["Pagamento registrado"],
    PaymentProcessStage.ARCHIVED: ["Processo arquivado"],
}


class PaymentProcessService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.repository = PaymentProcessRepository(db)
        self.audit = AuditService(db)

    async def list(
        self,
        *,
        page: int,
        limit: int,
        current_user: User,
        kind: PaymentProcessKind | None = None,
        stage: PaymentProcessStage | None = None,
        status_filter: str | None = None,
        organization_id: UUID | None = None,
        supplier_id: UUID | None = None,
        contract_id: UUID | None = None,
        competence_month: date | None = None,
        due_from: date | None = None,
        due_to: date | None = None,
        search: str | None = None,
    ) -> PaginatedResponse[dict]:
        if production_scope_is_empty(current_user):
            return PaginatedResponse[dict](data=[], pagination=build_pagination(page, limit, 0))

        scoped_org_id = scoped_organization_id(current_user, organization_id)
        records, total = await self.repository.list_paginated(
            page=page,
            limit=limit,
            kind=kind,
            stage=stage,
            status_filter=status_filter,
            organization_id=scoped_org_id,
            supplier_id=supplier_id,
            contract_id=contract_id,
            competence_month=self._month_start(competence_month),
            due_from=due_from,
            due_to=due_to,
            search=search,
        )
        return PaginatedResponse[dict](
            data=[self._serialize(item) for item in records],
            pagination=build_pagination(page, limit, total),
        )

    async def get(self, process_id: UUID, *, current_user: User) -> dict:
        record = await self.repository.get_by_id(process_id)
        if not record:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Processo de pagamento não encontrado")
        self._ensure_process_visible(record, current_user)
        return self._serialize(record)

    async def create(self, data: PaymentProcessCreate, current_user: User) -> dict:
        payload = data.model_dump()
        references = payload.pop("references", [])
        payload = await self._prepare_process_payload(payload, current_user=current_user)
        now = datetime.now(timezone.utc)
        record = PaymentProcess(
            import_key=self._manual_import_key(payload),
            created_by_user_id=current_user.id,
            updated_by_user_id=current_user.id,
            created_at=now,
            updated_at=now,
        )
        self._assign_process_fields(record, payload)
        self.repository.db.add(record)

        try:
            await self.repository.flush()
            await self._normalize_legacy_contract_values({record.contract_id})
            self._replace_references(record, references)
            self._ensure_default_checklist(record)
            await self.audit.record(
                actor=current_user,
                action="CREATE",
                entity_type="PAYMENT_PROCESS",
                entity_id=record.id,
                entity_label=record.process_number,
                details={"stage": record.stage, "amount": record.amount},
            )
            await self.db.commit()
        except IntegrityError as exc:
            await self.db.rollback()
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Não foi possível criar o processo") from exc

        return await self.get(record.id, current_user=current_user)

    async def update(self, process_id: UUID, data: PaymentProcessUpdate, current_user: User) -> dict:
        record = await self.repository.get_by_id(process_id)
        if not record:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Processo de pagamento não encontrado")
        self._ensure_process_visible(record, current_user)

        payload = data.model_dump(exclude_unset=True)
        references = payload.pop("references", None)
        previous_contract_id = record.contract_id
        payload = await self._prepare_process_payload(payload, current_user=current_user, existing=record)
        previous = self._serialize(record)
        self._assign_process_fields(record, payload)
        record.updated_by_user_id = current_user.id
        record.updated_at = datetime.now(timezone.utc)
        if references is not None:
            self._replace_references(record, references)
        self._ensure_default_checklist(record)
        await self.repository.flush()
        await self._normalize_legacy_contract_values({previous_contract_id, record.contract_id})

        try:
            await self.audit.record(
                actor=current_user,
                action="UPDATE",
                entity_type="PAYMENT_PROCESS",
                entity_id=record.id,
                entity_label=record.process_number,
                details={
                    "before_stage": previous.get("stage"),
                    "after_stage": record.stage,
                    "before_amount": previous.get("amount"),
                    "after_amount": record.amount,
                },
            )
            await self.db.commit()
        except IntegrityError as exc:
            await self.db.rollback()
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Não foi possível atualizar o processo") from exc

        return await self.get(record.id, current_user=current_user)

    async def delete(self, process_id: UUID, data: PaymentProcessDelete, current_user: User) -> dict:
        record = await self.repository.get_by_id(process_id)
        if not record:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Processo de pagamento não encontrado")
        self._ensure_process_visible(record, current_user)

        details = self._delete_audit_details(record, reason=data.reason)
        contract_id = record.contract_id

        try:
            await self.audit.record(
                actor=current_user,
                action="DELETE",
                entity_type="PAYMENT_PROCESS",
                entity_id=record.id,
                entity_label=record.process_number,
                details=details,
            )
            await self.repository.delete(record)
            await self._normalize_legacy_contract_values({contract_id})
            await self.db.commit()
        except IntegrityError as exc:
            await self.db.rollback()
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Não foi possível excluir o processo") from exc

        return {"message": "Processo de pagamento excluído"}

    async def change_stage(self, process_id: UUID, data: PaymentProcessStageUpdate, current_user: User) -> dict:
        record = await self.repository.get_by_id(process_id)
        if not record:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Processo de pagamento não encontrado")
        self._ensure_process_visible(record, current_user)

        previous_stage = record.stage
        record.stage = data.stage
        record.status_note = data.comment or record.status_note
        record.updated_by_user_id = current_user.id
        record.updated_at = datetime.now(timezone.utc)
        self._ensure_default_checklist(record)
        alerts = self._process_alerts(record)
        event = PaymentProcessStageEvent(
            process_id=record.id,
            from_stage=previous_stage,
            to_stage=data.stage,
            comment=data.comment,
            alerts_snapshot=json.dumps(alerts, ensure_ascii=True),
            created_by_user_id=current_user.id,
        )
        self.db.add(event)

        try:
            await self.audit.record(
                actor=current_user,
                action="STAGE",
                entity_type="PAYMENT_PROCESS",
                entity_id=record.id,
                entity_label=record.process_number,
                details={"from_stage": previous_stage, "to_stage": data.stage, "alerts": alerts},
            )
            await self.db.commit()
        except IntegrityError as exc:
            await self.db.rollback()
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Não foi possível avançar a etapa") from exc

        return await self.get(record.id, current_user=current_user)

    async def update_checklist(self, process_id: UUID, data: PaymentProcessChecklistUpdate, current_user: User) -> dict:
        record = await self.repository.get_by_id(process_id)
        if not record:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Processo de pagamento não encontrado")
        self._ensure_process_visible(record, current_user)

        await self.db.execute(delete(PaymentProcessChecklistItem).where(PaymentProcessChecklistItem.process_id == record.id))
        record.checklist_items.clear()
        now = datetime.now(timezone.utc)
        for item in data.items:
            record.checklist_items.append(
                PaymentProcessChecklistItem(
                    stage=item.stage,
                    item_label=item.item_label,
                    status=item.status,
                    notes=item.notes,
                    updated_by_user_id=current_user.id,
                    updated_at=now,
                    created_at=now,
                )
            )
        self._ensure_default_checklist(record)
        record.updated_by_user_id = current_user.id
        record.updated_at = now

        try:
            await self.audit.record(
                actor=current_user,
                action="CHECKLIST",
                entity_type="PAYMENT_PROCESS",
                entity_id=record.id,
                entity_label=record.process_number,
                details={"items": len(data.items)},
            )
            await self.db.commit()
        except IntegrityError as exc:
            await self.db.rollback()
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Não foi possível atualizar o checklist") from exc

        return await self.get(record.id, current_user=current_user)

    async def dashboard(self, *, current_user: User) -> dict:
        if production_scope_is_empty(current_user):
            return self._empty_dashboard()

        organization_id = scoped_organization_id(current_user)
        records = await self.repository.list_for_export(organization_id=organization_id, limit=10000)
        today = date.today()
        soon_limit = today + timedelta(days=7)
        total_amount = Decimal("0")
        paid_amount = Decimal("0")
        pending_amount = Decimal("0")
        alerts_count = 0
        stage_stats = {stage: {"count": 0, "amount": Decimal("0")} for stage in PaymentProcessStage}

        for record in records:
            amount = Decimal(str(record.amount or 0))
            total_amount += amount
            if record.stage in {PaymentProcessStage.PAID, PaymentProcessStage.ARCHIVED}:
                paid_amount += amount
            elif record.stage not in TERMINAL_STAGES:
                pending_amount += amount
            stage_stats[record.stage]["count"] += 1
            stage_stats[record.stage]["amount"] += amount
            alerts_count += len(self._process_alerts(record))

        contracts = await self.list_contracts(current_user=current_user)
        open_processes = len([record for record in records if record.stage not in TERMINAL_STAGES | {PaymentProcessStage.PAID, PaymentProcessStage.ARCHIVED}])
        overdue_processes = len([record for record in records if record.due_date and record.due_date < today and record.stage not in TERMINAL_STAGES | {PaymentProcessStage.PAID, PaymentProcessStage.ARCHIVED}])
        due_soon_processes = len(
            [
                record
                for record in records
                if record.due_date and today <= record.due_date <= soon_limit and record.stage not in TERMINAL_STAGES | {PaymentProcessStage.PAID, PaymentProcessStage.ARCHIVED}
            ]
        )
        return {
            "total_processes": len(records),
            "open_processes": open_processes,
            "overdue_processes": overdue_processes,
            "due_soon_processes": due_soon_processes,
            "total_amount": total_amount,
            "paid_amount": paid_amount,
            "pending_amount": pending_amount,
            "alerts_count": alerts_count,
            "stages": [
                {"stage": stage, "label": STAGE_LABELS[stage], "count": values["count"], "amount": values["amount"]}
                for stage, values in stage_stats.items()
            ],
            "contracts": contracts[:8],
        }

    async def list_suppliers(self, *, search: str | None = None, active_only: bool = False) -> list[dict]:
        suppliers = await self.repository.list_suppliers(search=search, active_only=active_only)
        return [self._serialize_supplier(supplier) for supplier in suppliers]

    async def get_supplier(self, supplier_id: UUID) -> dict:
        supplier = await self.repository.get_supplier(supplier_id)
        if not supplier:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Fornecedor não encontrado")
        return self._serialize_supplier(supplier)

    async def create_supplier(self, data: PaymentSupplierCreate, current_user: User) -> dict:
        supplier = PaymentSupplier(
            name=self._normalize_business_name(data.name),
            cnpj=self._digits(data.cnpj) or None,
            active=data.active,
            notes=data.notes,
        )
        try:
            await self.repository.create_supplier(supplier)
            await self.audit.record(
                actor=current_user,
                action="CREATE",
                entity_type="PAYMENT_SUPPLIER",
                entity_id=supplier.id,
                entity_label=supplier.name,
                details={"cnpj": supplier.cnpj},
            )
            await self.db.commit()
        except IntegrityError as exc:
            await self.db.rollback()
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Fornecedor já cadastrado") from exc
        return self._serialize_supplier(supplier)

    async def update_supplier(self, supplier_id: UUID, data: PaymentSupplierUpdate, current_user: User) -> dict:
        supplier = await self.repository.get_supplier(supplier_id)
        if not supplier:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Fornecedor não encontrado")
        payload = data.model_dump(exclude_unset=True)
        if "name" in payload and payload["name"] is not None:
            supplier.name = self._normalize_business_name(payload["name"])
        if "cnpj" in payload:
            supplier.cnpj = self._digits(payload["cnpj"]) or None
        for field in ("active", "notes"):
            if field in payload:
                setattr(supplier, field, payload[field])
        supplier.updated_at = datetime.now(timezone.utc)
        try:
            await self.audit.record(actor=current_user, action="UPDATE", entity_type="PAYMENT_SUPPLIER", entity_id=supplier.id, entity_label=supplier.name, details={})
            await self.db.commit()
        except IntegrityError as exc:
            await self.db.rollback()
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Não foi possível atualizar o fornecedor") from exc
        return self._serialize_supplier(supplier)

    async def delete_supplier(self, supplier_id: UUID, current_user: User) -> dict:
        supplier = await self.repository.get_supplier(supplier_id)
        if not supplier:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Fornecedor não encontrado")
        supplier.active = False
        supplier.updated_at = datetime.now(timezone.utc)
        await self.audit.record(actor=current_user, action="DISABLE", entity_type="PAYMENT_SUPPLIER", entity_id=supplier.id, entity_label=supplier.name, details={})
        await self.db.commit()
        return {"message": "Fornecedor inativado"}

    async def list_contracts(
        self,
        *,
        current_user: User | None = None,
        supplier_id: UUID | None = None,
        status_filter: PaymentContractStatus | None = None,
        kind: PaymentProcessKind | None = None,
        search: str | None = None,
    ) -> list[dict]:
        contracts = await self.repository.list_contracts(supplier_id=supplier_id, status_filter=status_filter, kind=kind, search=search)
        return [await self._serialize_contract(contract) for contract in contracts]

    async def get_contract(self, contract_id: UUID) -> dict:
        contract = await self.repository.get_contract(contract_id)
        if not contract:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Contrato não encontrado")
        return await self._serialize_contract(contract)

    async def create_contract(self, data: PaymentContractCreate, current_user: User) -> dict:
        supplier = await self.repository.get_supplier(data.supplier_id)
        if not supplier:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Fornecedor não encontrado")
        contract = PaymentContract(**data.model_dump())
        contract.number = self._normalize_identifier(contract.number)
        if contract.value_updated is None and contract.value_initial is not None:
            contract.value_updated = contract.value_initial
        try:
            await self.repository.create_contract(contract)
            await self.audit.record(
                actor=current_user,
                action="CREATE",
                entity_type="PAYMENT_CONTRACT",
                entity_id=contract.id,
                entity_label=contract.number,
                details={"supplier_id": str(contract.supplier_id)},
            )
            await self.db.commit()
        except IntegrityError as exc:
            await self.db.rollback()
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Contrato já cadastrado para este fornecedor") from exc
        return await self.get_contract(contract.id)

    async def update_contract(self, contract_id: UUID, data: PaymentContractUpdate, current_user: User) -> dict:
        contract = await self.repository.get_contract(contract_id)
        if not contract:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Contrato não encontrado")
        payload = data.model_dump(exclude_unset=True)
        if "supplier_id" in payload and payload["supplier_id"]:
            supplier = await self.repository.get_supplier(payload["supplier_id"])
            if not supplier:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Fornecedor não encontrado")
        for field, value in payload.items():
            if field == "number" and value is not None:
                value = self._normalize_identifier(value)
            setattr(contract, field, value)
        if contract.value_updated is None and contract.value_initial is not None:
            contract.value_updated = contract.value_initial
        contract.updated_at = datetime.now(timezone.utc)
        try:
            await self.audit.record(actor=current_user, action="UPDATE", entity_type="PAYMENT_CONTRACT", entity_id=contract.id, entity_label=contract.number, details={})
            await self.db.commit()
        except IntegrityError as exc:
            await self.db.rollback()
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Não foi possível atualizar o contrato") from exc
        return await self.get_contract(contract.id)

    async def delete_contract(self, contract_id: UUID, current_user: User) -> dict:
        contract = await self.repository.get_contract(contract_id)
        if not contract:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Contrato não encontrado")
        contract.status = PaymentContractStatus.CANCELLED
        contract.updated_at = datetime.now(timezone.utc)
        await self.audit.record(actor=current_user, action="CANCEL", entity_type="PAYMENT_CONTRACT", entity_id=contract.id, entity_label=contract.number, details={})
        await self.db.commit()
        return {"message": "Contrato cancelado"}

    async def create_contract_amendment(self, contract_id: UUID, data: PaymentContractAmendmentCreate, current_user: User) -> dict:
        contract = await self.repository.get_contract(contract_id)
        if not contract:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Contrato não encontrado")
        amendment = PaymentContractAmendment(contract_id=contract.id, **data.model_dump())
        if amendment.value_delta is not None:
            base = contract.value_updated if contract.value_updated is not None else contract.value_initial or Decimal("0")
            contract.value_updated = base + amendment.value_delta
        if amendment.valid_until:
            contract.valid_until = amendment.valid_until
        contract.updated_at = datetime.now(timezone.utc)
        try:
            await self.repository.create_contract_amendment(amendment)
            await self.audit.record(
                actor=current_user,
                action="AMEND",
                entity_type="PAYMENT_CONTRACT",
                entity_id=contract.id,
                entity_label=contract.number,
                details={"amendment_id": str(amendment.id), "value_delta": str(amendment.value_delta)},
            )
            await self.db.commit()
        except IntegrityError as exc:
            await self.db.rollback()
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Não foi possível adicionar o aditivo") from exc
        return await self.get_contract(contract.id)

    async def contract_management(
        self,
        contract_id: UUID,
        *,
        horizon_months: int = 6,
    ) -> dict:
        contract = await self.repository.get_contract(contract_id)
        if not contract:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Contrato não encontrado")
        horizon = self._normalize_horizon(horizon_months)
        return await self._build_contract_management(contract, horizon_months=horizon, include_related=True)

    async def contract_management_summary(
        self,
        *,
        current_user: User,
        horizon_months: int = 6,
        supplier_id: UUID | None = None,
        status_filter: PaymentContractStatus | None = None,
        kind: PaymentProcessKind | None = None,
        search: str | None = None,
    ) -> dict:
        horizon = self._normalize_horizon(horizon_months)
        contracts = await self.repository.list_contracts(supplier_id=supplier_id, status_filter=status_filter, kind=kind, search=search)
        items = []
        total_available = Decimal("0")
        total_consumed = Decimal("0")
        total_average = Decimal("0")

        for contract in contracts:
            serialized = await self._serialize_contract(contract)
            processes = await self.repository.list_contract_processes(contract.id)
            average = self._average_monthly_from_processes(processes, horizon)
            total_consumed += Decimal(str(serialized["consumed_amount"] or 0))
            total_average += average
            if serialized["available_balance"] is not None:
                total_available += Decimal(str(serialized["available_balance"]))
            projected_date, _projected_label = self._project_depletion_date(serialized["available_balance"], average, horizon)
            items.append(
                {
                    "contract_id": contract.id,
                    "contract_number": contract.number,
                    "supplier_name": serialized.get("supplier_name"),
                    "kind": contract.kind,
                    "effective_value": serialized.get("effective_value"),
                    "consumed_amount": serialized["consumed_amount"],
                    "paid_amount": serialized["paid_amount"],
                    "pending_amount": serialized["pending_amount"],
                    "available_balance": serialized["available_balance"],
                    "average_monthly_consumption": average,
                    "projected_depletion_date": projected_date,
                    "alerts_count": len(serialized.get("alerts") or []),
                    "status": contract.status,
                }
            )

        def ranking_key(item: dict):
            balance = item["available_balance"]
            balance_value = Decimal(str(balance)) if balance is not None else Decimal("999999999")
            projected = item["projected_depletion_date"] or date.max
            return (-item["alerts_count"], projected, balance_value)

        ranking = sorted(items, key=ranking_key)[:12]
        return {
            "total_contracts": len(contracts),
            "active_contracts": len([contract for contract in contracts if contract.status == PaymentContractStatus.ACTIVE]),
            "critical_contracts": len([item for item in items if item["alerts_count"] > 0 or (item["available_balance"] is not None and item["available_balance"] < 0)]),
            "total_available_balance": total_available,
            "total_consumed_amount": total_consumed,
            "average_monthly_consumption": total_average,
            "ranking": ranking,
        }

    async def import_xlsx(self, upload: UploadFile, current_user: User) -> dict:
        filename = upload.filename or "processos-pagamento.xlsx"
        suffix = Path(filename).suffix.lower()
        if suffix != ".xlsx":
            await upload.close()
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Envie uma planilha XLSX")

        if load_workbook is None:
            await upload.close()
            raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Leitura XLSX indisponível")

        content = await upload.read()
        await upload.close()
        if not content:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Arquivo vazio")

        parsed_rows = self._parse_workbook(content, filename)
        if not parsed_rows:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Nenhuma linha de processo encontrada")
        if len(parsed_rows) > MAX_IMPORT_ROWS:
            raise HTTPException(
                status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                detail=f"Importação limitada a {MAX_IMPORT_ROWS} registros por arquivo",
            )

        organizations = await self._organization_lookup()
        seen_keys: set[str] = set()
        created = 0
        updated = 0
        skipped = 0
        errors = 0
        row_results = []
        now = datetime.now(timezone.utc)
        imported_contract_ids: set[UUID | None] = set()

        try:
            for parsed in parsed_rows:
                data = dict(parsed["data"])
                data["organization_id"] = self._resolve_organization_id(data.get("unit_name"), organizations)
                data = self._apply_user_scope_to_row(data, current_user)
                process_number = data.get("process_number")
                row_info = {
                    "row_number": parsed["row_number"],
                    "sheet": parsed["sheet"],
                    "process_number": process_number,
                }

                if parsed.get("validation_error"):
                    errors += 1
                    row_results.append({**row_info, "action": "ERROR", "detail": parsed["validation_error"]})
                    continue
                if data.get("scope_error"):
                    errors += 1
                    row_results.append({**row_info, "action": "ERROR", "detail": data["scope_error"]})
                    continue

                data = await self._prepare_process_payload(data, current_user=current_user, from_import=True)
                if data.get("contract_id"):
                    imported_contract_ids.add(data.get("contract_id"))
                import_key = self._build_import_key(data)
                if import_key in seen_keys:
                    skipped += 1
                    row_results.append({**row_info, "action": "SKIPPED", "detail": "Linha duplicada no arquivo"})
                    continue
                seen_keys.add(import_key)

                existing = await self.repository.get_by_import_key(import_key)
                if existing:
                    self._fill_model(existing, data, current_user=current_user, now=now, filename=filename, sheet=parsed["sheet"], import_key=import_key)
                    updated += 1
                    action = "UPDATE"
                    record = existing
                else:
                    record = PaymentProcess(import_key=import_key)
                    self._fill_model(record, data, current_user=current_user, now=now, filename=filename, sheet=parsed["sheet"], import_key=import_key)
                    await self.repository.create(record)
                    created += 1
                    action = "CREATE"

                self._ensure_default_checklist(record)
                await self.audit.record(
                    actor=current_user,
                    action=action,
                    entity_type="PAYMENT_PROCESS",
                    entity_id=record.id,
                    entity_label=record.process_number,
                    details={
                        "source_filename": filename,
                        "source_sheet": parsed["sheet"],
                        "row_number": parsed["row_number"],
                        "amount": record.amount,
                        "status": record.status,
                    },
                )
                row_results.append({**row_info, "action": action, "detail": None})

            await self.repository.flush()
            await self._normalize_legacy_contract_values(imported_contract_ids)
            await self.db.commit()
        except IntegrityError as exc:
            await self.db.rollback()
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Não foi possível importar a planilha") from exc

        return {
            "total_rows": len(parsed_rows),
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": errors,
            "rows": row_results,
        }

    async def export_xlsx(
        self,
        *,
        current_user: User,
        kind: PaymentProcessKind | None = None,
        stage: PaymentProcessStage | None = None,
        status_filter: str | None = None,
        organization_id: UUID | None = None,
        supplier_id: UUID | None = None,
        contract_id: UUID | None = None,
        competence_month: date | None = None,
        due_from: date | None = None,
        due_to: date | None = None,
        search: str | None = None,
    ) -> tuple[str, bytes]:
        if production_scope_is_empty(current_user):
            records = []
        else:
            records = await self.repository.list_for_export(
                kind=kind,
                stage=stage,
                status_filter=status_filter,
                organization_id=scoped_organization_id(current_user, organization_id),
                supplier_id=supplier_id,
                contract_id=contract_id,
                competence_month=self._month_start(competence_month),
                due_from=due_from,
                due_to=due_to,
                search=search,
            )

        rows = [self._serialize(item) for item in records]
        filename = f"processos-pagamento-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}.xlsx"
        return filename, self._build_workbook_bytes(
            rows=[
                [
                    self._kind_label(row["kind"]),
                    row["process_number"],
                    row["system"],
                    row["status"],
                    row["stage_label"],
                    row["billing_number"],
                    row["invoice_number"],
                    row["invoice_type"],
                    row["unit_name"],
                    row["issue_date"].isoformat() if isinstance(row["issue_date"], date) else row["issue_date"],
                    row["competence_month"].isoformat() if isinstance(row["competence_month"], date) else row["competence_month"],
                    row["due_date"].isoformat() if isinstance(row["due_date"], date) else row["due_date"],
                    row["process_type"],
                    row["amount"],
                    row["supplier_name"],
                    "",
                    row["contract_number"],
                    row.get("contract_type"),
                    row.get("contract_object"),
                    row.get("contract_valid_from"),
                    row.get("contract_valid_until"),
                    row.get("contract_value_initial"),
                    row.get("contract_value_updated"),
                    row["contract_balance"],
                    row["commitment_number"],
                    row["commitment_date"].isoformat() if isinstance(row["commitment_date"], date) else row["commitment_date"],
                    row["liquidation_number"],
                    row["liquidation_date"].isoformat() if isinstance(row["liquidation_date"], date) else row["liquidation_date"],
                    row["payment_order_number"],
                    row["payment_order_date"].isoformat() if isinstance(row["payment_order_date"], date) else row["payment_order_date"],
                    row["paid_at"].isoformat() if isinstance(row["paid_at"], date) else row["paid_at"],
                    row["stage_owner"],
                    row["location"],
                    "; ".join(row["alerts"]) if row.get("alerts") else row["notes"],
                ]
                for row in rows
            ],
            headers=TEMPLATE_HEADERS,
            sheet_name="Processos",
        )

    def template_xlsx(self) -> tuple[str, bytes]:
        return "modelo-processos-pagamento.xlsx", self._build_workbook_bytes(
            headers=TEMPLATE_HEADERS,
            rows=[
                [
                    "Combustível",
                    "PMTF-PR-00000/2026",
                    "IGOV",
                    "EM ANALISE",
                    "Conferência",
                    "2721579",
                    "45154",
                    "Consumo combustivel",
                    "SECRETARIA MUNICIPAL DE ADMINISTRACAO",
                    date.today(),
                    date.today().replace(day=1),
                    date.today(),
                    "Composta Combustível",
                    0,
                    "POSTO L J",
                    "",
                    "2-914-2025",
                    "Fornecimento",
                    "Fornecimento de combustíveis",
                    date.today().replace(month=1, day=1),
                    date.today().replace(month=12, day=31),
                    0,
                    0,
                    0,
                    "",
                    None,
                    "",
                    None,
                    "",
                    None,
                    None,
                    "Setor de Frotas",
                    "Setor de Frotas",
                    "Linha exemplo. Substitua pelos dados reais.",
                ]
            ],
            sheet_name="Modelo",
        )

    def _parse_workbook(self, content: bytes, filename: str) -> list[dict]:
        try:
            workbook = load_workbook(io.BytesIO(content), data_only=True)
        except (InvalidFileException, OSError, ValueError) as exc:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Planilha XLSX inválida") from exc

        parsed_rows: list[dict] = []
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            header_info = self._detect_header(rows)
            if not header_info:
                continue

            header_index, headers, column_map = header_info
            metadata = self._extract_sheet_metadata(rows, sheet.title)
            for row_offset, values in enumerate(rows[header_index + 1 :], start=header_index + 2):
                data = self._extract_process_row(values, column_map, metadata, filename, sheet.title)
                if not self._row_has_process_content(data):
                    continue

                validation_error = None
                if not data.get("process_number"):
                    validation_error = "Linha sem número do processo"
                parsed_rows.append(
                    {
                        "sheet": sheet.title,
                        "row_number": row_offset,
                        "headers": headers,
                        "data": data,
                        "validation_error": validation_error,
                    }
                )

        return parsed_rows

    def _detect_header(self, rows: list[tuple]) -> tuple[int, list[str], dict[str, int]] | None:
        best = None
        for index, row in enumerate(rows[:40]):
            headers = [self._clean_header(value) for value in row]
            column_map = self._build_column_map(headers)
            score = len({"process_number", "status", "unit_name", "issue_date", "amount"} & set(column_map))
            if score >= 4 and (best is None or score > best[0]):
                best = (score, index, headers, column_map)
        if not best:
            return None
        return best[1], best[2], best[3]

    def _build_column_map(self, headers: list[str]) -> dict[str, int]:
        normalized = [self._norm_key(header) for header in headers]
        column_map: dict[str, int] = {}
        tipo_indexes: list[int] = []

        for index, key in enumerate(normalized):
            if not key:
                continue
            if key in {"tipo", "tipo processo"}:
                if key == "tipo processo":
                    column_map["kind"] = index
                else:
                    tipo_indexes.append(index)
                continue
            field = self._header_field(key)
            if field and field not in column_map:
                column_map[field] = index

        unit_index = column_map.get("unit_name")
        issue_index = column_map.get("issue_date")
        for index in tipo_indexes:
            if "invoice_type" not in column_map and (unit_index is None or index < unit_index):
                column_map["invoice_type"] = index
            elif "process_type" not in column_map and (issue_index is None or index > issue_index):
                column_map["process_type"] = index
            elif "invoice_type" not in column_map:
                column_map["invoice_type"] = index
            elif "process_type" not in column_map:
                column_map["process_type"] = index

        return column_map

    def _header_field(self, key: str) -> str | None:
        aliases = {
            "process_number": {"n do processo", "no do processo", "numero do processo", "processo", "n processo"},
            "system": {"sistema"},
            "status": {"status", "situacao", "situacao do processo"},
            "stage": {"etapa", "fase", "andamento etapa"},
            "billing_number": {"fatura"},
            "invoice_number": {"nf", "nota fiscal", "numero nf", "n nf"},
            "invoice_type": {"tipo nota", "tipo da nota"},
            "unit_name": {"unidade", "secretaria", "orgao", "lotacao"},
            "issue_date": {"emissao nf", "emissao da nf", "data emissao nf", "data emissao"},
            "competence_month": {"competencia", "mes competencia", "competencia mes"},
            "due_date": {"vencimento", "data vencimento", "data de vencimento"},
            "amount": {"valor", "valor nf", "valor total"},
            "supplier_name": {"fornecedor", "prestador", "empresa"},
            "supplier_cnpj": {"cnpj fornecedor", "cnpj prestador", "cnpj empresa"},
            "contract_number": {"contrato", "numero contrato", "contrato numero"},
            "contract_type": {"tipo contrato", "tipo do contrato"},
            "contract_object": {"objeto contrato", "objeto do contrato"},
            "contract_valid_from": {"inicio vigencia", "vigencia inicio", "data inicio contrato"},
            "contract_valid_until": {"fim vigencia", "vigencia fim", "data fim contrato"},
            "contract_value_initial": {"valor inicial contrato", "valor contratado", "valor inicial"},
            "contract_value_updated": {"valor atualizado contrato", "valor atual contrato", "valor atualizado"},
            "contract_balance": {"saldo restante", "saldo", "saldo do contrato"},
            "commitment_number": {"empenho", "numero empenho", "n empenho"},
            "commitment_date": {"data empenho"},
            "liquidation_number": {"liquidacao", "numero liquidacao", "n liquidacao"},
            "liquidation_date": {"data liquidacao"},
            "payment_order_number": {"ordem pagamento", "ordem de pagamento", "op", "numero op"},
            "payment_order_date": {"data ordem pagamento", "data op"},
            "paid_at": {"data pagamento", "pagamento em", "pago em"},
            "stage_owner": {"responsavel etapa", "setor responsavel", "responsavel"},
            "location": {"localizacao", "local", "andamento", "setor"},
            "notes": {"observacoes", "observacao", "obs", "nota"},
        }
        for field, values in aliases.items():
            if key in values:
                return field
        return None

    def _extract_sheet_metadata(self, rows: list[tuple], sheet_name: str) -> dict:
        text_cells = []
        balance = None
        for row in rows[:12]:
            row_values = list(row or [])
            for index, value in enumerate(row_values):
                text = self._string_value(value)
                if text:
                    text_cells.append(text)
                    if "SALDO RESTANTE" in self._strip_accents(text).upper():
                        for candidate in row_values[index + 1 : index + 4]:
                            balance = self._decimal_value(candidate)
                            if balance is not None:
                                break

        joined = " ".join(text_cells)
        contract = self._extract_contract(joined)
        supplier = self._extract_supplier(joined)
        if not supplier and "PRIME" in self._strip_accents(sheet_name).upper():
            supplier = "PRIME CONSULTORIA E ASSESSORIA EMPRESARIAL LTDA"
        if not supplier and "L J" in self._strip_accents(sheet_name).upper():
            supplier = "L J POSTO DE COMBUSTIVEIS LTDA"
        kind = self._kind_from_value(joined)
        return {
            "supplier_name": supplier,
            "contract_number": contract,
            "contract_balance": balance,
            "kind": kind,
        }

    def _extract_process_row(self, values: tuple, column_map: dict[str, int], metadata: dict, filename: str, sheet_name: str) -> dict:
        def cell(field: str):
            index = column_map.get(field)
            if index is None or index >= len(values):
                return None
            return values[index]

        data = {
            "process_number": self._string_value(cell("process_number"), uppercase=True),
            "kind": self._kind_from_value(cell("kind")) or metadata.get("kind") or PaymentProcessKind.FUEL,
            "system": self._string_value(cell("system"), uppercase=True),
            "status": self._string_value(cell("status"), uppercase=True),
            "stage": self._stage_from_value(cell("stage")) or self._stage_from_status(cell("status")),
            "billing_number": self._string_value(cell("billing_number")),
            "invoice_number": self._string_value(cell("invoice_number")),
            "invoice_type": self._string_value(cell("invoice_type")),
            "unit_name": self._string_value(cell("unit_name")),
            "issue_date": self._date_value(cell("issue_date")),
            "competence_month": self._month_start(self._date_value(cell("competence_month"))),
            "due_date": self._date_value(cell("due_date")),
            "process_type": self._string_value(cell("process_type")),
            "amount": self._decimal_value(cell("amount")),
            "supplier_name": self._string_value(cell("supplier_name")) or metadata.get("supplier_name"),
            "supplier_cnpj": self._string_value(cell("supplier_cnpj")),
            "contract_number": self._string_value(cell("contract_number")) or metadata.get("contract_number"),
            "contract_type": self._string_value(cell("contract_type")),
            "contract_object": self._string_value(cell("contract_object")),
            "contract_valid_from": self._date_value(cell("contract_valid_from")),
            "contract_valid_until": self._date_value(cell("contract_valid_until")),
            "contract_value_initial": self._decimal_value(cell("contract_value_initial")),
            "contract_value_updated": self._decimal_value(cell("contract_value_updated")),
            "contract_balance": self._decimal_value(cell("contract_balance")) or metadata.get("contract_balance"),
            "commitment_number": self._string_value(cell("commitment_number")),
            "commitment_date": self._date_value(cell("commitment_date")),
            "liquidation_number": self._string_value(cell("liquidation_number")),
            "liquidation_date": self._date_value(cell("liquidation_date")),
            "payment_order_number": self._string_value(cell("payment_order_number")),
            "payment_order_date": self._date_value(cell("payment_order_date")),
            "paid_at": self._date_value(cell("paid_at")),
            "stage_owner": self._string_value(cell("stage_owner")),
            "location": self._string_value(cell("location")),
            "notes": self._string_value(cell("notes")),
            "source_filename": filename,
            "source_sheet": sheet_name,
        }
        return data

    async def _organization_lookup(self) -> dict[str, UUID]:
        organizations = list((await self.db.execute(select(Organization))).scalars().all())
        return {self._norm_lookup(organization.name): organization.id for organization in organizations if self._norm_lookup(organization.name)}

    def _resolve_organization_id(self, unit_name: str | None, lookup: dict[str, UUID]) -> UUID | None:
        key = self._norm_lookup(unit_name)
        if not key:
            return None
        if key in lookup:
            return lookup[key]
        for organization_key, organization_id in lookup.items():
            if organization_key and (organization_key in key or key in organization_key):
                return organization_id
        return None

    def _apply_user_scope_to_row(self, data: dict, current_user: User) -> dict:
        scoped_org_id = scoped_organization_id(current_user)
        if scoped_org_id is None:
            if production_scope_is_empty(current_user):
                data["scope_error"] = "Usuário de produção sem órgão vinculado"
            return data

        row_org_id = data.get("organization_id")
        if is_production_user(current_user) and row_org_id is None:
            data["scope_error"] = "Linha sem órgão reconhecido para usuário de produção"
            return data
        if row_org_id and row_org_id != scoped_org_id:
            data["scope_error"] = "Linha pertence a outro órgão"
            return data
        data["organization_id"] = scoped_org_id
        return data

    def _fill_model(
        self,
        record: PaymentProcess,
        data: dict,
        *,
        current_user: User,
        now: datetime,
        filename: str,
        sheet: str,
        import_key: str,
    ) -> None:
        record.import_key = import_key
        record.process_number = data["process_number"]
        record.kind = data["kind"]
        record.system = data.get("system")
        record.status = data.get("status")
        record.stage = data.get("stage") or self._stage_from_status(data.get("status"))
        record.billing_number = data.get("billing_number")
        record.invoice_number = data.get("invoice_number")
        record.invoice_type = data.get("invoice_type")
        record.unit_name = data.get("unit_name")
        record.organization_id = data.get("organization_id")
        record.supplier_id = data.get("supplier_id")
        record.contract_id = data.get("contract_id")
        record.issue_date = data.get("issue_date")
        record.competence_month = data.get("competence_month") or self._month_start(data.get("issue_date"))
        record.due_date = data.get("due_date")
        record.paid_at = data.get("paid_at")
        record.process_type = data.get("process_type")
        record.amount = data.get("amount")
        record.supplier_name = data.get("supplier_name")
        record.contract_number = data.get("contract_number")
        record.contract_balance = data.get("contract_balance")
        record.location = data.get("location")
        record.notes = data.get("notes")
        record.assigned_to_user_id = data.get("assigned_to_user_id")
        record.stage_owner = data.get("stage_owner")
        record.status_note = data.get("status_note")
        record.commitment_number = data.get("commitment_number")
        record.commitment_date = data.get("commitment_date")
        record.liquidation_number = data.get("liquidation_number")
        record.liquidation_date = data.get("liquidation_date")
        record.payment_order_number = data.get("payment_order_number")
        record.payment_order_date = data.get("payment_order_date")
        record.source_filename = filename
        record.source_sheet = sheet
        record.updated_by_user_id = current_user.id
        record.updated_at = now
        if not record.created_by_user_id:
            record.created_by_user_id = current_user.id

    def _build_import_key(self, data: dict) -> str:
        parts = [
            data.get("process_number"),
            data.get("supplier_name"),
            data.get("contract_number"),
            data.get("billing_number"),
            data.get("invoice_number"),
            data.get("unit_name"),
        ]
        key = "|".join(self._norm_lookup(value) or "_" for value in parts)
        return key[:320]

    def _row_has_process_content(self, data: dict) -> bool:
        keys = ("process_number", "system", "status", "billing_number", "invoice_number", "unit_name", "amount")
        return any(self._present(data.get(key)) for key in keys)

    def _serialize(self, item: PaymentProcess) -> dict:
        return {
            "id": item.id,
            "import_key": item.import_key,
            "process_number": item.process_number,
            "kind": item.kind,
            "system": item.system,
            "status": item.status,
            "stage": item.stage,
            "stage_label": STAGE_LABELS.get(item.stage, str(item.stage)),
            "billing_number": item.billing_number,
            "invoice_number": item.invoice_number,
            "invoice_type": item.invoice_type,
            "unit_name": item.unit_name,
            "organization_id": item.organization_id,
            "organization_name": item.organization.name if item.organization else None,
            "supplier_id": item.supplier_id,
            "supplier_name": item.supplier.name if item.supplier else item.supplier_name,
            "contract_id": item.contract_id,
            "contract_number": item.contract.number if item.contract else item.contract_number,
            "contract_supplier_name": item.contract.supplier.name if item.contract and item.contract.supplier else None,
            "issue_date": item.issue_date,
            "competence_month": item.competence_month,
            "due_date": item.due_date,
            "paid_at": item.paid_at,
            "process_type": item.process_type,
            "amount": float(item.amount) if item.amount is not None else None,
            "contract_balance": float(item.contract_balance) if item.contract_balance is not None else None,
            "contract_type": item.contract.contract_type if item.contract else None,
            "contract_object": item.contract.object_description if item.contract else None,
            "contract_valid_from": item.contract.valid_from if item.contract else None,
            "contract_valid_until": item.contract.valid_until if item.contract else None,
            "contract_value_initial": float(item.contract.value_initial) if item.contract and item.contract.value_initial is not None else None,
            "contract_value_updated": float(item.contract.value_updated) if item.contract and item.contract.value_updated is not None else None,
            "location": item.location,
            "notes": item.notes,
            "assigned_to_user_id": item.assigned_to_user_id,
            "assigned_to_name": item.assigned_to.name if item.assigned_to else None,
            "stage_owner": item.stage_owner,
            "status_note": item.status_note,
            "commitment_number": item.commitment_number,
            "commitment_date": item.commitment_date,
            "liquidation_number": item.liquidation_number,
            "liquidation_date": item.liquidation_date,
            "payment_order_number": item.payment_order_number,
            "payment_order_date": item.payment_order_date,
            "source_filename": item.source_filename,
            "source_sheet": item.source_sheet,
            "alerts": self._process_alerts(item),
            "checklist": [self._serialize_checklist_item(checklist_item) for checklist_item in sorted(item.checklist_items, key=lambda row: (self._stage_rank(row.stage), row.item_label))],
            "references": [self._serialize_reference(reference) for reference in sorted(item.references, key=lambda row: row.created_at or datetime.min.replace(tzinfo=timezone.utc))],
            "stage_events": [self._serialize_stage_event(event) for event in sorted(item.stage_events, key=lambda row: row.created_at or datetime.min.replace(tzinfo=timezone.utc), reverse=True)],
            "created_at": item.created_at,
            "updated_at": item.updated_at,
        }

    def _delete_audit_details(self, item: PaymentProcess, *, reason: str) -> dict:
        return {
            "reason": reason,
            "process_number": item.process_number,
            "kind": item.kind,
            "stage": item.stage,
            "stage_label": STAGE_LABELS.get(item.stage, str(item.stage)),
            "status": item.status,
            "supplier_id": str(item.supplier_id) if item.supplier_id else None,
            "supplier_name": item.supplier.name if item.supplier else item.supplier_name,
            "contract_id": str(item.contract_id) if item.contract_id else None,
            "contract_number": item.contract.number if item.contract else item.contract_number,
            "organization_id": str(item.organization_id) if item.organization_id else None,
            "organization_name": item.organization.name if item.organization else None,
            "invoice_number": item.invoice_number,
            "billing_number": item.billing_number,
            "competence_month": item.competence_month.isoformat() if item.competence_month else None,
            "due_date": item.due_date.isoformat() if item.due_date else None,
            "amount": str(item.amount) if item.amount is not None else None,
            "source_filename": item.source_filename,
            "source_sheet": item.source_sheet,
            "references_count": len(item.references),
            "checklist_count": len(item.checklist_items),
            "stage_events_count": len(item.stage_events),
        }

    async def _normalize_legacy_contract_values(self, contract_ids: set[UUID | None]) -> None:
        for contract_id in {item for item in contract_ids if item}:
            contract = await self.repository.get_contract(contract_id)
            if not contract:
                continue
            if contract.value_initial is not None or contract.value_updated is not None or contract.imported_balance is None:
                continue
            totals = await self.repository.contract_process_totals(contract.id)
            contract.value_updated = Decimal(str(contract.imported_balance)) + totals["consumed_amount"]
            contract.updated_at = datetime.now(timezone.utc)

    async def _prepare_process_payload(
        self,
        payload: dict,
        *,
        current_user: User,
        existing: PaymentProcess | None = None,
        from_import: bool = False,
    ) -> dict:
        prepared = dict(payload)
        if prepared.get("process_number"):
            prepared["process_number"] = self._normalize_identifier(prepared["process_number"])
        if prepared.get("system"):
            prepared["system"] = self._normalize_identifier(prepared["system"])
        if prepared.get("status"):
            prepared["status"] = self._normalize_identifier(prepared["status"])

        if "stage" not in prepared or prepared.get("stage") is None:
            prepared["stage"] = self._stage_from_status(prepared.get("status") or (existing.status if existing else None))
        elif not isinstance(prepared["stage"], PaymentProcessStage):
            prepared["stage"] = self._stage_from_value(prepared["stage"]) or PaymentProcessStage.ASSEMBLY

        if prepared.get("competence_month"):
            prepared["competence_month"] = self._month_start(prepared["competence_month"])
        elif prepared.get("issue_date"):
            prepared["competence_month"] = self._month_start(prepared["issue_date"])

        supplier = None
        if prepared.get("supplier_id"):
            supplier = await self.repository.get_supplier(prepared["supplier_id"])
            if not supplier:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Fornecedor não encontrado")
            prepared["supplier_name"] = supplier.name
        elif prepared.get("supplier_name"):
            supplier = await self._get_or_create_supplier(
                prepared.get("supplier_name"),
                cnpj=prepared.get("supplier_cnpj"),
                current_user=current_user,
            )
            prepared["supplier_id"] = supplier.id
            prepared["supplier_name"] = supplier.name

        contract = None
        if prepared.get("contract_id"):
            contract = await self.repository.get_contract(prepared["contract_id"])
            if not contract:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Contrato não encontrado")
            prepared["contract_number"] = contract.number
            prepared["supplier_id"] = contract.supplier_id
            prepared["supplier_name"] = contract.supplier.name if contract.supplier else prepared.get("supplier_name")
        elif prepared.get("contract_number") and supplier:
            contract = await self._get_or_create_contract(
                supplier=supplier,
                number=prepared.get("contract_number"),
                kind=prepared.get("kind") or (existing.kind if existing else PaymentProcessKind.FUEL),
                contract_type=prepared.get("contract_type"),
                object_description=prepared.get("contract_object"),
                valid_from=prepared.get("contract_valid_from"),
                valid_until=prepared.get("contract_valid_until"),
                value_initial=prepared.get("contract_value_initial"),
                value_updated=prepared.get("contract_value_updated"),
                imported_balance=prepared.get("contract_balance"),
                current_user=current_user,
                from_import=from_import,
            )
            prepared["contract_id"] = contract.id
            prepared["contract_number"] = contract.number

        for transient in (
            "supplier_cnpj",
            "contract_type",
            "contract_object",
            "contract_valid_from",
            "contract_valid_until",
            "contract_value_initial",
            "contract_value_updated",
        ):
            prepared.pop(transient, None)
        return prepared

    def _assign_process_fields(self, record: PaymentProcess, payload: dict) -> None:
        fields = (
            "process_number",
            "kind",
            "system",
            "status",
            "stage",
            "billing_number",
            "invoice_number",
            "invoice_type",
            "unit_name",
            "organization_id",
            "supplier_id",
            "contract_id",
            "issue_date",
            "competence_month",
            "due_date",
            "paid_at",
            "process_type",
            "amount",
            "supplier_name",
            "contract_number",
            "contract_balance",
            "location",
            "notes",
            "assigned_to_user_id",
            "stage_owner",
            "status_note",
            "commitment_number",
            "commitment_date",
            "liquidation_number",
            "liquidation_date",
            "payment_order_number",
            "payment_order_date",
        )
        for field in fields:
            if field in payload:
                setattr(record, field, payload[field])

    async def _get_or_create_supplier(self, name: str | None, *, cnpj: str | None, current_user: User) -> PaymentSupplier:
        normalized_name = self._normalize_business_name(name)
        if not normalized_name:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Fornecedor inválido")
        digits = self._digits(cnpj)
        supplier = await self.repository.get_supplier_by_cnpj(digits) if digits else None
        if supplier is None:
            supplier = await self.repository.get_supplier_by_name(normalized_name)
        if supplier:
            changed = False
            if digits and not supplier.cnpj:
                supplier.cnpj = digits
                changed = True
            if not supplier.active:
                supplier.active = True
                changed = True
            if changed:
                supplier.updated_at = datetime.now(timezone.utc)
            return supplier

        supplier = PaymentSupplier(name=normalized_name, cnpj=digits or None, active=True)
        await self.repository.create_supplier(supplier)
        await self.audit.record(
            actor=current_user,
            action="CREATE",
            entity_type="PAYMENT_SUPPLIER",
            entity_id=supplier.id,
            entity_label=supplier.name,
            details={"auto_created": True},
        )
        return supplier

    async def _get_or_create_contract(
        self,
        *,
        supplier: PaymentSupplier,
        number: str | None,
        kind: PaymentProcessKind | str | None,
        contract_type: str | None,
        object_description: str | None,
        valid_from: date | None,
        valid_until: date | None,
        value_initial: Decimal | None,
        value_updated: Decimal | None,
        imported_balance: Decimal | None,
        current_user: User,
        from_import: bool,
    ) -> PaymentContract:
        normalized_number = self._normalize_identifier(number)
        if not normalized_number:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Contrato inválido")
        contract = await self.repository.get_contract_by_supplier_number(supplier.id, normalized_number)
        normalized_kind = kind if isinstance(kind, PaymentProcessKind) else self._kind_from_value(kind) or PaymentProcessKind.FUEL
        if contract:
            self._fill_contract_missing_data(
                contract,
                kind=normalized_kind,
                contract_type=contract_type,
                object_description=object_description,
                valid_from=valid_from,
                valid_until=valid_until,
                value_initial=value_initial,
                value_updated=value_updated,
                imported_balance=imported_balance,
                allow_overwrite=not from_import,
            )
            return contract

        contract = PaymentContract(
            supplier_id=supplier.id,
            number=normalized_number,
            kind=normalized_kind,
            contract_type=contract_type,
            object_description=object_description,
            valid_from=valid_from,
            valid_until=valid_until,
            value_initial=value_initial,
            value_updated=value_updated if value_updated is not None else value_initial,
            imported_balance=imported_balance,
            status=PaymentContractStatus.ACTIVE,
        )
        await self.repository.create_contract(contract)
        await self.audit.record(
            actor=current_user,
            action="CREATE",
            entity_type="PAYMENT_CONTRACT",
            entity_id=contract.id,
            entity_label=contract.number,
            details={"auto_created": True, "supplier_id": str(supplier.id)},
        )
        return contract

    def _fill_contract_missing_data(
        self,
        contract: PaymentContract,
        *,
        kind: PaymentProcessKind | None,
        contract_type: str | None,
        object_description: str | None,
        valid_from: date | None,
        valid_until: date | None,
        value_initial: Decimal | None,
        value_updated: Decimal | None,
        imported_balance: Decimal | None,
        allow_overwrite: bool = False,
    ) -> None:
        updates = {
            "kind": kind,
            "contract_type": contract_type,
            "object_description": object_description,
            "valid_from": valid_from,
            "valid_until": valid_until,
            "value_initial": value_initial,
            "value_updated": value_updated,
            "imported_balance": imported_balance,
        }
        changed = False
        for field, value in updates.items():
            if value is None:
                continue
            if allow_overwrite or getattr(contract, field) is None:
                setattr(contract, field, value)
                changed = True
        if changed:
            contract.updated_at = datetime.now(timezone.utc)

    def _replace_references(self, record: PaymentProcess, references: list[dict]) -> None:
        record.references.clear()
        for item in references or []:
            payload = item.model_dump() if hasattr(item, "model_dump") else dict(item)
            if not payload.get("label"):
                continue
            record.references.append(
                PaymentProcessReference(
                    reference_type=payload.get("reference_type") or PaymentProcessReferenceType.OTHER,
                    external_id=payload.get("external_id"),
                    label=payload["label"],
                    description=payload.get("description"),
                )
            )

    def _ensure_default_checklist(self, record: PaymentProcess) -> None:
        existing = {(item.stage, self._norm_lookup(item.item_label)) for item in record.checklist_items}
        now = datetime.now(timezone.utc)
        for stage, labels in CHECKLIST_DEFAULTS.items():
            for label in labels:
                key = (stage, self._norm_lookup(label))
                if key not in existing:
                    record.checklist_items.append(
                        PaymentProcessChecklistItem(
                            stage=stage,
                            item_label=label,
                            status=PaymentChecklistStatus.PENDING,
                            updated_at=now,
                            created_at=now,
                        )
                    )
                    existing.add(key)

    def _process_alerts(self, item: PaymentProcess) -> list[str]:
        alerts: list[str] = []
        if not item.supplier_id and not item.supplier_name:
            alerts.append("Fornecedor não vinculado.")
        if not item.contract_id and not item.contract_number:
            alerts.append("Contrato não vinculado.")
        if not item.competence_month:
            alerts.append("Competência não informada.")
        if not item.billing_number and not item.invoice_number:
            alerts.append("Fatura ou NF não informada.")
        if item.amount is None:
            alerts.append("Valor do processo não informado.")
        if not item.due_date and item.stage not in {PaymentProcessStage.PAID, PaymentProcessStage.ARCHIVED, PaymentProcessStage.CANCELLED}:
            alerts.append("Vencimento não informado.")

        current_pending = [
            checklist_item.item_label
            for checklist_item in item.checklist_items
            if checklist_item.stage == item.stage and checklist_item.status == PaymentChecklistStatus.PENDING
        ]
        if current_pending:
            alerts.append(f"Checklist pendente na etapa {STAGE_LABELS.get(item.stage, item.stage.value)}: {', '.join(current_pending[:3])}.")

        rank = self._stage_rank(item.stage)
        if rank >= self._stage_rank(PaymentProcessStage.COMMITMENT) and not item.commitment_number:
            alerts.append("Empenho não informado.")
        if rank >= self._stage_rank(PaymentProcessStage.LIQUIDATION) and not item.liquidation_number:
            alerts.append("Liquidação não informada.")
        if rank >= self._stage_rank(PaymentProcessStage.PAYMENT) and not item.payment_order_number:
            alerts.append("Ordem de pagamento não informada.")
        if item.stage in {PaymentProcessStage.PAID, PaymentProcessStage.ARCHIVED} and not item.paid_at:
            alerts.append("Data de pagamento não informada.")

        contract = item.contract
        if contract:
            if contract.status != PaymentContractStatus.ACTIVE:
                alerts.append("Contrato não está ativo.")
            if contract.valid_until and contract.valid_until < date.today():
                alerts.append("Contrato vencido.")
            balance = self._contract_balance_from_loaded(contract)
            if balance is not None and balance < Decimal("0"):
                alerts.append("Processos vinculados excedem o saldo do contrato.")
        return alerts

    def _serialize_supplier(self, supplier: PaymentSupplier) -> dict:
        return {
            "id": supplier.id,
            "name": supplier.name,
            "cnpj": supplier.cnpj,
            "active": supplier.active,
            "notes": supplier.notes,
            "created_at": supplier.created_at,
            "updated_at": supplier.updated_at,
        }

    async def _serialize_contract(self, contract: PaymentContract) -> dict:
        totals = await self.repository.contract_process_totals(contract.id)
        effective_value = self._contract_effective_value(contract, totals=totals)
        available_balance = effective_value - totals["consumed_amount"] if effective_value is not None else None
        alerts = []
        if contract.status != PaymentContractStatus.ACTIVE:
            alerts.append("Contrato não está ativo.")
        if contract.valid_until and contract.valid_until < date.today():
            alerts.append("Contrato vencido.")
        if available_balance is not None and available_balance < Decimal("0"):
            alerts.append("Saldo contratual excedido.")
        return {
            "id": contract.id,
            "supplier_id": contract.supplier_id,
            "supplier_name": contract.supplier.name if contract.supplier else None,
            "number": contract.number,
            "kind": contract.kind,
            "contract_type": contract.contract_type,
            "object_description": contract.object_description,
            "valid_from": contract.valid_from,
            "valid_until": contract.valid_until,
            "value_initial": contract.value_initial,
            "value_updated": contract.value_updated,
            "imported_balance": contract.imported_balance,
            "effective_value": effective_value,
            "consumed_amount": totals["consumed_amount"],
            "paid_amount": totals["paid_amount"],
            "pending_amount": totals["pending_amount"],
            "available_balance": available_balance,
            "status": contract.status,
            "notes": contract.notes,
            "alerts": alerts,
            "amendments": [
                {
                    "id": amendment.id,
                    "contract_id": amendment.contract_id,
                    "amendment_type": amendment.amendment_type,
                    "number": amendment.number,
                    "signed_at": amendment.signed_at,
                    "value_delta": amendment.value_delta,
                    "valid_until": amendment.valid_until,
                    "notes": amendment.notes,
                    "created_at": amendment.created_at,
                    "updated_at": amendment.updated_at,
                }
                for amendment in sorted(contract.amendments, key=lambda item: item.signed_at or date.min)
            ],
            "created_at": contract.created_at,
            "updated_at": contract.updated_at,
        }

    def _serialize_reference(self, reference: PaymentProcessReference) -> dict:
        return {
            "id": reference.id,
            "process_id": reference.process_id,
            "reference_type": reference.reference_type,
            "external_id": reference.external_id,
            "label": reference.label,
            "description": reference.description,
            "created_at": reference.created_at,
        }

    def _serialize_checklist_item(self, item: PaymentProcessChecklistItem) -> dict:
        return {
            "id": item.id,
            "process_id": item.process_id,
            "stage": item.stage,
            "item_label": item.item_label,
            "status": item.status,
            "notes": item.notes,
            "updated_by_user_id": item.updated_by_user_id,
            "updated_by_name": item.updater.name if item.updater else None,
            "updated_at": item.updated_at,
            "created_at": item.created_at,
        }

    def _serialize_stage_event(self, event: PaymentProcessStageEvent) -> dict:
        alerts = []
        if event.alerts_snapshot:
            try:
                parsed = json.loads(event.alerts_snapshot)
                if isinstance(parsed, list):
                    alerts = [str(item) for item in parsed]
            except json.JSONDecodeError:
                alerts = [event.alerts_snapshot]
        return {
            "id": event.id,
            "process_id": event.process_id,
            "from_stage": event.from_stage,
            "to_stage": event.to_stage,
            "comment": event.comment,
            "alerts": alerts,
            "created_by_user_id": event.created_by_user_id,
            "created_by_name": event.creator.name if event.creator else None,
            "created_at": event.created_at,
        }

    def _contract_effective_value(self, contract: PaymentContract, *, totals: dict[str, Decimal] | None = None) -> Decimal | None:
        if contract.value_updated is not None:
            return Decimal(str(contract.value_updated))
        if contract.value_initial is not None:
            base = Decimal(str(contract.value_initial))
            delta = sum((Decimal(str(item.value_delta or 0)) for item in contract.amendments), Decimal("0"))
            return base + delta
        return None

    def _contract_balance_from_loaded(self, contract: PaymentContract) -> Decimal | None:
        effective_value = self._contract_effective_value(contract)
        if effective_value is None:
            return None
        return effective_value - self._contract_loaded_consumed(contract)

    def _contract_loaded_consumed(self, contract: PaymentContract) -> Decimal:
        return sum(
            (
                Decimal(str(process.amount or 0))
                for process in contract.processes
                if process.stage not in {PaymentProcessStage.CANCELLED, PaymentProcessStage.RETURNED}
            ),
            Decimal("0"),
        )

    async def _build_contract_management(self, contract: PaymentContract, *, horizon_months: int, include_related: bool) -> dict:
        contract_data = await self._serialize_contract(contract)
        processes = await self.repository.list_contract_processes(contract.id)
        organization_ids = await self.repository.contract_organization_ids(contract.id)
        history_months = max(6, horizon_months)
        today = date.today()
        current_month = today.replace(day=1)
        start_month = self._add_months(current_month, -(history_months - 1))
        start_at = datetime.combine(start_month, time.min, tzinfo=timezone.utc)
        history = {month: self._empty_management_month(month) for month in self._month_range(start_month, history_months)}

        for process in processes:
            month = self._process_month(process)
            if month not in history:
                continue
            amount = Decimal(str(process.amount or 0))
            if process.stage not in TERMINAL_STAGES:
                history[month]["process_amount"] += amount
                history[month]["total_amount"] += amount
                history[month]["records_count"] += 1
            if process.stage in {PaymentProcessStage.PAID, PaymentProcessStage.ARCHIVED}:
                history[month]["paid_amount"] += amount
            elif process.stage not in TERMINAL_STAGES:
                history[month]["pending_amount"] += amount

        operations = []
        source_quality = "Previsao baseada nos processos financeiros vinculados ao contrato."
        contract_kind = contract.kind or self._infer_contract_kind(processes)
        if contract_kind == PaymentProcessKind.FUEL:
            fuel_rows = await self.repository.list_fuel_operations_for_supplier(
                supplier_cnpj=contract.supplier.cnpj if contract.supplier else None,
                supplier_name=contract.supplier.name if contract.supplier else None,
                organization_ids=organization_ids or None,
                start_at=start_at,
            )
            for row in fuel_rows:
                month = row.supplied_at.date().replace(day=1)
                if month not in history:
                    continue
                amount = Decimal(str(row.total_amount or 0))
                liters = Decimal(str(row.liters or 0))
                history[month]["operational_amount"] += amount
                history[month]["liters"] += liters
                history[month]["records_count"] += 1
                operations.append(row)
            if fuel_rows:
                source_quality = (
                    "Combustível cruzado por CNPJ do fornecedor e posto."
                    if contract.supplier and contract.supplier.cnpj
                    else "Combustível cruzado por nome normalizado do fornecedor/posto."
                )
        elif contract_kind == PaymentProcessKind.MAINTENANCE:
            maintenance_rows = await self.repository.list_maintenance_operations_for_organizations(
                organization_ids=organization_ids or None,
                start_at=start_at,
            )
            for row in maintenance_rows:
                month = row.start_date.date().replace(day=1)
                if month not in history:
                    continue
                amount = Decimal(str(row.total_cost or 0))
                history[month]["maintenance_amount"] += amount
                history[month]["records_count"] += 1
                operations.append(row)
            if maintenance_rows:
                source_quality = "Manutenção simplificada por histórico operacional dos órgãos relacionados."

        monthly_history = [self._serialize_management_month(item) for item in history.values()]
        forecast_values = self._forecast_values(monthly_history, contract_kind)
        average = self._average_decimal(forecast_values)
        variation = self._monthly_variation(forecast_values)
        depletion_date, depletion_label = self._project_depletion_date(contract_data["available_balance"], average, horizon_months)
        monthly_projection = self._build_projection(current_month, horizon_months, contract_data["available_balance"], average)
        alerts = list(contract_data.get("alerts") or [])
        if not forecast_values:
            alerts.append("Histórico insuficiente para previsão mensal consistente.")
        if contract_data["available_balance"] is not None and contract_data["available_balance"] < Decimal("0"):
            alerts.append("Saldo contratual ja esta negativo.")

        return {
            "contract": contract_data,
            "horizon_months": horizon_months,
            "generated_at": datetime.now(timezone.utc),
            "source_quality": source_quality,
            "average_monthly_consumption": average,
            "monthly_variation_percentage": variation,
            "projected_depletion_date": depletion_date,
            "projected_depletion_label": depletion_label,
            "kpis": self._contract_management_kpis(contract_data, average, variation, depletion_label, alerts),
            "monthly_history": monthly_history,
            "monthly_projection": monthly_projection,
            "related_processes": self._related_process_items(processes) if include_related else [],
            "related_operations": self._related_operation_items(operations, contract_kind) if include_related else [],
            "alerts": alerts,
        }

    def _contract_management_kpis(
        self,
        contract_data: dict,
        average: Decimal,
        variation: Decimal | None,
        depletion_label: str,
        alerts: list[str],
    ) -> list[dict]:
        return [
            {
                "key": "available_balance",
                "label": "Saldo",
                "value": contract_data["available_balance"],
                "formatted": self._format_money(contract_data["available_balance"]),
                "tone": "danger" if contract_data["available_balance"] is not None and contract_data["available_balance"] < 0 else "success",
                "formula": "Valor efetivo do contrato - processos vinculados não devolvidos/cancelados.",
                "source": "Processos de pagamento vinculados ao contrato.",
                "detail_type": "processes",
            },
            {
                "key": "consumed_amount",
                "label": "Consumido",
                "value": contract_data["consumed_amount"],
                "formatted": self._format_money(contract_data["consumed_amount"]),
                "tone": "neutral",
                "formula": "Soma de processos com contrato e valor, exceto Devolvido e Cancelado.",
                "source": "Processos de pagamento.",
                "detail_type": "processes",
            },
            {
                "key": "paid_amount",
                "label": "Pago",
                "value": contract_data["paid_amount"],
                "formatted": self._format_money(contract_data["paid_amount"]),
                "tone": "success",
                "formula": "Soma dos processos nas etapas Pago e Arquivado.",
                "source": "Workflow financeiro.",
                "detail_type": "processes",
            },
            {
                "key": "pending_amount",
                "label": "Pendente",
                "value": contract_data["pending_amount"],
                "formatted": self._format_money(contract_data["pending_amount"]),
                "tone": "warning",
                "formula": "Consumido - Pago.",
                "source": "Processos em aberto ou ainda não pagos.",
                "detail_type": "processes",
            },
            {
                "key": "monthly_average",
                "label": "Media mensal",
                "value": average,
                "formatted": self._format_money(average),
                "tone": "neutral",
                "formula": "Media dos meses com consumo no horizonte selecionado.",
                "source": "Processos financeiros e histórico operacional quando disponível.",
                "detail_type": "operations",
            },
            {
                "key": "depletion_forecast",
                "label": "Fim do saldo",
                "value": depletion_label,
                "formatted": depletion_label,
                "tone": "danger" if depletion_label not in {"Sem previsão", "Sem saldo definido"} else "neutral",
                "formula": "Saldo disponível / média mensal de consumo.",
                "source": "Saldo calculado e média mensal.",
                "detail_type": "projection",
            },
            {
                "key": "monthly_variation",
                "label": "Variação mensal",
                "value": variation,
                "formatted": self._format_percentage(variation),
                "tone": "warning" if variation is not None and variation > 0 else "neutral",
                "formula": "Comparação entre os dois últimos meses do horizonte.",
                "source": "Série histórica mensal.",
                "detail_type": "projection",
            },
            {
                "key": "alerts",
                "label": "Alertas",
                "value": len(alerts),
                "formatted": str(len(alerts)),
                "tone": "danger" if alerts else "success",
                "formula": "Alertas de saldo, vigência, status e qualidade da previsão.",
                "source": "Regras do módulo de pagamentos.",
                "detail_type": "alerts",
            },
        ]

    def _average_monthly_from_processes(self, processes: list[PaymentProcess], horizon_months: int) -> Decimal:
        today = date.today().replace(day=1)
        start_month = self._add_months(today, -(max(3, horizon_months) - 1))
        buckets: defaultdict[date, Decimal] = defaultdict(lambda: Decimal("0"))
        for process in processes:
            if process.stage in TERMINAL_STAGES:
                continue
            month = self._process_month(process)
            if month >= start_month:
                buckets[month] += Decimal(str(process.amount or 0))
        return self._average_decimal([value for value in buckets.values() if value > 0])

    def _forecast_values(self, monthly_history: list[dict], contract_kind: PaymentProcessKind | None) -> list[Decimal]:
        values: list[Decimal] = []
        for month in monthly_history:
            process_amount = Decimal(str(month.get("process_amount") or 0))
            operational_amount = Decimal(str(month.get("operational_amount") or 0))
            maintenance_amount = Decimal(str(month.get("maintenance_amount") or 0))
            if contract_kind == PaymentProcessKind.FUEL and operational_amount > 0:
                values.append(operational_amount)
            elif contract_kind == PaymentProcessKind.MAINTENANCE and process_amount > 0:
                values.append(process_amount)
            elif contract_kind == PaymentProcessKind.MAINTENANCE and maintenance_amount > 0:
                values.append(maintenance_amount)
            elif process_amount > 0:
                values.append(process_amount)
        return values

    def _build_projection(self, current_month: date, horizon_months: int, available_balance, average: Decimal) -> list[dict]:
        balance = Decimal(str(available_balance)) if available_balance is not None else None
        rows = []
        for offset in range(1, horizon_months + 1):
            month = self._add_months(current_month, offset)
            if balance is not None:
                balance -= average
            item = self._empty_management_month(month)
            item["projected_amount"] = average
            item["projected_balance"] = balance
            rows.append(self._serialize_management_month(item))
        return rows

    def _project_depletion_date(self, available_balance, average: Decimal, horizon_months: int) -> tuple[date | None, str]:
        if available_balance is None:
            return None, "Sem saldo definido"
        balance = Decimal(str(available_balance))
        if balance < 0:
            return date.today(), "Saldo negativo"
        if average <= 0:
            return None, "Sem previsão"
        months_until = int((balance / average).to_integral_value(rounding=ROUND_FLOOR)) + 1
        projected = self._add_months(date.today().replace(day=1), months_until)
        label = self._month_label(projected)
        if months_until > horizon_months:
            label = f"Apos {horizon_months} meses"
        return projected, label

    def _monthly_variation(self, values: list[Decimal]) -> Decimal | None:
        if len(values) < 2:
            return None
        previous = values[-2]
        current = values[-1]
        if previous <= 0:
            return None
        return ((current - previous) / previous * Decimal("100")).quantize(Decimal("0.01"))

    def _average_decimal(self, values: list[Decimal]) -> Decimal:
        values = [Decimal(str(value)) for value in values if Decimal(str(value or 0)) > 0]
        if not values:
            return Decimal("0")
        return (sum(values, Decimal("0")) / Decimal(len(values))).quantize(Decimal("0.01"))

    def _related_process_items(self, processes: list[PaymentProcess]) -> list[dict]:
        rows = sorted(processes, key=lambda item: item.updated_at or item.created_at, reverse=True)[:60]
        return [
            {
                "id": item.id,
                "kind": "process",
                "label": item.process_number,
                "date": item.competence_month or item.issue_date or item.due_date,
                "amount": item.amount,
                "status": STAGE_LABELS.get(item.stage, item.stage.value),
                "detail": item.supplier.name if item.supplier else item.supplier_name,
            }
            for item in rows
        ]

    def _related_operation_items(self, operations: list, contract_kind: PaymentProcessKind | None) -> list[dict]:
        rows = sorted(operations, key=lambda item: getattr(item, "supplied_at", None) or getattr(item, "start_date", None) or datetime.min.replace(tzinfo=timezone.utc), reverse=True)[:60]
        result = []
        for item in rows:
            if contract_kind == PaymentProcessKind.FUEL:
                result.append(
                    {
                        "id": item.id,
                        "kind": "fuel_supply",
                        "label": item.vehicle.plate if item.vehicle else item.fuel_station or "Abastecimento",
                        "date": item.supplied_at,
                        "amount": Decimal(str(item.total_amount or 0)),
                        "status": item.fuel_type,
                        "detail": f"{item.liters or 0} L - {item.fuel_station_ref.name if item.fuel_station_ref else item.fuel_station or '-'}",
                    }
                )
            else:
                result.append(
                    {
                        "id": item.id,
                        "kind": "maintenance",
                        "label": item.vehicle.plate if item.vehicle else "Manutenção",
                        "date": item.start_date,
                        "amount": item.total_cost,
                        "status": "Aberta" if not item.end_date else "Concluída",
                        "detail": item.service_description,
                    }
                )
        return result

    def _empty_management_month(self, month: date) -> dict:
        return {
            "month": month,
            "label": self._month_label(month),
            "process_amount": Decimal("0"),
            "operational_amount": Decimal("0"),
            "maintenance_amount": Decimal("0"),
            "total_amount": Decimal("0"),
            "paid_amount": Decimal("0"),
            "pending_amount": Decimal("0"),
            "liters": Decimal("0"),
            "records_count": 0,
            "projected_amount": None,
            "projected_balance": None,
        }

    def _serialize_management_month(self, item: dict) -> dict:
        return {
            "month": item["month"],
            "label": item["label"],
            "process_amount": item["process_amount"],
            "operational_amount": item["operational_amount"],
            "maintenance_amount": item["maintenance_amount"],
            "total_amount": item["total_amount"] or item["operational_amount"] or item["maintenance_amount"],
            "paid_amount": item["paid_amount"],
            "pending_amount": item["pending_amount"],
            "liters": item["liters"],
            "records_count": item["records_count"],
            "projected_amount": item["projected_amount"],
            "projected_balance": item["projected_balance"],
        }

    def _month_range(self, start_month: date, count: int) -> list[date]:
        return [self._add_months(start_month, offset) for offset in range(count)]

    def _add_months(self, value: date, months: int) -> date:
        month_index = value.month - 1 + months
        year = value.year + month_index // 12
        month = month_index % 12 + 1
        return date(year, month, 1)

    def _process_month(self, process: PaymentProcess) -> date:
        value = process.competence_month or process.issue_date or process.due_date
        if isinstance(value, datetime):
            value = value.date()
        if isinstance(value, date):
            return value.replace(day=1)
        created = process.created_at.date() if isinstance(process.created_at, datetime) else date.today()
        return created.replace(day=1)

    def _month_label(self, value: date) -> str:
        return f"{value.month:02d}/{value.year}"

    def _normalize_horizon(self, value: int) -> int:
        if value <= 3:
            return 3
        if value >= 12:
            return 12
        return 6

    def _infer_contract_kind(self, processes: list[PaymentProcess]) -> PaymentProcessKind | None:
        counts: defaultdict[PaymentProcessKind, int] = defaultdict(int)
        for process in processes:
            counts[process.kind] += 1
        if not counts:
            return None
        return max(counts.items(), key=lambda item: item[1])[0]

    def _format_money(self, value) -> str:
        if value is None:
            return "-"
        number = Decimal(str(value)).quantize(Decimal("0.01"))
        text = f"{number:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")
        return f"R$ {text}"

    def _format_percentage(self, value) -> str:
        if value is None:
            return "-"
        number = Decimal(str(value)).quantize(Decimal("0.01"))
        return f"{number}%"

    def _empty_dashboard(self) -> dict:
        return {
            "total_processes": 0,
            "open_processes": 0,
            "overdue_processes": 0,
            "due_soon_processes": 0,
            "total_amount": Decimal("0"),
            "paid_amount": Decimal("0"),
            "pending_amount": Decimal("0"),
            "alerts_count": 0,
            "stages": [{"stage": stage, "label": STAGE_LABELS[stage], "count": 0, "amount": Decimal("0")} for stage in PaymentProcessStage],
            "contracts": [],
        }

    def _ensure_process_visible(self, record: PaymentProcess, current_user: User) -> None:
        organization_id = scoped_organization_id(current_user)
        if organization_id is None:
            if production_scope_is_empty(current_user):
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Processo de pagamento não encontrado")
            return
        if record.organization_id != organization_id:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Processo de pagamento não encontrado")

    def _manual_import_key(self, data: dict) -> str:
        stable = self._build_import_key(data)
        return f"manual|{stable}|{uuid4()}"[:320]

    def _stage_rank(self, stage: PaymentProcessStage | str | None) -> int:
        normalized = stage if isinstance(stage, PaymentProcessStage) else self._stage_from_value(stage)
        if normalized in STAGE_ORDER:
            return STAGE_ORDER.index(normalized)
        return -1

    def _month_start(self, value) -> date | None:
        if value is None:
            return None
        if isinstance(value, datetime):
            value = value.date()
        if isinstance(value, date):
            return date(value.year, value.month, 1)
        parsed = self._date_value(value)
        return self._month_start(parsed)

    def _stage_from_status(self, value) -> PaymentProcessStage:
        text = self._strip_accents(str(value or "")).upper()
        if not text:
            return PaymentProcessStage.ASSEMBLY
        if "CANCEL" in text or "REPROV" in text:
            return PaymentProcessStage.CANCELLED
        if "DEVOL" in text:
            return PaymentProcessStage.RETURNED
        if "PAGO" in text or "CONCL" in text:
            return PaymentProcessStage.PAID
        if "LIQUID" in text:
            return PaymentProcessStage.LIQUIDATION
        if "EMPENH" in text:
            return PaymentProcessStage.COMMITMENT
        if "PAGAMENTO" in text or "FINAN" in text or "TESOUR" in text:
            return PaymentProcessStage.PAYMENT
        if "ANAL" in text or "CONFER" in text:
            return PaymentProcessStage.REVIEW
        return PaymentProcessStage.ASSEMBLY

    def _stage_from_value(self, value) -> PaymentProcessStage | None:
        text = self._strip_accents(str(value or "")).upper().strip()
        if not text:
            return None
        for stage in PaymentProcessStage:
            if text == stage.value:
                return stage
        label_map = {self._strip_accents(label).upper(): stage for stage, label in STAGE_LABELS.items()}
        return label_map.get(text) or self._stage_from_status(text)

    def _normalize_business_name(self, value: str | None) -> str:
        return self._normalize_spaces(str(value or "")).upper()

    def _normalize_identifier(self, value: str | None) -> str:
        return self._normalize_spaces(str(value or "")).upper()

    def _digits(self, value: str | None) -> str:
        return re.sub(r"\D+", "", str(value or ""))

    def _build_workbook_bytes(self, *, headers: list[str], rows: list[list], sheet_name: str) -> bytes:
        if Workbook is None:
            raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Geração XLSX indisponível")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        sheet.freeze_panes = "A2"
        for column_cells in sheet.columns:
            max_length = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=10)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 42)
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _extract_contract(self, text: str) -> str | None:
        match = re.search(r"\b\d+-\d+-\d{4}\b", text or "")
        return match.group(0) if match else None

    def _extract_supplier(self, text: str) -> str | None:
        normalized = self._strip_accents(text or "")
        match = re.search(r"Pagamento em favor de\s+(.+?)\s+\d+-\d+-\d{4}", normalized, flags=re.IGNORECASE)
        if match:
            return self._normalize_spaces(match.group(1)).upper()
        match = re.search(r"SALDO\s*-\s*(.+?)\s*-\s*CONTRATO", normalized, flags=re.IGNORECASE)
        if match:
            return self._normalize_spaces(match.group(1)).upper()
        return None

    def _kind_from_value(self, value) -> PaymentProcessKind | None:
        text = self._strip_accents(str(value or "")).upper()
        if "MANUT" in text:
            return PaymentProcessKind.MAINTENANCE
        if "COMBUST" in text or "FUEL" in text:
            return PaymentProcessKind.FUEL
        return None

    def _clean_header(self, value) -> str:
        return self._normalize_spaces(str(value or ""))

    def _string_value(self, value, *, uppercase: bool = False) -> str | None:
        if value is None:
            return None
        if isinstance(value, datetime):
            text = value.date().isoformat()
        elif isinstance(value, date):
            text = value.isoformat()
        elif isinstance(value, float) and value.is_integer():
            text = str(int(value))
        else:
            text = str(value)
        text = self._normalize_spaces(text)
        if not text or text.upper() in MISSING_TEXTS or not text.strip("\u2014").strip():
            return None
        return text.upper() if uppercase else text

    def _date_value(self, value) -> date | None:
        parsed = self._datetime_value(value)
        return parsed.date() if parsed else None

    def _datetime_value(self, value) -> datetime | None:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.replace(tzinfo=value.tzinfo or timezone.utc)
        if isinstance(value, date):
            return datetime.combine(value, time.min, tzinfo=timezone.utc)
        text = self._string_value(value)
        if not text:
            return None
        if re.fullmatch(r"\d+(\.\d+)?", text):
            try:
                return datetime(1899, 12, 30, tzinfo=timezone.utc) + timedelta(days=float(text))
            except ValueError:
                return None
        for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y", "%Y-%m-%d", "%Y-%m-%dT%H:%M:%S"):
            try:
                return datetime.strptime(text, fmt).replace(tzinfo=timezone.utc)
            except ValueError:
                continue
        return None

    def _decimal_value(self, value) -> Decimal | None:
        if value is None:
            return None
        if isinstance(value, Decimal):
            return value.quantize(Decimal("0.01"))
        if isinstance(value, (int, float)):
            return Decimal(str(value)).quantize(Decimal("0.01"))
        text = self._string_value(value)
        if not text:
            return None
        normalized = text.replace("R$", "").replace(" ", "")
        if "," in normalized and "." in normalized:
            normalized = normalized.replace(".", "").replace(",", ".")
        else:
            normalized = normalized.replace(",", ".")
        normalized = re.sub(r"[^0-9.\-]", "", normalized)
        if not normalized:
            return None
        try:
            return Decimal(normalized).quantize(Decimal("0.01"))
        except (InvalidOperation, ValueError):
            return None

    def _present(self, value) -> bool:
        return self._string_value(value) is not None if not isinstance(value, Decimal) else True

    def _normalize_spaces(self, value: str) -> str:
        return re.sub(r"\s+", " ", value).strip()

    def _strip_accents(self, value: str) -> str:
        return "".join(ch for ch in unicodedata.normalize("NFD", value) if unicodedata.category(ch) != "Mn")

    def _norm_key(self, value) -> str:
        text = self._strip_accents(str(value or "")).lower()
        text = re.sub(r"[^a-z0-9]+", " ", text)
        return self._normalize_spaces(text)

    def _norm_lookup(self, value) -> str:
        return self._norm_key(value).upper()

    def _kind_label(self, kind: PaymentProcessKind | str | None) -> str:
        value = kind.value if isinstance(kind, PaymentProcessKind) else str(kind or "")
        return "Manutenção" if value == PaymentProcessKind.MAINTENANCE.value else "Combustível"
