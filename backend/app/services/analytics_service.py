from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from html import escape
from io import BytesIO
from typing import Any
from uuid import UUID

from fastapi import HTTPException, status
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.drawing.image import Image as WorksheetImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image, LongTable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import case, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.official_identity import (
    ADMINISTRATION_SECRETARIAT,
    COLOR_BORDER,
    COLOR_MUTED,
    COLOR_NAVY,
    COLOR_SURFACE,
    FLEET_DEPARTMENT,
    MUNICIPALITY_ADDRESS,
    MUNICIPALITY_CNPJ,
    MUNICIPALITY_NAME,
    crest_path,
    ensure_pdf_fonts,
    institutional_datetime,
)
from app.models.claim import Claim
from app.models.driver import Driver
from app.models.fine import Fine
from app.models.fleet_analytics_snapshot import FleetAnalyticsSnapshot
from app.models.fuel_supply import FuelSupply
from app.models.location_history import LocationHistory
from app.models.maintenance import MaintenanceRecord
from app.models.master_data import Allocation, Department
from app.models.vehicle import Vehicle, VehicleStatus
from app.repositories.analytics_repository import AnalyticsRepository


ANALYTICS_EXPORT_NO_CACHE_HEADERS = {
    "Cache-Control": "private, no-store, no-cache, max-age=0, must-revalidate",
    "Pragma": "no-cache",
    "Expires": "0",
    "X-Content-Type-Options": "nosniff",
}

ANALYTICS_EXPORT_COLUMNS = (
    ("metric", "metric", 28),
    ("severity", "severity", 16),
    ("current_value", "current_value", 18),
    ("category_average", "category_average", 22),
    ("variance_percentage", "variance_percentage", 22),
    ("recommended_action", "recommended_action", 54),
)

ANALYTICS_METRIC_LABELS = {
    "consumption_l_100km": "Consumo de combustível (L/100 km)",
    "tco_cost_per_km": "Custo total por quilômetro",
    "driver_risk_score": "Índice de risco do condutor",
}

ANALYTICS_SEVERITY_LABELS = {
    "CRITICAL": "Crítica",
    "HIGH": "Alta",
    "MEDIUM": "Média",
    "LOW": "Baixa",
}

MARKET_TCO_BENCHMARK_BY_TYPE = {
    "SEDAN": 1.35,
    "HATCH": 1.20,
    "PICAPE": 1.75,
    "SUV": 1.90,
    "PERUA_SW": 1.55,
    "VAN": 2.30,
    "MICRO_ONIBUS": 2.80,
    "ONIBUS": 3.20,
    "CAMINHAO": 3.80,
    "MOTOCICLETA": 0.60,
    "MAQUINA": 5.50,
}


def neutralize_analytics_spreadsheet_text(value: str) -> str:
    """Prevent spreadsheet formula execution, including whitespace-prefixed payloads."""
    probe = value.lstrip(" \t\r\n")
    if probe.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _analytics_xlsx_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (int, float)):
        return value
    return neutralize_analytics_spreadsheet_text(str(value))


def _analytics_pdf_text(value: Any, *, fallback: str = "-") -> str:
    return escape(str(fallback if value is None or value == "" else value))


def _institutional_timestamp(value: datetime) -> str:
    local_value = institutional_datetime(value)
    offset = local_value.strftime("%z")
    formatted_offset = f"{offset[:3]}:{offset[3:]}" if offset else ""
    suffix = f" (UTC{formatted_offset})" if formatted_offset else ""
    return f"{local_value.strftime('%d/%m/%Y às %H:%M:%S')}{suffix}"


def calculate_consumption_l_100km(total_liters: float, total_km: float) -> float | None:
    if total_km <= 0:
        return None
    return (total_liters / total_km) * 100


def calculate_tco_per_km(fuel_cost: float, maintenance_cost: float, fines_cost: float, total_km: float) -> float | None:
    if total_km <= 0:
        return None
    return (fuel_cost + maintenance_cost + fines_cost) / total_km


def calculate_driver_risk_score(*, fines_count: int, claims_count: int, anomalies_count: int) -> float:
    return (fines_count * 0.3) + (claims_count * 0.5) + (anomalies_count * 0.2)


def calculate_variance_percentage(current_value: float | None, baseline: float | None) -> float | None:
    if current_value is None or baseline is None or baseline == 0:
        return None
    return ((current_value - baseline) / baseline) * 100


@dataclass
class _VehicleAggregate:
    vehicle_id: UUID
    vehicle_type: str
    total_km: float = 0.0
    total_liters: float = 0.0
    fuel_cost: float = 0.0
    maintenance_cost: float = 0.0
    fines_cost: float = 0.0
    anomalies_count: int = 0


class AnalyticsService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.analytics_repo = AnalyticsRepository(db)

    def _period_bounds(self, period_days: int) -> tuple[datetime, datetime]:
        reference = datetime.now(timezone.utc)
        end = reference.replace(hour=0, minute=0, second=0, microsecond=0)
        start = end - timedelta(days=period_days)
        return start, end

    @staticmethod
    def _unique_vehicle_snapshots(rows: list[FleetAnalyticsSnapshot]) -> list[FleetAnalyticsSnapshot]:
        unique: dict[UUID, FleetAnalyticsSnapshot] = {}
        for row in rows:
            if row.scope != "VEHICLE" or row.vehicle_id is None:
                continue
            unique.setdefault(row.vehicle_id, row)
        return list(unique.values())

    @staticmethod
    def _unique_driver_snapshots(rows: list[FleetAnalyticsSnapshot]) -> list[FleetAnalyticsSnapshot]:
        unique: dict[UUID, FleetAnalyticsSnapshot] = {}
        for row in rows:
            if row.scope != "DRIVER" or row.driver_id is None:
                continue
            unique.setdefault(row.driver_id, row)
        return list(unique.values())

    async def _vehicle_ids_for_organization(self, organization_id: UUID | None) -> set[UUID] | None:
        if not organization_id:
            return None

        result = await self.db.execute(
            select(LocationHistory.vehicle_id)
            .join(Allocation, Allocation.id == LocationHistory.allocation_id)
            .join(Department, Department.id == Allocation.department_id)
            .where(
                LocationHistory.end_date.is_(None),
                Department.organization_id == organization_id,
            )
        )
        return set(result.scalars().all())

    async def _driver_ids_for_organization(self, organization_id: UUID | None) -> set[UUID] | None:
        if not organization_id:
            return None

        result = await self.db.execute(select(Driver.id).where(Driver.organization_id == organization_id))
        return set(result.scalars().all())

    async def _filter_vehicle_snapshots_by_organization(
        self,
        rows: list[FleetAnalyticsSnapshot],
        organization_id: UUID | None,
    ) -> list[FleetAnalyticsSnapshot]:
        vehicle_ids = await self._vehicle_ids_for_organization(organization_id)
        if vehicle_ids is None:
            return rows
        return [row for row in rows if row.vehicle_id in vehicle_ids]

    async def _filter_driver_snapshots_by_organization(
        self,
        rows: list[FleetAnalyticsSnapshot],
        organization_id: UUID | None,
    ) -> list[FleetAnalyticsSnapshot]:
        driver_ids = await self._driver_ids_for_organization(organization_id)
        if driver_ids is None:
            return rows
        return [row for row in rows if row.driver_id in driver_ids]

    async def _ensure_snapshots(self, period_days: int) -> list[FleetAnalyticsSnapshot]:
        period_start, period_end = self._period_bounds(period_days)

        vehicle_rows = (
            await self.db.execute(
                select(Vehicle.id, Vehicle.vehicle_type).where(Vehicle.status != VehicleStatus.INATIVO)
            )
        ).all()
        vehicle_map = {row.id: str(row.vehicle_type.value if hasattr(row.vehicle_type, "value") else row.vehicle_type) for row in vehicle_rows}
        aggregates = {vehicle_id: _VehicleAggregate(vehicle_id=vehicle_id, vehicle_type=vehicle_type) for vehicle_id, vehicle_type in vehicle_map.items()}

        fuel_rows = (
            await self.db.execute(
                select(
                    FuelSupply.vehicle_id,
                    func.max(FuelSupply.odometer_km).label("max_odometer"),
                    func.min(FuelSupply.odometer_km).label("min_odometer"),
                    func.sum(FuelSupply.liters).label("sum_liters"),
                    func.sum(func.coalesce(FuelSupply.total_amount, 0)).label("sum_fuel_cost"),
                    func.sum(case((FuelSupply.is_consumption_anomaly.is_(True), 1), else_=0)).label("anomalies_count"),
                )
                .where(FuelSupply.supplied_at >= period_start, FuelSupply.supplied_at <= period_end)
                .group_by(FuelSupply.vehicle_id)
            )
        ).all()
        for row in fuel_rows:
            item = aggregates.get(row.vehicle_id)
            if not item:
                continue
            km = float((row.max_odometer or 0) - (row.min_odometer or 0))
            item.total_km = max(km, 0.0)
            item.total_liters = float(row.sum_liters or 0)
            item.fuel_cost = float(row.sum_fuel_cost or 0)
            item.anomalies_count = int(row.anomalies_count or 0)

        maint_rows = (
            await self.db.execute(
                select(MaintenanceRecord.vehicle_id, func.sum(MaintenanceRecord.total_cost).label("sum_cost"))
                .where(MaintenanceRecord.start_date >= period_start, MaintenanceRecord.start_date <= period_end)
                .group_by(MaintenanceRecord.vehicle_id)
            )
        ).all()
        for row in maint_rows:
            item = aggregates.get(row.vehicle_id)
            if item:
                item.maintenance_cost = float(row.sum_cost or 0)

        fine_rows = (
            await self.db.execute(
                select(Fine.vehicle_id, func.sum(Fine.amount).label("sum_cost"))
                .where(Fine.infraction_date >= period_start.date(), Fine.infraction_date <= period_end.date())
                .group_by(Fine.vehicle_id)
            )
        ).all()
        for row in fine_rows:
            item = aggregates.get(row.vehicle_id)
            if item:
                item.fines_cost = float(row.sum_cost or 0)

        type_consumptions: dict[str, list[float]] = defaultdict(list)
        type_tco: dict[str, list[float]] = defaultdict(list)

        per_vehicle_metrics: dict[UUID, tuple[float | None, float | None]] = {}
        for vehicle_id, agg in aggregates.items():
            consumption = calculate_consumption_l_100km(agg.total_liters, agg.total_km)
            tco = calculate_tco_per_km(agg.fuel_cost, agg.maintenance_cost, agg.fines_cost, agg.total_km)
            per_vehicle_metrics[vehicle_id] = (consumption, tco)
            if consumption is not None:
                type_consumptions[agg.vehicle_type].append(consumption)
            if tco is not None:
                type_tco[agg.vehicle_type].append(tco)

        snapshots: list[FleetAnalyticsSnapshot] = []
        for vehicle_id, agg in aggregates.items():
            consumption, tco = per_vehicle_metrics[vehicle_id]
            avg_consumption = (
                sum(type_consumptions[agg.vehicle_type]) / len(type_consumptions[agg.vehicle_type]) if type_consumptions[agg.vehicle_type] else None
            )
            avg_tco = sum(type_tco[agg.vehicle_type]) / len(type_tco[agg.vehicle_type]) if type_tco[agg.vehicle_type] else None
            snapshots.append(
                FleetAnalyticsSnapshot(
                    period_start=period_start,
                    period_end=period_end,
                    scope="VEHICLE",
                    vehicle_id=vehicle_id,
                    vehicle_type=agg.vehicle_type,
                    total_km=agg.total_km,
                    total_liters=agg.total_liters,
                    fuel_cost=Decimal(str(round(agg.fuel_cost, 2))),
                    maintenance_cost=Decimal(str(round(agg.maintenance_cost, 2))),
                    fines_cost=Decimal(str(round(agg.fines_cost, 2))),
                    consumption_l_100km=consumption,
                    tco_cost_per_km=tco,
                    anomalies_count=agg.anomalies_count,
                    category_average_consumption=avg_consumption,
                    category_average_tco=avg_tco,
                    market_benchmark_tco=MARKET_TCO_BENCHMARK_BY_TYPE.get(agg.vehicle_type, avg_tco),
                )
            )

        driver_rows = (
            await self.db.execute(select(Driver.id, Driver.nome_completo).where(Driver.ativo.is_(True)))
        ).all()
        for driver in driver_rows:
            fines_count = int(
                (
                    await self.db.execute(
                        select(func.count(Fine.id)).where(
                            Fine.driver_id == driver.id,
                            Fine.infraction_date >= period_start.date(),
                            Fine.infraction_date <= period_end.date(),
                        )
                    )
                ).scalar_one()
                or 0
            )
            claims_count = int(
                (
                    await self.db.execute(
                        select(func.count(Claim.id)).where(
                            Claim.driver_id == driver.id,
                            Claim.data_ocorrencia >= period_start,
                            Claim.data_ocorrencia <= period_end,
                        )
                    )
                ).scalar_one()
                or 0
            )
            anomalies_count = int(
                (
                    await self.db.execute(
                        select(func.count(FuelSupply.id)).where(
                            FuelSupply.driver_id == driver.id,
                            FuelSupply.supplied_at >= period_start,
                            FuelSupply.supplied_at <= period_end,
                            FuelSupply.is_consumption_anomaly.is_(True),
                        )
                    )
                ).scalar_one()
                or 0
            )

            raw_score = calculate_driver_risk_score(
                fines_count=fines_count,
                claims_count=claims_count,
                anomalies_count=anomalies_count,
            )
            normalized = min(round(raw_score * 10, 2), 100)
            snapshots.append(
                FleetAnalyticsSnapshot(
                    period_start=period_start,
                    period_end=period_end,
                    scope="DRIVER",
                    driver_id=driver.id,
                    vehicle_type="N/A",
                    total_km=0,
                    total_liters=0,
                    fuel_cost=Decimal("0"),
                    maintenance_cost=Decimal("0"),
                    fines_cost=Decimal("0"),
                    driver_risk_score=normalized,
                    anomalies_count=anomalies_count,
                    notes=driver.nome_completo,
                    extra_payload={
                        "fines_count": fines_count,
                        "claims_count": claims_count,
                        "anomalies_count": anomalies_count,
                        "raw_score": raw_score,
                    },
                )
            )

        if not snapshots:
            return []

        await self.analytics_repo.replace_period_snapshots(period_start=period_start, period_end=period_end, items=snapshots)
        await self.db.commit()
        return snapshots

    async def overview(self, period_days: int, organization_id: UUID | None = None) -> dict:
        snapshots = await self._ensure_snapshots(period_days)
        vehicle_rows = await self._filter_vehicle_snapshots_by_organization(
            self._unique_vehicle_snapshots(snapshots),
            organization_id,
        )
        driver_rows = await self._filter_driver_snapshots_by_organization(
            self._unique_driver_snapshots(snapshots),
            organization_id,
        )
        insights = self._build_insights(vehicle_rows, driver_rows)
        consumptions = [item.consumption_l_100km for item in vehicle_rows if item.consumption_l_100km is not None]
        tcos = [item.tco_cost_per_km for item in vehicle_rows if item.tco_cost_per_km is not None]

        return {
            "period_days": period_days,
            "fleet_active": len(vehicle_rows),
            "average_consumption_l_100km": round(sum(consumptions) / len(consumptions), 2) if consumptions else 0,
            "average_tco_per_km": round(sum(tcos) / len(tcos), 2) if tcos else 0,
            "active_alerts": len(insights),
            "generated_at": datetime.now(timezone.utc),
        }

    async def efficiency(
        self,
        period_days: int,
        vehicle_type: str | None = None,
        organization_id: UUID | None = None,
    ) -> list[dict]:
        rows = await self._filter_vehicle_snapshots_by_organization(
            self._unique_vehicle_snapshots(await self._ensure_snapshots(period_days)),
            organization_id,
        )
        if vehicle_type:
            rows = [row for row in rows if row.vehicle_type == vehicle_type]
        payload = []
        for row in rows:
            variance = calculate_variance_percentage(row.consumption_l_100km, row.category_average_consumption)
            payload.append(
                {
                    "vehicle_type": row.vehicle_type,
                    "vehicle_id": row.vehicle_id,
                    "total_km": row.total_km,
                    "total_liters": row.total_liters,
                    "consumption_l_100km": row.consumption_l_100km,
                    "category_average": row.category_average_consumption,
                    "variance_percentage": round(variance, 2) if variance is not None else None,
                }
            )
        return sorted(payload, key=lambda item: (item["variance_percentage"] or 0), reverse=True)

    async def tco(
        self,
        period_days: int,
        vehicle_type: str | None = None,
        organization_id: UUID | None = None,
    ) -> list[dict]:
        rows = await self._filter_vehicle_snapshots_by_organization(
            self._unique_vehicle_snapshots(await self._ensure_snapshots(period_days)),
            organization_id,
        )
        if vehicle_type:
            rows = [row for row in rows if row.vehicle_type == vehicle_type]

        payload = []
        for row in rows:
            variance = calculate_variance_percentage(row.tco_cost_per_km, row.market_benchmark_tco)
            payload.append(
                {
                    "vehicle_type": row.vehicle_type,
                    "vehicle_id": row.vehicle_id,
                    "total_km": row.total_km,
                    "tco_cost_per_km": row.tco_cost_per_km,
                    "category_average": row.category_average_tco,
                    "market_benchmark": row.market_benchmark_tco,
                    "variance_percentage": round(variance, 2) if variance is not None else None,
                }
            )
        return sorted(payload, key=lambda item: (item["tco_cost_per_km"] or 0), reverse=True)

    async def driver_risk(self, period_days: int, organization_id: UUID | None = None) -> list[dict]:
        rows = await self._filter_driver_snapshots_by_organization(
            self._unique_driver_snapshots(await self._ensure_snapshots(period_days)),
            organization_id,
        )
        payload = []
        for row in rows:
            extra = row.extra_payload or {}
            payload.append(
                {
                    "driver_id": row.driver_id,
                    "driver_name": row.notes or "Condutor",
                    "fines_count": int(extra.get("fines_count", 0)),
                    "claims_count": int(extra.get("claims_count", 0)),
                    "anomalies_count": int(extra.get("anomalies_count", 0)),
                    "risk_score": float(extra.get("raw_score", 0)),
                    "normalized_risk_score": row.driver_risk_score or 0,
                }
            )
        return sorted(payload, key=lambda item: item["normalized_risk_score"], reverse=True)

    def _build_insights(self, vehicle_rows: list[FleetAnalyticsSnapshot], driver_rows: list[FleetAnalyticsSnapshot]) -> list[dict]:
        insights: list[dict] = []
        now = datetime.now(timezone.utc)

        for row in vehicle_rows:
            variance = calculate_variance_percentage(row.consumption_l_100km, row.category_average_consumption)
            if variance is not None and abs(variance) > 20:
                severity = "HIGH" if abs(variance) >= 40 else "MEDIUM"
                insights.append(
                    {
                        "vehicle_id": row.vehicle_id,
                        "vehicle_type": row.vehicle_type,
                        "metric": "consumption_l_100km",
                        "current_value": round(row.consumption_l_100km or 0, 2),
                        "category_average": round(row.category_average_consumption or 0, 2),
                        "variance_percentage": round(variance, 2),
                        "severity": severity,
                        "message": (
                            f"Veículo tipo {row.vehicle_type} apresenta consumo {abs(variance):.1f}% "
                            f"{'superior' if variance > 0 else 'inferior'} a média da categoria "
                            f"({(row.category_average_consumption or 0):.1f} L/100km)."
                        ),
                        "recommended_action": "Agendar inspeção mecânica preventiva",
                        "generated_at": now,
                    }
                )

            tco_var = calculate_variance_percentage(row.tco_cost_per_km, row.market_benchmark_tco)
            if tco_var is not None and abs(tco_var) > 30:
                insights.append(
                    {
                        "vehicle_id": row.vehicle_id,
                        "vehicle_type": row.vehicle_type,
                        "metric": "tco_cost_per_km",
                        "current_value": round(row.tco_cost_per_km or 0, 2),
                        "category_average": round(row.market_benchmark_tco or 0, 2),
                        "variance_percentage": round(tco_var, 2),
                        "severity": "HIGH" if abs(tco_var) >= 50 else "MEDIUM",
                        "message": (
                            f"TCO por km de {row.vehicle_type} está {abs(tco_var):.1f}% fora da referência "
                            f"de mercado ({(row.market_benchmark_tco or 0):.2f}/km)."
                        ),
                        "recommended_action": "Revisar plano de custos e manutenção",
                        "generated_at": now,
                    }
                )

        for row in driver_rows:
            if (row.driver_risk_score or 0) >= 70:
                extra = row.extra_payload or {}
                insights.append(
                    {
                        "driver_id": row.driver_id,
                        "vehicle_id": None,
                        "vehicle_type": "N/A",
                        "metric": "driver_risk_score",
                        "current_value": round(row.driver_risk_score or 0, 2),
                        "category_average": 70,
                        "variance_percentage": round((row.driver_risk_score or 0) - 70, 2),
                        "severity": "CRITICAL" if (row.driver_risk_score or 0) >= 85 else "HIGH",
                        "message": (
                            f"Condutor {row.notes or ''} com score de risco {row.driver_risk_score:.1f}/100 "
                            f"(multas={extra.get('fines_count', 0)}, sinistros={extra.get('claims_count', 0)}, anomalias={extra.get('anomalies_count', 0)})."
                        ),
                        "recommended_action": "Aplicar treinamento de direção defensiva e monitoramento semanal",
                        "generated_at": now,
                    }
                )

        severity_rank = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2}
        return sorted(insights, key=lambda item: severity_rank.get(item["severity"], 99))

    async def insights(
        self,
        period_days: int,
        vehicle_type: str | None = None,
        organization_id: UUID | None = None,
    ) -> list[dict]:
        rows = await self._ensure_snapshots(period_days)
        vehicle_rows = await self._filter_vehicle_snapshots_by_organization(
            self._unique_vehicle_snapshots(rows),
            organization_id,
        )
        if vehicle_type:
            vehicle_rows = [row for row in vehicle_rows if row.vehicle_type == vehicle_type]
        driver_rows = await self._filter_driver_snapshots_by_organization(
            self._unique_driver_snapshots(rows),
            organization_id,
        )
        return self._build_insights(
            vehicle_rows,
            driver_rows,
        )


    async def costs_trend(
        self,
        months: int = 12,
        vehicle_type: str | None = None,
        organization_id: UUID | None = None,
    ) -> list[dict]:
        now = datetime.now(timezone.utc)
        month_end = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        timeline: list[dict] = []
        organization_vehicle_ids = await self._vehicle_ids_for_organization(organization_id)

        for offset in range(months - 1, -1, -1):
            month_start = (month_end - timedelta(days=offset * 31)).replace(day=1)
            next_month = (month_start + timedelta(days=32)).replace(day=1)

            vehicle_filter = []
            if vehicle_type:
                vehicle_filter = [Vehicle.vehicle_type == vehicle_type]

            fuel_stmt = select(func.sum(func.coalesce(FuelSupply.total_amount, 0))).where(
                FuelSupply.supplied_at >= month_start,
                FuelSupply.supplied_at < next_month,
            )
            maint_stmt = select(func.sum(func.coalesce(MaintenanceRecord.total_cost, 0))).where(
                MaintenanceRecord.start_date >= month_start,
                MaintenanceRecord.start_date < next_month,
            )
            fine_stmt = select(func.sum(func.coalesce(Fine.amount, 0))).where(
                Fine.infraction_date >= month_start.date(),
                Fine.infraction_date < next_month.date(),
            )

            if vehicle_filter:
                fuel_stmt = fuel_stmt.join(Vehicle, Vehicle.id == FuelSupply.vehicle_id).where(*vehicle_filter)
                maint_stmt = maint_stmt.join(Vehicle, Vehicle.id == MaintenanceRecord.vehicle_id).where(*vehicle_filter)
                fine_stmt = fine_stmt.join(Vehicle, Vehicle.id == Fine.vehicle_id).where(*vehicle_filter)

            if organization_vehicle_ids is not None:
                fuel_stmt = fuel_stmt.where(FuelSupply.vehicle_id.in_(organization_vehicle_ids))
                maint_stmt = maint_stmt.where(MaintenanceRecord.vehicle_id.in_(organization_vehicle_ids))
                fine_stmt = fine_stmt.where(Fine.vehicle_id.in_(organization_vehicle_ids))

            fuel_cost = float((await self.db.execute(fuel_stmt)).scalar_one() or 0)
            maintenance_cost = float((await self.db.execute(maint_stmt)).scalar_one() or 0)
            fines_cost = float((await self.db.execute(fine_stmt)).scalar_one() or 0)
            total_cost = fuel_cost + maintenance_cost + fines_cost

            timeline.append(
                {
                    "month": month_start.strftime("%m/%Y"),
                    "fuel_cost": round(fuel_cost, 2),
                    "maintenance_cost": round(maintenance_cost, 2),
                    "fines_cost": round(fines_cost, 2),
                    "total_cost": round(total_cost, 2),
                }
            )

        return timeline

    @staticmethod
    def _build_export_xlsx(insights: list[dict], period_days: int, issued_at: datetime) -> bytes:
        workbook = Workbook()
        workbook.properties.creator = MUNICIPALITY_NAME
        workbook.properties.title = "Relatório de análises da frota"
        workbook.properties.subject = f"{ADMINISTRATION_SECRETARIAT} · {FLEET_DEPARTMENT}"
        workbook.properties.description = "Documento oficial de gestão da frota municipal"

        worksheet = workbook.active
        worksheet.title = "Análises"
        worksheet.sheet_view.showGridLines = False
        worksheet.merge_cells("B1:F1")
        worksheet.merge_cells("B2:F2")
        worksheet.merge_cells("A4:F4")
        worksheet.merge_cells("A5:F5")
        worksheet.merge_cells("A6:F6")

        crest = WorksheetImage(str(crest_path()))
        crest.width = 48
        crest.height = 60
        worksheet.add_image(crest, "A1")
        worksheet.row_dimensions[1].height = 24
        worksheet.row_dimensions[2].height = 21
        worksheet.row_dimensions[3].height = 8
        worksheet["B1"] = MUNICIPALITY_NAME.upper()
        worksheet["B1"].font = Font(name="Roboto", size=13, bold=True, color=COLOR_NAVY.lstrip("#"))
        worksheet["B1"].alignment = Alignment(vertical="bottom")
        worksheet["B2"] = f"{ADMINISTRATION_SECRETARIAT} · {FLEET_DEPARTMENT}"
        worksheet["B2"].font = Font(name="Roboto", size=9, color=COLOR_MUTED.lstrip("#"))
        worksheet["B2"].alignment = Alignment(vertical="top")

        worksheet["A4"] = "RELATÓRIO DE ANÁLISES DA FROTA"
        worksheet["A4"].font = Font(name="Roboto", size=16, bold=True, color=COLOR_NAVY.lstrip("#"))
        worksheet["A4"].alignment = Alignment(horizontal="left")
        worksheet["A5"] = f"Período analisado: últimos {period_days} dias"
        worksheet["A6"] = f"Emissão: {_institutional_timestamp(issued_at)} · Alertas identificados: {len(insights)}"
        for coordinate in ("A5", "A6"):
            worksheet[coordinate].font = Font(name="Roboto", size=9, color=COLOR_MUTED.lstrip("#"))

        header_row = 8
        navy_fill = PatternFill("solid", fgColor=COLOR_NAVY.lstrip("#"))
        alternate_fill = PatternFill("solid", fgColor=COLOR_SURFACE.lstrip("#"))
        border = Border(
            left=Side(style="thin", color=COLOR_BORDER.lstrip("#")),
            right=Side(style="thin", color=COLOR_BORDER.lstrip("#")),
            top=Side(style="thin", color=COLOR_BORDER.lstrip("#")),
            bottom=Side(style="thin", color=COLOR_BORDER.lstrip("#")),
        )
        for column_index, (_, header, width) in enumerate(ANALYTICS_EXPORT_COLUMNS, start=1):
            cell = worksheet.cell(row=header_row, column=column_index, value=header)
            cell.font = Font(name="Roboto", size=9, bold=True, color="FFFFFF")
            cell.fill = navy_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
            worksheet.column_dimensions[get_column_letter(column_index)].width = width
        worksheet.row_dimensions[header_row].height = 28

        for row_index, item in enumerate(insights, start=header_row + 1):
            for column_index, (key, _, _) in enumerate(ANALYTICS_EXPORT_COLUMNS, start=1):
                value = _analytics_xlsx_value(item.get(key))
                cell = worksheet.cell(row=row_index, column=column_index, value=value)
                cell.font = Font(name="Roboto", size=9, color="20304A")
                cell.alignment = Alignment(vertical="top", wrap_text=column_index in {1, 2, 6})
                cell.border = border
                if row_index % 2 == 0:
                    cell.fill = alternate_fill
                if key in {"current_value", "category_average", "variance_percentage"} and isinstance(value, (int, float)):
                    cell.number_format = "0.00"

        last_row = max(header_row, header_row + len(insights))
        worksheet.freeze_panes = f"A{header_row + 1}"
        worksheet.auto_filter.ref = f"A{header_row}:F{last_row}"
        worksheet.print_title_rows = f"1:{header_row}"
        worksheet.print_area = f"A1:F{last_row}"
        worksheet.page_setup.orientation = "landscape"
        worksheet.page_setup.fitToWidth = 1
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.oddFooter.center.text = f"{MUNICIPALITY_NAME} · CNPJ {MUNICIPALITY_CNPJ}"
        worksheet.oddFooter.right.text = "Página &P de &N"

        metadata = workbook.create_sheet("Informações")
        metadata.sheet_view.showGridLines = False
        metadata_rows = (
            ("Órgão emissor", MUNICIPALITY_NAME),
            ("Unidade", f"{ADMINISTRATION_SECRETARIAT} · {FLEET_DEPARTMENT}"),
            ("CNPJ", MUNICIPALITY_CNPJ),
            ("Endereço", MUNICIPALITY_ADDRESS),
            ("Documento", "Relatório de análises da frota"),
            ("Período", f"Últimos {period_days} dias"),
            ("Emissão", _institutional_timestamp(issued_at)),
            ("Alertas identificados", len(insights)),
        )
        for row_index, (label, value) in enumerate(metadata_rows, start=1):
            label_cell = metadata.cell(row=row_index, column=1, value=label)
            value_cell = metadata.cell(row=row_index, column=2, value=_analytics_xlsx_value(value))
            label_cell.font = Font(name="Roboto", size=9, bold=True, color=COLOR_NAVY.lstrip("#"))
            value_cell.font = Font(name="Roboto", size=9, color="20304A")
            value_cell.alignment = Alignment(wrap_text=True, vertical="top")
        metadata.column_dimensions["A"].width = 24
        metadata.column_dimensions["B"].width = 86

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    @staticmethod
    def _build_export_pdf(insights: list[dict], period_days: int, issued_at: datetime) -> bytes:
        output = BytesIO()
        font_regular, font_bold = ensure_pdf_fonts()
        page_size = landscape(A4)
        document = SimpleDocTemplate(
            output,
            pagesize=page_size,
            leftMargin=12 * mm,
            rightMargin=12 * mm,
            topMargin=11 * mm,
            bottomMargin=21 * mm,
            title="Relatório de análises da frota",
            author=MUNICIPALITY_NAME,
            subject=f"{ADMINISTRATION_SECRETARIAT} · {FLEET_DEPARTMENT}",
        )
        styles = getSampleStyleSheet()
        body_style = ParagraphStyle(
            "AnalyticsBody",
            parent=styles["BodyText"],
            fontName=font_regular,
            fontSize=7.2,
            leading=9.2,
            textColor=colors.HexColor("#20304A"),
        )
        small_style = ParagraphStyle(
            "AnalyticsSmall",
            parent=body_style,
            fontSize=6.5,
            leading=8,
            textColor=colors.HexColor(COLOR_MUTED),
        )
        institution_style = ParagraphStyle(
            "AnalyticsInstitution",
            parent=body_style,
            fontName=font_bold,
            fontSize=9,
            leading=11,
            textColor=colors.HexColor(COLOR_NAVY),
        )
        header_style = ParagraphStyle(
            "AnalyticsHeader",
            parent=body_style,
            fontName=font_bold,
            textColor=colors.white,
        )
        title_style = ParagraphStyle(
            "AnalyticsTitle",
            parent=styles["Title"],
            fontName=font_bold,
            fontSize=15,
            leading=18,
            textColor=colors.HexColor(COLOR_NAVY),
            spaceAfter=2 * mm,
        )
        available_width = page_size[0] - 24 * mm
        crest = Image(str(crest_path()), width=10 * mm, height=12.6 * mm)
        identity = [
            Paragraph(MUNICIPALITY_NAME.upper(), institution_style),
            Paragraph(
                _analytics_pdf_text(f"{ADMINISTRATION_SECRETARIAT} · {FLEET_DEPARTMENT}"),
                small_style,
            ),
        ]
        identity_header = Table([[crest, identity]], colWidths=[14 * mm, available_width - 14 * mm], hAlign="LEFT")
        identity_header.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3 * mm),
                    ("LINEBELOW", (0, 0), (-1, -1), 0.8, colors.HexColor(COLOR_NAVY)),
                ]
            )
        )
        story = [
            identity_header,
            Spacer(1, 3 * mm),
            Paragraph("RELATÓRIO DE ANÁLISES DA FROTA", title_style),
            Paragraph(
                f"Período analisado: últimos {period_days} dias · "
                f"Emissão: {_analytics_pdf_text(_institutional_timestamp(issued_at))} · "
                f"Alertas identificados: {len(insights)}",
                body_style,
            ),
            Spacer(1, 5 * mm),
        ]

        if insights:
            table_rows = [
                [
                    Paragraph("Severidade", header_style),
                    Paragraph("Métrica", header_style),
                    Paragraph("Diagnóstico", header_style),
                    Paragraph("Ação recomendada", header_style),
                ]
            ]
            for item in insights:
                severity = ANALYTICS_SEVERITY_LABELS.get(str(item.get("severity", "")), item.get("severity", "-"))
                metric = ANALYTICS_METRIC_LABELS.get(str(item.get("metric", "")), item.get("metric", "-"))
                measurements = (
                    f"Valor atual: {_analytics_pdf_text(item.get('current_value'))} · "
                    f"Referência: {_analytics_pdf_text(item.get('category_average'))} · "
                    f"Variação: {_analytics_pdf_text(item.get('variance_percentage'))}%"
                )
                diagnosis = (
                    f"{_analytics_pdf_text(item.get('message'))}<br/>"
                    f"<font color='{COLOR_MUTED}' size='6.5'>{measurements}</font>"
                )
                table_rows.append(
                    [
                        Paragraph(_analytics_pdf_text(severity), body_style),
                        Paragraph(_analytics_pdf_text(metric), body_style),
                        Paragraph(diagnosis, body_style),
                        Paragraph(_analytics_pdf_text(item.get("recommended_action")), body_style),
                    ]
                )
            table = LongTable(
                table_rows,
                colWidths=[25 * mm, 44 * mm, 100 * mm, available_width - (169 * mm)],
                repeatRows=1,
                hAlign="LEFT",
            )
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(COLOR_NAVY)),
                        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor(COLOR_BORDER)),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(COLOR_SURFACE)]),
                        ("LEFTPADDING", (0, 0), (-1, -1), 4),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                        ("TOPPADDING", (0, 0), (-1, -1), 4),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ]
                )
            )
            story.append(table)
        else:
            story.append(
                Paragraph(
                    "Não foram identificados alertas analíticos para o período e os filtros informados.",
                    body_style,
                )
            )

        def footer(canvas, doc):
            canvas.saveState()
            canvas.setStrokeColor(colors.HexColor(COLOR_NAVY))
            canvas.setLineWidth(0.45)
            canvas.line(12 * mm, 15 * mm, page_size[0] - 12 * mm, 15 * mm)
            canvas.setFillColor(colors.HexColor(COLOR_MUTED))
            canvas.setFont(font_regular, 6.2)
            canvas.drawString(12 * mm, 11.5 * mm, f"{MUNICIPALITY_NAME} · CNPJ {MUNICIPALITY_CNPJ}")
            canvas.drawString(12 * mm, 8.5 * mm, MUNICIPALITY_ADDRESS)
            canvas.drawRightString(page_size[0] - 12 * mm, 11.5 * mm, f"Página {doc.page}")
            canvas.restoreState()

        document.build(story, onFirstPage=footer, onLaterPages=footer)
        return output.getvalue()

    async def export(
        self,
        period_days: int,
        export_format: str,
        vehicle_type: str | None = None,
        organization_id: UUID | None = None,
    ) -> Response:
        if export_format not in {"pdf", "xlsx"}:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Formato de exportação suportado: pdf ou xlsx")

        insights = await self.insights(period_days, vehicle_type=vehicle_type, organization_id=organization_id)
        timestamp = datetime.now(timezone.utc)
        headers = {
            **ANALYTICS_EXPORT_NO_CACHE_HEADERS,
            "Content-Disposition": f'attachment; filename="analytics-{period_days}d.{export_format}"',
        }
        if export_format == "xlsx":
            return Response(
                content=self._build_export_xlsx(insights, period_days, timestamp),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers=headers,
            )

        return Response(
            content=self._build_export_pdf(insights, period_days, timestamp),
            media_type="application/pdf",
            headers=headers,
        )
