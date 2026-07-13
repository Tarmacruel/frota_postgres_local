from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO
from time import perf_counter
from typing import Any

from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.organization_scope import production_scope_is_empty, scoped_organization_id
from app.models.possession import VehiclePossession
from app.models.possession_trip import VehiclePossessionTrip
from app.models.user import User, UserRole
from app.repositories.possession_report_repository import PossessionReportRepository
from app.repositories.user_report_preference_repository import UserReportPreferenceRepository
from app.schemas.possession_report import (
    PossessionReportMode,
    PossessionReportOrientation,
    PossessionReportPreferenceIn,
    PossessionReportPreset,
    PossessionReportRequest,
)
from app.services.audit_service import AuditService
from app.services.possession_report_registry import (
    REPORT_COLUMN_BY_KEY,
    ReportColumn,
    ReportRowContext,
    ReportValueKind,
    columns_for_role_and_mode,
    preset_columns,
)


REPORT_NO_CACHE_HEADERS = {
    "Cache-Control": "private, no-store, no-cache, max-age=0, must-revalidate",
    "Pragma": "no-cache",
    "Expires": "0",
    "X-Content-Type-Options": "nosniff",
}
PDF_MAX_ROWS = 1_500
XLSX_MAX_ROWS = 5_000
PDF_MAX_PORTRAIT_COLUMNS = 10
PDF_MAX_LANDSCAPE_COLUMNS = 18
REPORT_PREFERENCE_TYPE = "possession"


@dataclass(frozen=True, slots=True)
class PreparedCell:
    raw: Any
    display: str


@dataclass(frozen=True, slots=True)
class PreparedReport:
    request: PossessionReportRequest
    columns: tuple[ReportColumn, ...]
    rows: tuple[tuple[PreparedCell, ...], ...]

    @property
    def row_count(self) -> int:
        return len(self.rows)


def neutralize_spreadsheet_text(value: str) -> str:
    """Prevent spreadsheet formula execution, including whitespace-prefixed payloads."""
    probe = value.lstrip(" \t\r\n")
    if probe.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


class PossessionReportService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.reports = PossessionReportRepository(db)
        self.preferences = UserReportPreferenceRepository(db)
        self.audit = AuditService(db)

    @classmethod
    def metadata(cls, current_user: User) -> dict:
        role = cls._role(current_user)
        modes = []
        for mode in PossessionReportMode:
            columns = columns_for_role_and_mode(role, mode)
            presets = [
                preset
                for preset in PossessionReportPreset
                if cls._preset_is_available(role, preset)
            ]
            modes.append(
                {
                    "key": mode.value,
                    "title": "Por posse" if mode == PossessionReportMode.POSSESSION else "Por rota",
                    "columns": [cls._column_metadata(column) for column in columns],
                    "presets": [
                        {
                            "key": preset.value,
                            "title": cls._preset_title(preset),
                            "column_keys": [
                                column.key for column in preset_columns(role, mode, preset)
                            ] if preset != PossessionReportPreset.CUSTOM else [],
                        }
                        for preset in presets
                    ],
                }
            )
        return {
            "default_mode": PossessionReportMode.POSSESSION.value,
            "default_preset": PossessionReportPreset.SUMMARY.value,
            "can_export_xlsx": role in {UserRole.ADMIN, UserRole.PRODUCAO},
            "limits": {
                "maximum_period_days": 366,
                "pdf_rows": PDF_MAX_ROWS,
                "xlsx_rows": XLSX_MAX_ROWS,
                "pdf_portrait_columns": PDF_MAX_PORTRAIT_COLUMNS,
                "pdf_landscape_columns": PDF_MAX_LANDSCAPE_COLUMNS,
            },
            "modes": modes,
        }

    async def render_pdf(self, data: PossessionReportRequest, current_user: User) -> tuple[bytes, str]:
        started = perf_counter()
        try:
            prepared = await self.prepare(data, current_user, row_limit=PDF_MAX_ROWS)
            max_columns = (
                PDF_MAX_PORTRAIT_COLUMNS
                if data.orientation == PossessionReportOrientation.PORTRAIT
                else PDF_MAX_LANDSCAPE_COLUMNS
            )
            if len(prepared.columns) > max_columns:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail={
                        "code": "REPORT_PDF_COLUMN_LIMIT_EXCEEDED",
                        "message": f"Reduza a seleção para no máximo {max_columns} colunas neste formato.",
                    },
                )
            content = self._build_pdf(prepared, current_user)
        except Exception as exc:
            await self._audit_report_failure(
                action="REPORT_PREVIEW",
                request=data,
                current_user=current_user,
                duration_ms=self._duration_ms(started),
                exc=exc,
            )
            raise
        await self._audit_report(
            action="REPORT_PREVIEW",
            prepared=prepared,
            current_user=current_user,
            duration_ms=self._duration_ms(started),
            outcome="SUCCESS",
        )
        return content, self._filename(prepared.request.mode, "pdf")

    async def render_xlsx(self, data: PossessionReportRequest, current_user: User) -> tuple[bytes, str]:
        started = perf_counter()
        role = self._role(current_user)
        if role not in {UserRole.ADMIN, UserRole.PRODUCAO}:
            exc = HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail={
                    "code": "REPORT_OPERATIONAL_EXPORT_FORBIDDEN",
                    "message": "Seu perfil pode previsualizar o relatório resumido, mas não exportar XLSX.",
                },
            )
            await self._audit_report_failure(
                action="REPORT_EXPORT_XLSX",
                request=data,
                current_user=current_user,
                duration_ms=self._duration_ms(started),
                exc=exc,
            )
            raise exc
        try:
            prepared = await self.prepare(data, current_user, row_limit=XLSX_MAX_ROWS)
            content = self._build_xlsx(prepared, current_user)
        except Exception as exc:
            await self._audit_report_failure(
                action="REPORT_EXPORT_XLSX",
                request=data,
                current_user=current_user,
                duration_ms=self._duration_ms(started),
                exc=exc,
            )
            raise
        await self._audit_report(
            action="REPORT_EXPORT_XLSX",
            prepared=prepared,
            current_user=current_user,
            duration_ms=self._duration_ms(started),
            outcome="SUCCESS",
        )
        return content, self._filename(prepared.request.mode, "xlsx")

    async def prepare(
        self,
        data: PossessionReportRequest,
        current_user: User,
        *,
        row_limit: int,
    ) -> PreparedReport:
        role = self._role(current_user)
        columns = self._resolve_columns(data, role)
        if production_scope_is_empty(current_user):
            records: list[VehiclePossession | VehiclePossessionTrip] = []
        else:
            records = await self.reports.load(
                mode=data.mode,
                filters=data.filters,
                organization_id=scoped_organization_id(current_user, data.filters.organization_id),
                limit=row_limit,
                include_operational_search=role in {UserRole.ADMIN, UserRole.PRODUCAO},
            )
        if len(records) > row_limit:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail={
                    "code": "REPORT_VOLUME_LIMIT_EXCEEDED",
                    "message": f"O relatório excedeu {row_limit} registros. Reduza o período ou aplique mais filtros.",
                    "limit": row_limit,
                },
            )
        contexts = [
            ReportRowContext(possession=record, trip=None)
            if isinstance(record, VehiclePossession)
            else ReportRowContext(possession=record.possession, trip=record)
            for record in records
        ]
        rows = tuple(
            tuple(self._prepare_cell(column, context) for column in columns)
            for context in contexts
        )
        return PreparedReport(request=data, columns=tuple(columns), rows=rows)

    async def get_preference(self, current_user: User) -> dict:
        role = self._role(current_user)
        stored = await self.preferences.get(user_id=current_user.id, report_type=REPORT_PREFERENCE_TYPE)
        if stored is None:
            return self._default_preference(role, sanitized=False)
        config = stored.config if isinstance(stored.config, dict) else {}
        try:
            preference = PossessionReportPreferenceIn.model_validate(config)
            normalized = self._normalize_preference(preference, role)
            sanitized = normalized != preference.model_dump(mode="json")
            return {**normalized, "sanitized": sanitized}
        except (ValueError, HTTPException):
            return self._default_preference(role, sanitized=True)

    async def update_preference(self, data: PossessionReportPreferenceIn, current_user: User) -> dict:
        role = self._role(current_user)
        config = self._normalize_preference(data, role)
        await self.preferences.upsert(
            user_id=current_user.id,
            report_type=REPORT_PREFERENCE_TYPE,
            config=config,
        )
        await self.audit.record(
            actor=current_user,
            action="REPORT_PREFERENCE_UPDATE",
            entity_type="USER_REPORT_PREFERENCE",
            entity_id=current_user.id,
            entity_label="Preferência de relatório de posses",
            details={
                "report_type": REPORT_PREFERENCE_TYPE,
                "mode": config["mode"],
                "preset": config["preset"],
                "column_keys": config["column_keys"],
            },
        )
        await self.db.commit()
        return {**config, "sanitized": False}

    def _resolve_columns(self, data: PossessionReportRequest, role: UserRole) -> list[ReportColumn]:
        if not self._preset_is_available(role, data.preset):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail={"code": "REPORT_PRESET_FORBIDDEN", "message": "Preset não autorizado para este perfil."},
            )
        if data.preset != PossessionReportPreset.CUSTOM:
            return preset_columns(role, data.mode, data.preset)
        columns: list[ReportColumn] = []
        for key in data.column_keys or []:
            column = REPORT_COLUMN_BY_KEY.get(key)
            if column is None:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail={"code": "REPORT_COLUMN_UNKNOWN", "message": "A seleção contém uma coluna desconhecida."},
                )
            if role not in column.roles:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail={"code": "REPORT_COLUMN_FORBIDDEN", "message": "A seleção contém uma coluna restrita."},
                )
            if data.mode not in column.modes:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail={"code": "REPORT_COLUMN_MODE_MISMATCH", "message": "A coluna não é compatível com o modo escolhido."},
                )
            columns.append(column)
        return columns

    def _normalize_preference(self, data: PossessionReportPreferenceIn, role: UserRole) -> dict:
        if not self._preset_is_available(role, data.preset):
            raise HTTPException(status_code=403, detail={"code": "REPORT_PRESET_FORBIDDEN"})
        if data.preset == PossessionReportPreset.CUSTOM:
            request = PossessionReportRequest(
                mode=data.mode,
                preset=data.preset,
                column_keys=data.column_keys,
            )
            columns = self._resolve_columns(request, role)
        else:
            columns = preset_columns(role, data.mode, data.preset)
        return {
            "mode": data.mode.value,
            "preset": data.preset.value,
            "column_keys": [column.key for column in columns],
        }

    def _default_preference(self, role: UserRole, *, sanitized: bool) -> dict:
        columns = preset_columns(role, PossessionReportMode.POSSESSION, PossessionReportPreset.SUMMARY)
        return {
            "mode": PossessionReportMode.POSSESSION.value,
            "preset": PossessionReportPreset.SUMMARY.value,
            "column_keys": [column.key for column in columns],
            "sanitized": sanitized,
        }

    async def _audit_report(
        self,
        *,
        action: str,
        prepared: PreparedReport,
        current_user: User,
        duration_ms: int,
        outcome: str,
    ) -> None:
        await self.audit.record(
            actor=current_user,
            action=action,
            entity_type="POSSESSION_REPORT",
            entity_id=current_user.id,
            entity_label="Relatório configurável de posses",
            details={
                "mode": prepared.request.mode.value,
                "preset": prepared.request.preset.value,
                "column_keys": [column.key for column in prepared.columns],
                "filters": self._audit_filters(prepared.request),
                "record_count": prepared.row_count,
                "duration_ms": duration_ms,
                "outcome": outcome,
            },
        )
        await self.db.commit()

    async def _audit_report_failure(
        self,
        *,
        action: str,
        request: PossessionReportRequest,
        current_user: User,
        duration_ms: int,
        exc: Exception,
    ) -> None:
        detail = getattr(exc, "detail", None)
        error_code = detail.get("code") if isinstance(detail, dict) else type(exc).__name__
        try:
            await self.db.rollback()
            await self.audit.record(
                actor=current_user,
                action=action,
                entity_type="POSSESSION_REPORT",
                entity_id=current_user.id,
                entity_label="Relatório configurável de posses",
                details={
                    "mode": request.mode.value,
                    "preset": request.preset.value,
                    "column_keys": request.column_keys or [],
                    "filters": self._audit_filters(request),
                    "record_count": 0,
                    "duration_ms": duration_ms,
                    "outcome": "FAILURE",
                    "error_code": error_code,
                },
            )
            await self.db.commit()
        except Exception:
            await self.db.rollback()

    def _build_pdf(self, prepared: PreparedReport, current_user: User) -> bytes:
        buffer = BytesIO()
        page_size = A4 if prepared.request.orientation == PossessionReportOrientation.PORTRAIT else landscape(A4)
        document = SimpleDocTemplate(
            buffer,
            pagesize=page_size,
            leftMargin=12 * mm,
            rightMargin=12 * mm,
            topMargin=12 * mm,
            bottomMargin=12 * mm,
            title="Relatório configurável de posses",
            author="Frota PMTF",
        )
        styles = getSampleStyleSheet()
        small = ParagraphStyle("ReportSmall", parent=styles["BodyText"], fontSize=7, leading=9)
        header = ParagraphStyle("ReportHeader", parent=small, textColor=colors.white, fontName="Helvetica-Bold")
        story = [
            Paragraph("Relatório de posses e rotas", styles["Title"]),
            Paragraph(
                f"Modo: {self._mode_title(prepared.request.mode)} · Preset: {self._preset_title(prepared.request.preset)} · "
                f"Gerado em {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M UTC')} · "
                f"Usuário: {self._safe_pdf_text(current_user.name)} · Registros: {prepared.row_count}",
                styles["BodyText"],
            ),
            Paragraph(self._pdf_filter_summary(prepared.request), small),
            Spacer(1, 5 * mm),
        ]
        table_rows = [
            [Paragraph(self._safe_pdf_text(column.title), header) for column in prepared.columns],
            *[
                [Paragraph(self._safe_pdf_text(cell.display), small) for cell in row]
                for row in prepared.rows
            ],
        ]
        available_width = page_size[0] - 24 * mm
        widths = self._pdf_widths(prepared.columns, available_width)
        table = Table(table_rows, colWidths=widths, repeatRows=1, hAlign="LEFT")
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#243447")),
                    ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#B8C2CC")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7F9")]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        story.append(table)
        document.build(story)
        return buffer.getvalue()

    def _build_xlsx(self, prepared: PreparedReport, current_user: User) -> bytes:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Posses" if prepared.request.mode == PossessionReportMode.POSSESSION else "Rotas"
        for index, column in enumerate(prepared.columns, start=1):
            cell = worksheet.cell(row=1, column=index, value=column.title)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="243447")
            worksheet.column_dimensions[get_column_letter(index)].width = min(max(column.suggested_width, 10), 45)
        for row_index, row in enumerate(prepared.rows, start=2):
            for column_index, (column, prepared_cell) in enumerate(zip(prepared.columns, row, strict=True), start=1):
                value = self._xlsx_value(prepared_cell.raw, column.value_kind)
                cell = worksheet.cell(row=row_index, column=column_index, value=value)
                if column.value_kind == ReportValueKind.DATETIME and value is not None:
                    cell.number_format = "dd/mm/yyyy hh:mm"
                elif column.value_kind == ReportValueKind.DECIMAL and value is not None:
                    cell.number_format = "0.0"
        worksheet.freeze_panes = "A2"
        if prepared.columns:
            worksheet.auto_filter.ref = f"A1:{get_column_letter(len(prepared.columns))}{max(prepared.row_count + 1, 1)}"
        metadata = workbook.create_sheet("Metadados")
        metadata_rows = [
            ("Relatório", "Posses e rotas"),
            ("Modo", self._mode_title(prepared.request.mode)),
            ("Preset", self._preset_title(prepared.request.preset)),
            ("Gerado em", datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC")),
            ("Usuário", current_user.name),
            ("Registros", prepared.row_count),
            ("Filtros", self._plain_filter_summary(prepared.request)),
        ]
        for row in metadata_rows:
            metadata.append([neutralize_spreadsheet_text(str(row[0])), self._xlsx_text(row[1])])
        metadata.column_dimensions["A"].width = 22
        metadata.column_dimensions["B"].width = 80
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def _prepare_cell(column: ReportColumn, context: ReportRowContext) -> PreparedCell:
        raw = column.extractor(context)
        if raw is None or raw == "":
            return PreparedCell(raw=None, display="-")
        if column.value_kind == ReportValueKind.DATETIME and isinstance(raw, datetime):
            normalized = raw.astimezone(timezone.utc) if raw.utcoffset() is not None else raw.replace(tzinfo=timezone.utc)
            return PreparedCell(raw=normalized, display=normalized.strftime("%d/%m/%Y %H:%M UTC"))
        if column.value_kind == ReportValueKind.DECIMAL:
            number = Decimal(str(raw))
            return PreparedCell(raw=number, display=f"{number:.1f}".replace(".", ","))
        if column.value_kind == ReportValueKind.INTEGER:
            return PreparedCell(raw=int(raw), display=str(int(raw)))
        text = str(getattr(raw, "value", raw))
        return PreparedCell(raw=text, display=text)

    @staticmethod
    def _xlsx_value(raw: Any, value_kind: ReportValueKind) -> Any:
        if raw is None:
            return None
        if value_kind == ReportValueKind.DATETIME and isinstance(raw, datetime):
            normalized = raw.astimezone(timezone.utc) if raw.utcoffset() is not None else raw
            return normalized.replace(tzinfo=None)
        if value_kind == ReportValueKind.DECIMAL:
            return float(raw)
        if value_kind == ReportValueKind.INTEGER:
            return int(raw)
        return neutralize_spreadsheet_text(str(raw))

    @staticmethod
    def _xlsx_text(value: Any) -> Any:
        if isinstance(value, (int, float)):
            return value
        return neutralize_spreadsheet_text(str(value))

    @staticmethod
    def _column_metadata(column: ReportColumn) -> dict:
        return {
            "key": column.key,
            "title": column.title,
            "category": column.category.value,
            "value_type": column.value_kind.value,
            "classification": column.classification.value,
            "contains_personal_data": column.contains_personal_data,
            "masking_rule": column.masking_rule,
            "suggested_width": column.suggested_width,
            "manual_only": column.manual_only,
            "presets": [preset.value for preset in column.presets],
        }

    @staticmethod
    def _preset_is_available(role: UserRole, preset: PossessionReportPreset) -> bool:
        if preset in {PossessionReportPreset.SUMMARY, PossessionReportPreset.CUSTOM}:
            return role in {UserRole.ADMIN, UserRole.PRODUCAO, UserRole.PADRAO}
        return role in {UserRole.ADMIN, UserRole.PRODUCAO}

    @staticmethod
    def _role(current_user: User) -> UserRole:
        role = getattr(current_user, "role", None)
        return role if isinstance(role, UserRole) else UserRole(str(getattr(role, "value", role)))

    @staticmethod
    def _preset_title(preset: PossessionReportPreset) -> str:
        return {
            PossessionReportPreset.SUMMARY: "Resumido",
            PossessionReportPreset.OPERATIONAL: "Operacional",
            PossessionReportPreset.COMPLETE: "Completo",
            PossessionReportPreset.CUSTOM: "Personalizado",
        }[preset]

    @staticmethod
    def _mode_title(mode: PossessionReportMode) -> str:
        return "Por posse" if mode == PossessionReportMode.POSSESSION else "Por rota"

    @staticmethod
    def _duration_ms(started: float) -> int:
        return max(0, round((perf_counter() - started) * 1000))

    @staticmethod
    def _filename(mode: PossessionReportMode, extension: str) -> str:
        suffix = "posses" if mode == PossessionReportMode.POSSESSION else "rotas"
        return f"relatorio-{suffix}-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.{extension}"

    @staticmethod
    def _safe_pdf_text(value: Any) -> str:
        return str(value or "-").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    @staticmethod
    def _pdf_widths(columns: tuple[ReportColumn, ...], available_width: float) -> list[float]:
        total = sum(column.suggested_width for column in columns) or 1
        return [available_width * column.suggested_width / total for column in columns]

    def _pdf_filter_summary(self, request: PossessionReportRequest) -> str:
        return self._safe_pdf_text(f"Filtros: {self._plain_filter_summary(request)}")

    @staticmethod
    def _plain_filter_summary(request: PossessionReportRequest) -> str:
        filters = request.filters
        values = [f"critério temporal={filters.temporal_field.value}"]
        for label, value in (
            ("de", filters.date_from.isoformat() if filters.date_from else None),
            ("até", filters.date_to.isoformat() if filters.date_to else None),
            ("veículo", filters.vehicle_id),
            ("condutor", filters.driver_id),
            ("secretaria", filters.organization_id),
            ("status da posse", getattr(filters.possession_status, "value", None)),
            ("status da rota", getattr(filters.trip_status, "value", None)),
            ("com retorno", filters.has_return),
            ("com confirmação", filters.has_return_confirmation),
            ("busca", filters.search),
        ):
            if value is not None:
                values.append(f"{label}={value}")
        return "; ".join(values)

    @staticmethod
    def _audit_filters(request: PossessionReportRequest) -> dict:
        filters = request.filters
        return {
            "date_from": filters.date_from.isoformat() if filters.date_from else None,
            "date_to": filters.date_to.isoformat() if filters.date_to else None,
            "temporal_field": filters.temporal_field.value,
            "vehicle_id": str(filters.vehicle_id) if filters.vehicle_id else None,
            "driver_id": str(filters.driver_id) if filters.driver_id else None,
            "organization_id": str(filters.organization_id) if filters.organization_id else None,
            "possession_status": getattr(filters.possession_status, "value", None),
            "trip_status": getattr(filters.trip_status, "value", None),
            "has_return": filters.has_return,
            "has_return_confirmation": filters.has_return_confirmation,
            "search_used": bool(filters.search),
            "search_length": len(filters.search) if filters.search else 0,
        }
