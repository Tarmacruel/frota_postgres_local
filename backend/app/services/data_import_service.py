from __future__ import annotations

import csv
import hashlib
import io
import re
import unicodedata
from datetime import date, datetime, time, timedelta, timezone
from decimal import Decimal
from pathlib import Path
from uuid import UUID

from fastapi import HTTPException, UploadFile, status
from sqlalchemy import Text, cast, func, or_, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload, selectinload

from app.core.config import settings
from app.models.data_import import (
    DataImportBatch,
    DataImportBatchStatus,
    DataImportEntityType,
    DataImportRow,
    DataImportRowStatus,
    DataImportSuggestedAction,
)
from app.models.driver import Driver, DriverLicenseCategory
from app.models.fine import Fine, FineInfraction, FineStatus
from app.models.location_history import LocationHistory
from app.models.master_data import Allocation, Department, Organization
from app.models.user import User
from app.models.vehicle import Vehicle, VehicleOwnershipType, VehicleStatus, VehicleType
from app.schemas.common import PaginatedResponse, build_pagination
from app.schemas.data_import import DataImportRowUpdate
from app.services.audit_service import AuditService

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - validated by requirements/tests.
    load_workbook = None


MISSING_VALUES = {"", "-", "N/A", "NA", "SELECIONE", "INDISPONIVEL", "INDISPONÍVEL", "NULL", "NONE"}
FAKE_EMAILS = {"XXXXX@XXXXX.COM", "xxxxx@xxxxx.COM"}
FAKE_PHONES = {"(00) 0 0000-0000", "(73) 0 0000-0000"}

VEHICLE_IMPORTABLE_FIELDS = ["plate", "chassis_number", "brand", "model", "vehicle_type", "ownership_type", "status", "allocation_id"]
VEHICLE_OFFICIAL_EXTRA_FIELDS = [
    "year",
    "prefix",
    "patrimonio_numero_frota",
    "renavam",
    "color",
    "fuel_type",
    "tank_capacity_liters",
    "transmission",
    "city",
    "state",
    "registered_detran",
    "engine_spec",
]
VEHICLE_TRIAGE_EXTRA_COLUMNS = [
    "Foto",
    "IPVA",
    "Seguro",
    "Licenciamento",
    "Km Restrição",
    "KM/Hora por L Mínimo",
    "KM/Hora por L Máximo",
    "Km do Primeiro Abastecimento",
    "Primeiro Abastecimento",
    "KM / Horímetro do Último Abastecimento",
    "Último Abastecimento",
    "Data do Cadastro",
    "KM/L",
    "Bloqueio Temporario",
    "Data de Inativação",
    "Data de Reativação",
]

DRIVER_IMPORTABLE_FIELDS = ["nome_completo", "documento", "organization_id", "contato", "email", "cnh_categoria", "cnh_validade", "ativo"]
DRIVER_OFFICIAL_EXTRA_FIELDS = [
    "registro",
    "matricula",
    "cargo",
    "cnh_numero",
    "rg",
    "data_nascimento",
    "data_emissao_cnh",
    "ultimo_abastecimento",
]
DRIVER_TRIAGE_EXTRA_COLUMNS = ["Unidade", "Subunidade", "Lotação Condutor"]

FINE_IMPORTABLE_FIELDS = [
    "vehicle_id",
    "infraction_type_id",
    "ticket_number",
    "infraction_date",
    "amount",
    "description",
    "location",
    "status",
]
FINE_OFFICIAL_EXTRA_FIELDS = [
    "infraction_time",
    "communication_number",
    "sent_date",
    "process_number",
    "source_status",
    "imported_driver_name",
    "notes",
]
FINE_TRIAGE_EXTRA_COLUMNS = ["RENAVAM", "VINCULO", "MODELO", "SECRETARIA", "TIPO"]


class DataImportService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.audit = AuditService(db)

    async def list_batches(self) -> list[dict]:
        result = await self.db.execute(select(DataImportBatch).order_by(DataImportBatch.created_at.desc()))
        return [self._serialize_batch(batch) for batch in result.scalars().all()]

    async def get_batch(self, batch_id: UUID) -> dict:
        batch = await self._get_batch(batch_id, with_rows=True)
        preview = list(batch.rows[:25])
        payload = self._serialize_batch(batch)
        payload["rows_preview"] = [self._serialize_row(row) for row in preview]
        return payload

    async def list_rows(
        self,
        batch_id: UUID,
        *,
        page: int,
        limit: int,
        row_status: DataImportRowStatus | None = None,
        has_conflict: bool | None = None,
        has_error: bool | None = None,
        search: str | None = None,
    ) -> PaginatedResponse[dict]:
        await self._get_batch(batch_id)
        stmt = select(DataImportRow).where(DataImportRow.batch_id == batch_id)
        count_stmt = select(func.count(DataImportRow.id)).where(DataImportRow.batch_id == batch_id)

        if row_status:
            stmt = stmt.where(DataImportRow.status == row_status)
            count_stmt = count_stmt.where(DataImportRow.status == row_status)
        if has_conflict is True:
            stmt = stmt.where(func.json_array_length(DataImportRow.conflicts) > 0)
            count_stmt = count_stmt.where(func.json_array_length(DataImportRow.conflicts) > 0)
        if has_error is True:
            stmt = stmt.where(func.json_array_length(DataImportRow.validation_errors) > 0)
            count_stmt = count_stmt.where(func.json_array_length(DataImportRow.validation_errors) > 0)
        if search and search.strip():
            term = f"%{search.strip()}%"
            search_clause = or_(
                cast(DataImportRow.row_number, Text).ilike(term),
                DataImportRow.matched_by.ilike(term),
                DataImportRow.manager_notes.ilike(term),
                cast(DataImportRow.raw_data, Text).ilike(term),
                cast(DataImportRow.mapped_data, Text).ilike(term),
                cast(DataImportRow.official_extra_data, Text).ilike(term),
                cast(DataImportRow.triage_extra_data, Text).ilike(term),
            )
            stmt = stmt.where(search_clause)
            count_stmt = count_stmt.where(search_clause)

        stmt = stmt.order_by(DataImportRow.row_number.asc()).offset((page - 1) * limit).limit(limit)
        total = int((await self.db.execute(count_stmt)).scalar_one())
        rows = list((await self.db.execute(stmt)).scalars().all())
        return PaginatedResponse[dict](data=[self._serialize_row(row) for row in rows], pagination=build_pagination(page, limit, total))

    async def upload(self, upload: UploadFile, current_user: User) -> dict:
        filename = upload.filename or "importacao"
        suffix = Path(filename).suffix.lower()
        if suffix not in {".xlsx", ".csv"}:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Envie arquivo XLSX ou CSV")

        content = await upload.read()
        await upload.close()
        if not content:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Arquivo de importacao vazio")

        parsed_rows = self._read_upload_rows(content, suffix)
        entity_type, header_index, header = self._detect_header(parsed_rows)
        raw_records = self._build_raw_records(parsed_rows, header_index, header)
        if not raw_records:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Arquivo nao possui linhas de dados")

        analysis_context = await self._build_analysis_context()
        key_counts = self._build_key_counts(entity_type, raw_records)
        mapped_rows = []
        for row_number, raw_data in raw_records:
            mapped_rows.append(await self._analyze_row(entity_type, row_number, raw_data, key_counts, analysis_context))

        batch = DataImportBatch(
            entity_type=entity_type,
            status=DataImportBatchStatus.ANALYZED,
            source_filename=filename,
            header_row_index=header_index + 1,
            detected_columns=header,
            importable_fields=self._importable_fields(entity_type),
            official_extra_fields=self._official_extra_fields(entity_type),
            triage_extra_fields=self._triage_extra_fields(entity_type),
            summary=self._build_summary(mapped_rows),
            created_by_id=current_user.id,
        )
        self.db.add(batch)
        await self.db.flush()

        stored_path = self._store_import_file(batch.id, filename, content)
        batch.stored_path = stored_path

        for item in mapped_rows:
            self.db.add(DataImportRow(batch_id=batch.id, **item))

        await self.audit.record(
            actor=current_user,
            action="CREATE",
            entity_type="DATA_IMPORT",
            entity_id=batch.id,
            entity_label=filename,
            details={"entity_type": entity_type.value, "summary": batch.summary},
        )
        await self.db.commit()
        return await self.get_batch(batch.id)

    async def update_row(self, batch_id: UUID, row_id: UUID, data: DataImportRowUpdate, current_user: User) -> dict:
        batch = await self._get_batch(batch_id)
        row = await self._get_row(batch_id, row_id)
        payload = data.model_dump(exclude_unset=True)

        if "status" in payload and payload["status"] is not None:
            row.status = payload["status"]
        for field in ("mapped_data", "official_extra_data", "triage_extra_data", "manager_notes"):
            if field in payload:
                setattr(row, field, payload[field])

        batch.status = DataImportBatchStatus.REVIEWING
        batch.summary = await self._rebuild_batch_summary(batch.id)
        await self.audit.record(
            actor=current_user,
            action="UPDATE",
            entity_type="DATA_IMPORT_ROW",
            entity_id=row.id,
            entity_label=f"{batch.source_filename} linha {row.row_number}",
            details={"status": row.status.value, "manager_notes": row.manager_notes},
        )
        await self.db.commit()
        return self._serialize_row(row)

    async def apply(self, batch_id: UUID, current_user: User) -> dict:
        batch = await self._get_batch(batch_id, with_rows=True)
        created = 0
        updated = 0
        errors = 0
        skipped = 0
        now = datetime.now(timezone.utc)

        for row in batch.rows:
            if row.status != DataImportRowStatus.APPROVED:
                skipped += 1
                continue
            try:
                result = await self._apply_row(batch, row, current_user)
                if result["action"] == "CREATE":
                    created += 1
                elif result["action"] == "UPDATE":
                    updated += 1
                row.status = DataImportRowStatus.APPLIED
                row.applied_result = result
                row.applied_at = now
            except Exception as exc:
                errors += 1
                row.status = DataImportRowStatus.ERROR
                row.validation_errors = [*row.validation_errors, str(exc)]

        batch.status = DataImportBatchStatus.APPLIED if errors == 0 else DataImportBatchStatus.REVIEWING
        batch.applied_by_id = current_user.id
        batch.applied_at = now
        batch.summary = self._build_summary(batch.rows)
        await self.audit.record(
            actor=current_user,
            action="UPDATE",
            entity_type="DATA_IMPORT",
            entity_id=batch.id,
            entity_label=batch.source_filename,
            details={"event": "APPLY_IMPORT", "created": created, "updated": updated, "errors": errors, "skipped": skipped},
        )
        try:
            await self.db.commit()
        except IntegrityError as exc:
            await self.db.rollback()
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Nao foi possivel aplicar o lote de importacao") from exc

        return {"batch_id": batch.id, "created": created, "updated": updated, "errors": errors, "skipped": skipped, "applied_at": now}

    async def export_batch_csv(self, batch_id: UUID) -> tuple[str, str]:
        batch = await self._get_batch(batch_id, with_rows=True)
        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=";")
        writer.writerow([
            "linha",
            "status",
            "acao_sugerida",
            "match",
            "match_por",
            "dados_importaveis",
            "campos_oficiais_novos",
            "extras_triagem",
            "conflitos",
            "erros",
            "observacao_gestor",
        ])
        for row in batch.rows:
            writer.writerow([
                row.row_number,
                row.status.value,
                row.suggested_action.value,
                str(row.matched_entity_id) if row.matched_entity_id else "",
                row.matched_by or "",
                row.mapped_data,
                row.official_extra_data,
                row.triage_extra_data,
                row.conflicts,
                row.validation_errors,
                row.manager_notes or "",
            ])
        return f"importacao-{batch.entity_type.value.lower()}-{batch.id}.csv", buffer.getvalue()

    def export_template_csv(self, entity_type: DataImportEntityType) -> tuple[str, str]:
        fields = (
            ["Placa", "Chassi", "Marca", "Modelo", "Tipo", "Tipo Frota", "Status", "Unidade", "Subunidade", "Ano", "Prefixo", "Patrimônio/Núm. Frota", "Renavam", "Cor", "Combustível", "Capacidade", "Transmissao", "Cidade Veículo", "Estado Veículo", "Registrado no DETRAN", "Motorização"]
            if entity_type == DataImportEntityType.VEHICLE
            else ["Nome", "CPF", "Unidade", "Telefone", "Celular", "Email", "Categoria", "Vencimento", "Status", "Registro", "Matricula", "Cargo", "CNH", "RG", "Data Nascimento", "Data Emissão CNH", "Ultimo Abastecimento"]
        )
        if entity_type == DataImportEntityType.FINE:
            fields = ["PLACA", "RENAVAM", "VINCULO", "MODELO", "SECRETARIA", "A. INFRA??O", "TIPO DA INFRA??O", "DATA ", "HORA", "LOCAL", "C.I.", "ENVIADO", "PROCESSO", "V.MULTA", "SITUA??O", "TIPO", "MOTORISTA", "OBS"]
        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=";")
        writer.writerow(fields)
        return f"modelo-importacao-{entity_type.value.lower()}.csv", buffer.getvalue()

    def _read_upload_rows(self, content: bytes, suffix: str) -> list[list]:
        if suffix == ".csv":
            text = content.decode("utf-8-sig", errors="replace")
            sample = text[:2048]
            dialect = csv.Sniffer().sniff(sample, delimiters=";,|\t") if sample.strip() else csv.excel
            return [list(row) for row in csv.reader(io.StringIO(text), dialect)]

        if load_workbook is None:
            raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Leitura XLSX indisponivel")
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        return [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]

    def _detect_header(self, rows: list[list]) -> tuple[DataImportEntityType, int, list[str]]:
        best = None
        for index, row in enumerate(rows[:30]):
            values = [self._clean_header(value) for value in row]
            normalized = {self._norm_key(value) for value in values if value}
            vehicle_score = len(normalized & {self._norm_key(value) for value in ("Placa", "Chassi", "Marca", "Modelo", "Tipo Frota")})
            driver_score = len(normalized & {self._norm_key(value) for value in ("Nome", "CPF", "CNH", "Categoria", "Vencimento")})
            fine_score = len(normalized & {self._norm_key(value) for value in ("Placa", "A. Infração", "Tipo da Infração", "Data", "V.Multa")})
            score = max(vehicle_score, driver_score, fine_score)
            if score and (best is None or score > best[0]):
                if fine_score >= vehicle_score and fine_score >= driver_score:
                    entity = DataImportEntityType.FINE
                else:
                    entity = DataImportEntityType.VEHICLE if vehicle_score >= driver_score else DataImportEntityType.DRIVER
                best = (score, entity, index, values)
        if not best or best[0] < 3:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Nao foi possivel detectar o cabecalho do relatorio")
        return best[1], best[2], best[3]

    def _build_raw_records(self, rows: list[list], header_index: int, header: list[str]) -> list[tuple[int, dict]]:
        records = []
        header_length = len(header)
        for row_offset, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            values = list(row) + [""] * max(0, header_length - len(row))
            raw = {
                header[column_index]: self._json_value(values[column_index])
                for column_index in range(header_length)
                if header[column_index]
            }
            if any(self._present(value) for value in raw.values()):
                records.append((row_offset, raw))
        return records

    async def _build_analysis_context(self) -> dict:
        vehicles = list((await self.db.execute(select(Vehicle))).scalars().all())
        drivers = list((await self.db.execute(select(Driver))).scalars().all())
        fines = list((await self.db.execute(select(Fine))).scalars().all())
        infractions = list((await self.db.execute(select(FineInfraction))).scalars().all())
        allocations = list(
            (
                await self.db.execute(
                    select(Allocation).options(joinedload(Allocation.department).joinedload(Department.organization))
                )
            )
            .scalars()
            .unique()
            .all()
        )
        organizations = list((await self.db.execute(select(Organization))).scalars().all())
        return {
            "vehicles_by_plate": {self._norm_key(vehicle.plate): vehicle for vehicle in vehicles if vehicle.plate},
            "vehicles_by_chassis": {self._norm_key(vehicle.chassis_number): vehicle for vehicle in vehicles if vehicle.chassis_number},
            "vehicles_by_renavam": {self._digits(vehicle.renavam): vehicle for vehicle in vehicles if vehicle.renavam},
            "drivers_by_document": {self._digits(driver.documento): driver for driver in drivers if driver.documento},
            "drivers_by_cnh": {self._digits(driver.cnh_numero): driver for driver in drivers if driver.cnh_numero},
            "drivers_by_name": {self._norm_lookup(driver.nome_completo): driver for driver in drivers if driver.nome_completo},
            "fines_by_ticket_vehicle": {
                f"{self._norm_key(fine.ticket_number)}|{fine.vehicle_id}": fine
                for fine in fines
                if fine.ticket_number and fine.vehicle_id
            },
            "infractions_by_code": {
                f"{self._norm_key(infraction.code)}|{self._norm_key(infraction.desdobramento)}": infraction
                for infraction in infractions
                if infraction.code
            },
            "infractions_by_description": {
                self._normalize_infraction_description(infraction.description): infraction
                for infraction in infractions
                if infraction.description
            },
            "infractions": infractions,
            "allocations": self._allocation_lookup(allocations),
            "organizations": self._organization_lookup(organizations),
        }

    def _build_key_counts(self, entity_type: DataImportEntityType, raw_records: list[tuple[int, dict]]) -> dict:
        keys = {"primary": {}, "secondary": {}, "fine_signatures": {}}
        for _row_number, raw in raw_records:
            if entity_type == DataImportEntityType.VEHICLE:
                primary = self._norm_key(raw.get("Placa"))
                secondary = self._norm_key(raw.get("Chassi"))
            elif entity_type == DataImportEntityType.DRIVER:
                primary = self._digits(raw.get("CPF"))
                secondary = self._digits(raw.get("CNH"))
            else:
                primary = self._fine_duplicate_key(raw)
                secondary = self._norm_key(raw.get("A. INFRAÇÃO"))
                if primary:
                    keys["fine_signatures"].setdefault(primary, set()).add(self._fine_signature(raw))
            keys["primary"][primary] = keys["primary"].get(primary, 0) + (1 if primary else 0)
            keys["secondary"][secondary] = keys["secondary"].get(secondary, 0) + (1 if secondary else 0)
        return keys

    async def _analyze_row(self, entity_type: DataImportEntityType, row_number: int, raw: dict, key_counts: dict, context: dict) -> dict:
        if entity_type == DataImportEntityType.VEHICLE:
            mapped, official_extra, triage_extra, errors, conflicts = self._map_vehicle(raw, key_counts, context)
            matched, matched_by = self._match_vehicle(mapped, context)
        elif entity_type == DataImportEntityType.DRIVER:
            mapped, official_extra, triage_extra, errors, conflicts = self._map_driver(raw, key_counts, context)
            matched, matched_by = self._match_driver(mapped, official_extra, context)
        else:
            mapped, official_extra, triage_extra, errors, conflicts = self._map_fine(raw, key_counts, context)
            matched, matched_by = self._match_fine(mapped, context)

        action = DataImportSuggestedAction.UPDATE if matched else DataImportSuggestedAction.CREATE
        if errors or any("duplicidade conflitante" in conflict.lower() for conflict in conflicts):
            action = DataImportSuggestedAction.REVIEW
        return {
            "row_number": row_number,
            "status": DataImportRowStatus.PENDING,
            "suggested_action": action,
            "matched_entity_id": matched.id if matched else None,
            "matched_by": matched_by,
            "raw_data": raw,
            "mapped_data": mapped,
            "official_extra_data": official_extra,
            "triage_extra_data": triage_extra,
            "conflicts": conflicts,
            "validation_errors": errors,
        }

    def _map_vehicle(self, raw: dict, key_counts: dict, context: dict) -> tuple[dict, dict, dict, list[str], list[str]]:
        errors = []
        conflicts = []
        plate = self._normalize_plate(raw.get("Placa"))
        chassis = self._normalize_text(raw.get("Chassi"), uppercase=True)
        mapped = {
            "plate": plate,
            "chassis_number": chassis,
            "brand": self._normalize_text(raw.get("Marca"), uppercase=True),
            "model": self._normalize_text(raw.get("Modelo"), uppercase=True),
            "vehicle_type": self._map_vehicle_type(raw.get("Tipo"), conflicts),
            "ownership_type": self._map_ownership(raw.get("Tipo Frota"), conflicts),
            "status": self._map_vehicle_status(raw.get("Status")),
        }
        allocation_id = self._resolve_allocation(raw.get("Subunidade"), raw.get("Unidade"), context)
        if allocation_id:
            mapped["allocation_id"] = str(allocation_id)

        for field in ("plate", "brand", "model", "vehicle_type", "ownership_type", "status"):
            if not mapped.get(field):
                errors.append(f"Campo obrigatorio ausente: {field}")
        if not mapped.get("allocation_id"):
            errors.append("Lotacao nao encontrada; selecione uma lotacao antes de aprovar")
        if plate and key_counts["primary"].get(self._norm_key(plate), 0) > 1:
            conflicts.append("Placa duplicada no arquivo")
        if chassis and key_counts["secondary"].get(self._norm_key(chassis), 0) > 1:
            conflicts.append("Chassi duplicado no arquivo")

        official_extra = {
            "year": self._normalize_text(raw.get("Ano"), uppercase=True),
            "prefix": self._normalize_text(raw.get("Prefixo"), uppercase=True),
            "patrimonio_numero_frota": self._normalize_text(raw.get("Patrimônio/Núm. Frota"), uppercase=True),
            "renavam": self._normalize_text(raw.get("Renavam"), uppercase=True),
            "color": self._normalize_text(raw.get("Cor"), uppercase=True),
            "fuel_type": self._normalize_text(raw.get("Combustível"), uppercase=True),
            "tank_capacity_liters": self._parse_float(raw.get("Capacidade")),
            "transmission": self._normalize_text(raw.get("Transmissao"), uppercase=True),
            "city": self._normalize_text(raw.get("Cidade Veículo"), uppercase=True),
            "state": self._normalize_state(raw.get("Estado Veículo")),
            "registered_detran": self._parse_bool(raw.get("Registrado no DETRAN")),
            "engine_spec": self._normalize_text(raw.get("Motorização"), uppercase=True),
        }
        official_extra = {key: value for key, value in official_extra.items() if value is not None}
        triage_extra = {column: raw.get(column) for column in VEHICLE_TRIAGE_EXTRA_COLUMNS if self._present(raw.get(column))}
        return mapped, official_extra, triage_extra, errors, conflicts

    def _map_driver(self, raw: dict, key_counts: dict, context: dict) -> tuple[dict, dict, dict, list[str], list[str]]:
        errors = []
        conflicts = []
        document = self._normalize_document(raw.get("CPF"))
        mapped = {
            "nome_completo": self._normalize_text(raw.get("Nome"), uppercase=True),
            "documento": document,
            "contato": self._best_contact(raw.get("Celular"), raw.get("Telefone")),
            "email": self._normalize_email(raw.get("Email")),
            "cnh_categoria": self._map_driver_category(raw.get("Categoria")),
            "cnh_validade": self._parse_date(raw.get("Vencimento")),
            "ativo": self._map_active(raw.get("Status")),
        }
        organization_id = self._resolve_organization(raw.get("Unidade"), context)
        if organization_id:
            mapped["organization_id"] = str(organization_id)

        for field in ("nome_completo", "documento", "cnh_categoria"):
            if not mapped.get(field):
                errors.append(f"Campo obrigatorio ausente: {field}")
        if not mapped.get("organization_id"):
            errors.append("Secretaria nao encontrada; selecione uma secretaria antes de aprovar")
        if document and key_counts["primary"].get(self._digits(document), 0) > 1:
            conflicts.append("CPF duplicado no arquivo")
        cnh_numero = self._normalize_text(raw.get("CNH"), uppercase=True)
        if cnh_numero and key_counts["secondary"].get(self._digits(cnh_numero), 0) > 1:
            conflicts.append("CNH duplicada no arquivo")

        official_extra = {
            "registro": self._normalize_text(raw.get("Registro"), uppercase=True),
            "matricula": self._normalize_text(raw.get("Matricula"), uppercase=True),
            "cargo": self._normalize_text(raw.get("Cargo"), uppercase=True),
            "cnh_numero": cnh_numero,
            "rg": self._normalize_text(raw.get("RG"), uppercase=True),
            "data_nascimento": self._parse_date(raw.get("Data Nascimento")),
            "data_emissao_cnh": self._parse_date(raw.get("Data Emissão CNH")),
            "ultimo_abastecimento": self._parse_datetime(raw.get("Ultimo Abastecimento")),
        }
        official_extra = {key: value for key, value in official_extra.items() if value is not None}
        triage_extra = {column: raw.get(column) for column in DRIVER_TRIAGE_EXTRA_COLUMNS if self._present(raw.get(column))}
        return mapped, official_extra, triage_extra, errors, conflicts

    def _map_fine(self, raw: dict, key_counts: dict, context: dict) -> tuple[dict, dict, dict, list[str], list[str]]:
        errors = []
        conflicts = []
        plate = self._normalize_plate(self._raw_value(raw, "PLACA"))
        renavam = self._normalize_text(self._raw_value(raw, "RENAVAM"), uppercase=True)
        ticket_number = self._normalize_text(self._raw_value(raw, "A. INFRAÇÃO", "AUTO", "AUTO INFRAÇÃO"), uppercase=True)
        description = self._normalize_text(self._raw_value(raw, "TIPO DA INFRAÇÃO", "DESCRIÇÃO"), uppercase=True)
        infraction_date = self._parse_date(self._raw_value(raw, "DATA", "DATA "))
        amount = self._parse_money(self._raw_value(raw, "V.MULTA", "VALOR", "VALOR MULTA"))
        source_status = self._normalize_text(self._raw_value(raw, "SITUAÇÃO"), uppercase=True)

        vehicle = self._match_import_vehicle(plate, renavam, context)
        infraction = self._match_import_infraction(raw, description, context)
        driver = self._match_import_driver(self._raw_value(raw, "MOTORISTA"), context)

        mapped = {
            "ticket_number": ticket_number,
            "infraction_date": infraction_date,
            "amount": amount,
            "description": description or (infraction.description if infraction else "INFRAÇÃO NÃO INFORMADA NA IMPORTAÇÃO"),
            "location": self._normalize_text(self._raw_value(raw, "LOCAL"), uppercase=True),
            "status": self._map_fine_status(source_status),
        }
        if vehicle:
            mapped["vehicle_id"] = str(vehicle.id)
        elif plate:
            mapped["provisional_vehicle"] = {
                "plate": plate,
                "renavam": renavam,
                "model": self._normalize_text(self._raw_value(raw, "MODELO"), uppercase=True) or "IMPORTADO",
                "ownership_type": self._map_import_vehicle_ownership(self._raw_value(raw, "VINCULO")),
                "vehicle_type": self._map_vehicle_type(self._raw_value(raw, "TIPO"), conflicts),
                "organization_name": self._normalize_text(self._raw_value(raw, "SECRETARIA"), uppercase=True),
            }
            conflicts.append("Veiculo nao encontrado; sera criado como provisório ao aplicar")

        if infraction:
            mapped["infraction_type_id"] = str(infraction.id)
        else:
            provisional_description = description or "INFRAÇÃO NÃO INFORMADA NA IMPORTAÇÃO"
            mapped["provisional_infraction"] = {
                "description": provisional_description,
                "source": "Importacao de multas",
            }
            conflicts.append("Enquadramento nao encontrado; sera criado como provisório ao aplicar")

        if driver:
            mapped["driver_id"] = str(driver.id)

        for field in ("ticket_number", "infraction_date", "amount"):
            if not mapped.get(field):
                errors.append(f"Campo obrigatorio ausente: {field}")
        if not vehicle and not plate:
            errors.append("Campo obrigatorio ausente: plate")

        duplicate_key = self._fine_duplicate_key(raw)
        if duplicate_key and key_counts["primary"].get(duplicate_key, 0) > 1:
            signatures = key_counts.get("fine_signatures", {}).get(duplicate_key, set())
            if len(signatures) > 1:
                conflicts.append("Duplicidade conflitante no arquivo; revise antes de aprovar")
            else:
                conflicts.append("Duplicidade exata no arquivo; aplicar uma linha atualizara a mesma multa")

        official_extra = {
            "infraction_time": self._parse_time(self._raw_value(raw, "HORA")),
            "communication_number": self._normalize_text(self._raw_value(raw, "C.I.", "CI"), uppercase=True),
            "sent_date": self._parse_date(self._raw_value(raw, "ENVIADO")),
            "process_number": self._normalize_text(self._raw_value(raw, "PROCESSO"), uppercase=True),
            "source_status": source_status,
            "imported_driver_name": self._normalize_text(self._raw_value(raw, "MOTORISTA"), uppercase=True),
            "notes": self._normalize_text(self._raw_value(raw, "OBS"), uppercase=True),
        }
        official_extra = {key: value for key, value in official_extra.items() if value is not None}
        triage_extra = {column: self._raw_value(raw, column) for column in FINE_TRIAGE_EXTRA_COLUMNS if self._present(self._raw_value(raw, column))}
        return mapped, official_extra, triage_extra, errors, conflicts

    async def _apply_row(self, batch: DataImportBatch, row: DataImportRow, current_user: User) -> dict:
        errors = self._apply_validation_errors(batch.entity_type, row.mapped_data)
        if errors:
            raise ValueError("; ".join(errors))

        if batch.entity_type == DataImportEntityType.VEHICLE:
            return await self._apply_vehicle_row(row, current_user)
        if batch.entity_type == DataImportEntityType.DRIVER:
            return await self._apply_driver_row(row, current_user)
        return await self._apply_fine_row(row, current_user)

    async def _apply_vehicle_row(self, row: DataImportRow, current_user: User) -> dict:
        data = {**row.mapped_data, **row.official_extra_data}
        vehicle = await self._find_vehicle_for_apply(data)
        action = "UPDATE" if vehicle else "CREATE"
        allocation_id = UUID(str(data["allocation_id"]))

        if vehicle is None:
            vehicle = Vehicle(
                plate=data["plate"],
                chassis_number=data.get("chassis_number"),
                brand=data["brand"],
                model=data["model"],
                vehicle_type=VehicleType(data["vehicle_type"]),
                ownership_type=VehicleOwnershipType(data["ownership_type"]),
                status=VehicleStatus(data["status"]),
            )
            self._assign_vehicle_extra(vehicle, data)
            self.db.add(vehicle)
            await self.db.flush()
            self.db.add(LocationHistory(vehicle_id=vehicle.id, allocation_id=allocation_id, department="Importacao de dados"))
        else:
            for field in ("plate", "chassis_number", "brand", "model"):
                if data.get(field):
                    setattr(vehicle, field, data[field])
            for field, enum_class in (("vehicle_type", VehicleType), ("ownership_type", VehicleOwnershipType), ("status", VehicleStatus)):
                if data.get(field):
                    setattr(vehicle, field, enum_class(data[field]))
            self._assign_vehicle_extra(vehicle, data)
            active = await self._get_active_vehicle_history(vehicle.id)
            if not active or active.allocation_id != allocation_id:
                if active:
                    active.end_date = datetime.now(timezone.utc)
                self.db.add(LocationHistory(vehicle_id=vehicle.id, allocation_id=allocation_id, department="Importacao de dados"))

        await self.audit.record(
            actor=current_user,
            action="CREATE" if action == "CREATE" else "UPDATE",
            entity_type="VEHICLE",
            entity_id=vehicle.id,
            entity_label=vehicle.plate,
            details={"event": "DATA_IMPORT", "row_id": str(row.id), "action": action, "data": data},
        )
        await self.db.flush()
        return {"action": action, "entity_id": str(vehicle.id), "label": vehicle.plate}

    async def _apply_driver_row(self, row: DataImportRow, current_user: User) -> dict:
        data = {**row.mapped_data, **row.official_extra_data}
        driver = await self._find_driver_for_apply(data)
        action = "UPDATE" if driver else "CREATE"
        payload = self._driver_payload_for_model(data)

        if driver is None:
            driver = Driver(**payload)
            self.db.add(driver)
        else:
            for field, value in payload.items():
                setattr(driver, field, value)

        await self.db.flush()
        await self.audit.record(
            actor=current_user,
            action="CREATE" if action == "CREATE" else "UPDATE",
            entity_type="DRIVER",
            entity_id=driver.id,
            entity_label=driver.nome_completo,
            details={"event": "DATA_IMPORT", "row_id": str(row.id), "action": action, "data": data},
        )
        await self.db.flush()
        return {"action": action, "entity_id": str(driver.id), "label": driver.nome_completo}

    async def _apply_fine_row(self, row: DataImportRow, current_user: User) -> dict:
        data = {**row.mapped_data, **row.official_extra_data}
        vehicle = await self._find_or_create_fine_vehicle(data)
        infraction = await self._find_or_create_fine_infraction(data)
        fine = await self._find_fine_for_apply(data, vehicle.id)
        action = "UPDATE" if fine else "CREATE"

        payload = {
            "vehicle_id": vehicle.id,
            "driver_id": UUID(str(data["driver_id"])) if data.get("driver_id") else None,
            "infraction_type_id": infraction.id,
            "ticket_number": data["ticket_number"],
            "infraction_date": self._date_from_iso(data["infraction_date"]),
            "infraction_time": self._time_from_iso(data.get("infraction_time")),
            "due_date": self._date_from_iso(data.get("due_date")),
            "amount": Decimal(str(data["amount"])),
            "description": data.get("description") or infraction.description,
            "location": data.get("location"),
            "status": FineStatus(data.get("status") or FineStatus.PENDENTE.value),
            "communication_number": data.get("communication_number"),
            "sent_date": self._date_from_iso(data.get("sent_date")),
            "process_number": data.get("process_number"),
            "source_status": data.get("source_status"),
            "imported_driver_name": data.get("imported_driver_name"),
            "notes": data.get("notes"),
            "source_import_row_id": row.id,
        }

        if fine is None:
            fine = Fine(created_by=current_user.id, **payload)
            self.db.add(fine)
        else:
            for field, value in payload.items():
                setattr(fine, field, value)

        await self.db.flush()
        await self.audit.record(
            actor=current_user,
            action="CREATE" if action == "CREATE" else "UPDATE",
            entity_type="FINE",
            entity_id=fine.id,
            entity_label=f"{vehicle.plate} - {fine.ticket_number}",
            details={"event": "DATA_IMPORT", "row_id": str(row.id), "action": action, "data": data},
        )
        await self.db.flush()
        return {"action": action, "entity_id": str(fine.id), "label": f"{vehicle.plate} - {fine.ticket_number}"}

    def _assign_vehicle_extra(self, vehicle: Vehicle, data: dict) -> None:
        for field in VEHICLE_OFFICIAL_EXTRA_FIELDS:
            if field not in data:
                continue
            target_field = {
                "color": "color",
                "fuel_type": "fuel_type",
                "tank_capacity_liters": "tank_capacity_liters",
                "engine_spec": "engine_spec",
            }.get(field, field)
            setattr(vehicle, target_field, data[field])

    def _driver_payload_for_model(self, data: dict) -> dict:
        payload = {
            "nome_completo": data["nome_completo"],
            "documento": data["documento"],
            "organization_id": UUID(str(data["organization_id"])),
            "contato": data.get("contato"),
            "email": data.get("email"),
            "cnh_categoria": DriverLicenseCategory(data["cnh_categoria"]),
            "cnh_validade": self._date_from_iso(data.get("cnh_validade")),
            "ativo": bool(data.get("ativo", True)),
        }
        for field in DRIVER_OFFICIAL_EXTRA_FIELDS:
            if field not in data:
                continue
            if field in {"data_nascimento", "data_emissao_cnh"}:
                payload[field] = self._date_from_iso(data[field])
            elif field == "ultimo_abastecimento":
                payload[field] = self._datetime_from_iso(data[field])
            else:
                payload[field] = data[field]
        return payload

    def _apply_validation_errors(self, entity_type: DataImportEntityType, data: dict) -> list[str]:
        if entity_type == DataImportEntityType.VEHICLE:
            required = ("plate", "brand", "model", "vehicle_type", "ownership_type", "status", "allocation_id")
            return [f"Campo obrigatorio ausente: {field}" for field in required if not data.get(field)]
        if entity_type == DataImportEntityType.DRIVER:
            required = ("nome_completo", "documento", "organization_id", "cnh_categoria")
            return [f"Campo obrigatorio ausente: {field}" for field in required if not data.get(field)]
        errors = [f"Campo obrigatorio ausente: {field}" for field in ("ticket_number", "infraction_date", "amount") if not data.get(field)]
        if not data.get("vehicle_id") and not data.get("provisional_vehicle"):
            errors.append("Campo obrigatorio ausente: vehicle_id")
        if not data.get("infraction_type_id") and not data.get("provisional_infraction"):
            errors.append("Campo obrigatorio ausente: infraction_type_id")
        return errors

    async def _find_vehicle_for_apply(self, data: dict) -> Vehicle | None:
        if data.get("plate"):
            vehicle = (await self.db.execute(select(Vehicle).where(Vehicle.plate == data["plate"]))).scalar_one_or_none()
            if vehicle:
                return vehicle
        if data.get("chassis_number"):
            return (await self.db.execute(select(Vehicle).where(Vehicle.chassis_number == data["chassis_number"]))).scalar_one_or_none()
        return None

    async def _find_driver_for_apply(self, data: dict) -> Driver | None:
        if data.get("documento"):
            driver = (await self.db.execute(select(Driver).where(Driver.documento == data["documento"]))).scalar_one_or_none()
            if driver:
                return driver
        if data.get("cnh_numero"):
            return (await self.db.execute(select(Driver).where(Driver.cnh_numero == data["cnh_numero"]))).scalar_one_or_none()
        return None

    async def _find_fine_for_apply(self, data: dict, vehicle_id: UUID) -> Fine | None:
        if not data.get("ticket_number"):
            return None
        return (
            await self.db.execute(
                select(Fine).where(Fine.ticket_number == data["ticket_number"], Fine.vehicle_id == vehicle_id)
            )
        ).scalar_one_or_none()

    async def _find_or_create_fine_vehicle(self, data: dict) -> Vehicle:
        if data.get("vehicle_id"):
            vehicle = (await self.db.execute(select(Vehicle).where(Vehicle.id == UUID(str(data["vehicle_id"]))))).scalar_one_or_none()
            if vehicle:
                return vehicle

        provisional = data.get("provisional_vehicle") or {}
        plate = self._normalize_plate(provisional.get("plate"))
        if not plate:
            raise ValueError("Veiculo provisório sem placa")
        existing = (await self.db.execute(select(Vehicle).where(Vehicle.plate == plate))).scalar_one_or_none()
        if existing:
            return existing

        model_text = provisional.get("model") or "IMPORTADO"
        brand, model = self._split_import_vehicle_model(model_text)
        vehicle = Vehicle(
            plate=plate,
            renavam=provisional.get("renavam"),
            brand=brand,
            model=model,
            vehicle_type=VehicleType(provisional.get("vehicle_type") or VehicleType.SEDAN.value),
            ownership_type=VehicleOwnershipType(provisional.get("ownership_type") or VehicleOwnershipType.PROPRIO.value),
            status=VehicleStatus.ATIVO,
            is_provisional=True,
            provisional_source="Importacao de multas",
        )
        self.db.add(vehicle)
        await self.db.flush()
        return vehicle

    async def _find_or_create_fine_infraction(self, data: dict) -> FineInfraction:
        if data.get("infraction_type_id"):
            infraction = (
                await self.db.execute(select(FineInfraction).where(FineInfraction.id == UUID(str(data["infraction_type_id"]))))
            ).scalar_one_or_none()
            if infraction:
                return infraction

        provisional = data.get("provisional_infraction") or {}
        description = provisional.get("description") or data.get("description") or "INFRAÇÃO NÃO INFORMADA NA IMPORTAÇÃO"
        normalized = self._normalize_infraction_description(description)
        existing = (
            await self.db.execute(
                select(FineInfraction).where(FineInfraction.normalized_description == normalized).order_by(FineInfraction.is_provisional.asc())
            )
        ).scalars().first()
        if existing:
            return existing

        code_hash = hashlib.sha1(normalized.encode("utf-8")).hexdigest()[:10].upper()
        infraction = FineInfraction(
            code=f"PROV-{code_hash}",
            desdobramento="0",
            description=description,
            normalized_description=normalized,
            default_amount=Decimal(str(data["amount"])) if data.get("amount") else None,
            is_active=True,
            is_official=False,
            is_provisional=True,
            source=provisional.get("source") or "Importacao de multas",
        )
        self.db.add(infraction)
        await self.db.flush()
        return infraction

    async def _get_active_vehicle_history(self, vehicle_id: UUID) -> LocationHistory | None:
        return (
            await self.db.execute(
                select(LocationHistory)
                .where(LocationHistory.vehicle_id == vehicle_id, LocationHistory.end_date.is_(None))
                .order_by(LocationHistory.start_date.desc())
            )
        ).scalar_one_or_none()

    def _match_vehicle(self, mapped: dict, context: dict):
        if mapped.get("plate"):
            match = context["vehicles_by_plate"].get(self._norm_key(mapped["plate"]))
            if match:
                return match, "plate"
        if mapped.get("chassis_number"):
            match = context["vehicles_by_chassis"].get(self._norm_key(mapped["chassis_number"]))
            if match:
                return match, "chassis_number"
        return None, None

    def _match_driver(self, mapped: dict, official_extra: dict, context: dict):
        if mapped.get("documento"):
            match = context["drivers_by_document"].get(self._digits(mapped["documento"]))
            if match:
                return match, "documento"
        if official_extra.get("cnh_numero"):
            match = context["drivers_by_cnh"].get(self._digits(official_extra["cnh_numero"]))
            if match:
                return match, "cnh_numero"
        return None, None

    def _match_fine(self, mapped: dict, context: dict):
        vehicle_id = mapped.get("vehicle_id")
        ticket_number = mapped.get("ticket_number")
        if vehicle_id and ticket_number:
            match = context["fines_by_ticket_vehicle"].get(f"{self._norm_key(ticket_number)}|{vehicle_id}")
            if match:
                return match, "ticket_number_vehicle"
        return None, None

    def _match_import_vehicle(self, plate: str | None, renavam: str | None, context: dict) -> Vehicle | None:
        if plate:
            match = context["vehicles_by_plate"].get(self._norm_key(plate))
            if match:
                return match
        if renavam:
            return context["vehicles_by_renavam"].get(self._digits(renavam))
        return None

    def _match_import_driver(self, value, context: dict) -> Driver | None:
        name = self._normalize_text(value, uppercase=True)
        if not name:
            return None
        ignored = {"MOTORISTA NAO IDENTIFICADO", "MOTORISTA NÃO IDENTIFICADO"}
        if name in ignored or "MULTA ORIGINARIA" in name:
            return None
        return context["drivers_by_name"].get(self._norm_lookup(name))

    def _match_import_infraction(self, raw: dict, description: str | None, context: dict) -> FineInfraction | None:
        code = self._normalize_text(self._raw_value(raw, "CODIGO", "CÓDIGO", "CODIGO INFRACAO", "CÓDIGO INFRAÇÃO"), uppercase=True)
        desdobramento = self._normalize_text(self._raw_value(raw, "DESDOBRAMENTO", "DESDOB"), uppercase=True) or "0"
        if code:
            match = context["infractions_by_code"].get(f"{self._norm_key(code)}|{self._norm_key(desdobramento)}")
            if match:
                return match
            for key, infraction in context["infractions_by_code"].items():
                if key.startswith(f"{self._norm_key(code)}|"):
                    return infraction

        normalized = self._normalize_infraction_description(description)
        if not normalized:
            return None
        aliases = (
            ("TRANS VELOC SUP ATE 20", "VELOCIDADE SUPERIOR MAXIMA PERMITIDA EM ATE 20"),
            ("TRANS VELOC SUP 20", "VELOCIDADE SUPERIOR MAXIMA PERMITIDA EM ATE 20"),
            ("VELOCIDADE SUP 20 ATE 50", "VELOCIDADE SUPERIOR MAXIMA PERMITIDA EM MAIS DE 20 ATE 50"),
            ("VELOCIDADE SUP ATE 50", "VELOCIDADE SUPERIOR MAXIMA PERMITIDA EM MAIS DE 20 ATE 50"),
            ("TRANS VELOC SUP 50", "VELOCIDADE SUPERIOR MAXIMA PERMITIDA EM MAIS DE 50"),
            ("AVANCAR O SINAL VERMELHO", "AVANCAR O SINAL VERMELHO"),
            ("AVANCAR SINAL VERMELHO", "AVANCAR O SINAL VERMELHO"),
            ("CONDUTOR SEM CINTO", "DEIXAR O CONDUTOR DE USAR O CINTO SEGURANCA"),
            ("PASSAGEIRO SEM CINTO", "DEIXAR O PASSAGEIRO DE USAR O CINTO SEGURANCA"),
        )
        candidates = [normalized]
        for source, target in aliases:
            if source in normalized:
                candidates.append(target)
        for candidate in candidates:
            exact = context["infractions_by_description"].get(candidate)
            if exact:
                return exact
            for key, infraction in context["infractions_by_description"].items():
                if candidate in key or key in candidate:
                    return infraction
        return None

    def _allocation_lookup(self, allocations: list[Allocation]) -> dict[str, UUID]:
        lookup = {}
        for allocation in allocations:
            values = [
                allocation.name,
                allocation.display_name,
                allocation.department.name if allocation.department else None,
                allocation.organization_name,
            ]
            for value in values:
                key = self._norm_lookup(value)
                if key:
                    lookup[key] = allocation.id
        return lookup

    def _organization_lookup(self, organizations: list[Organization]) -> dict[str, UUID]:
        return {self._norm_lookup(organization.name): organization.id for organization in organizations if self._norm_lookup(organization.name)}

    def _resolve_allocation(self, subunit, unit, context: dict) -> UUID | None:
        for value in (subunit, unit):
            key = self._norm_lookup(value)
            if key in context["allocations"]:
                return context["allocations"][key]
        return None

    def _resolve_organization(self, unit, context: dict) -> UUID | None:
        key = self._norm_lookup(unit)
        return context["organizations"].get(key)

    def _build_summary(self, rows) -> dict:
        row_list = list(rows)
        statuses = {}
        actions = {}
        for row in row_list:
            status_value = row.status.value if hasattr(row, "status") else row["status"].value
            action_value = row.suggested_action.value if hasattr(row, "suggested_action") else row["suggested_action"].value
            statuses[status_value] = statuses.get(status_value, 0) + 1
            actions[action_value] = actions.get(action_value, 0) + 1
        conflicts = sum(1 for row in row_list if (row.conflicts if hasattr(row, "conflicts") else row["conflicts"]))
        errors = sum(1 for row in row_list if (row.validation_errors if hasattr(row, "validation_errors") else row["validation_errors"]))
        return {"total_rows": len(row_list), "statuses": statuses, "actions": actions, "conflicts": conflicts, "errors": errors}

    async def _rebuild_batch_summary(self, batch_id: UUID) -> dict:
        rows = list((await self.db.execute(select(DataImportRow).where(DataImportRow.batch_id == batch_id))).scalars().all())
        return self._build_summary(rows)

    def _store_import_file(self, batch_id: UUID, filename: str, content: bytes) -> str:
        safe_name = re.sub(r"[^A-Za-z0-9._-]+", "-", filename).strip(".-") or "importacao"
        relative = Path("data_imports") / f"{batch_id}-{safe_name}"
        absolute = Path(settings.STORAGE_DIR) / relative
        absolute.parent.mkdir(parents=True, exist_ok=True)
        absolute.write_bytes(content)
        return relative.as_posix()

    async def _get_batch(self, batch_id: UUID, *, with_rows: bool = False) -> DataImportBatch:
        stmt = select(DataImportBatch).where(DataImportBatch.id == batch_id)
        if with_rows:
            stmt = stmt.options(selectinload(DataImportBatch.rows))
        batch = (await self.db.execute(stmt)).scalar_one_or_none()
        if not batch:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Lote de importacao nao encontrado")
        return batch

    async def _get_row(self, batch_id: UUID, row_id: UUID) -> DataImportRow:
        row = (
            await self.db.execute(select(DataImportRow).where(DataImportRow.batch_id == batch_id, DataImportRow.id == row_id))
        ).scalar_one_or_none()
        if not row:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Linha de importacao nao encontrada")
        return row

    def _serialize_batch(self, batch: DataImportBatch) -> dict:
        return {
            "id": batch.id,
            "entity_type": batch.entity_type,
            "status": batch.status,
            "source_filename": batch.source_filename,
            "header_row_index": batch.header_row_index,
            "detected_columns": batch.detected_columns or [],
            "importable_fields": batch.importable_fields or [],
            "official_extra_fields": batch.official_extra_fields or [],
            "triage_extra_fields": batch.triage_extra_fields or [],
            "summary": batch.summary or {},
            "notes": batch.notes,
            "created_by_id": batch.created_by_id,
            "applied_by_id": batch.applied_by_id,
            "created_at": batch.created_at,
            "updated_at": batch.updated_at,
            "applied_at": batch.applied_at,
        }

    def _serialize_row(self, row: DataImportRow) -> dict:
        return {
            "id": row.id,
            "batch_id": row.batch_id,
            "row_number": row.row_number,
            "status": row.status,
            "suggested_action": row.suggested_action,
            "matched_entity_id": row.matched_entity_id,
            "matched_by": row.matched_by,
            "raw_data": row.raw_data or {},
            "mapped_data": row.mapped_data or {},
            "official_extra_data": row.official_extra_data or {},
            "triage_extra_data": row.triage_extra_data or {},
            "conflicts": row.conflicts or [],
            "validation_errors": row.validation_errors or [],
            "manager_notes": row.manager_notes,
            "applied_result": row.applied_result,
            "created_at": row.created_at,
            "updated_at": row.updated_at,
            "applied_at": row.applied_at,
        }

    def _clean_header(self, value) -> str:
        return str(value or "").strip()

    def _json_value(self, value):
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, date):
            return value.isoformat()
        return str(value).strip() if not isinstance(value, (int, float, bool)) else value

    def _present(self, value) -> bool:
        normalized = str(value or "").strip()
        return normalized.upper() not in MISSING_VALUES

    def _raw_value(self, raw: dict, *headers: str):
        normalized_raw = {self._norm_key(key): value for key, value in raw.items()}
        for header in headers:
            key = self._norm_key(header)
            if key in normalized_raw:
                return normalized_raw[key]
        return None

    def _normalize_text(self, value, *, uppercase: bool = False) -> str | None:
        if not self._present(value):
            return None
        normalized = str(value).strip()
        if normalized in FAKE_EMAILS or normalized in FAKE_PHONES:
            return None
        return normalized.upper() if uppercase else normalized

    def _normalize_plate(self, value) -> str | None:
        text = self._normalize_text(value, uppercase=True)
        return re.sub(r"[^A-Z0-9-]", "", text) if text else None

    def _normalize_document(self, value) -> str | None:
        text = self._normalize_text(value, uppercase=True)
        return text if text else None

    def _normalize_email(self, value) -> str | None:
        text = self._normalize_text(value)
        if not text or text in FAKE_EMAILS or "@" not in text:
            return None
        return text.lower()

    def _best_contact(self, *values) -> str | None:
        for value in values:
            text = self._normalize_text(value)
            if text and text not in FAKE_PHONES:
                return text
        return None

    def _normalize_state(self, value) -> str | None:
        text = self._normalize_text(value, uppercase=True)
        return text[:2] if text else None

    def _norm_key(self, value) -> str:
        return re.sub(r"[^A-Z0-9]+", "", self._strip_accents(str(value or "").upper()))

    def _norm_lookup(self, value) -> str:
        text = self._normalize_text(value, uppercase=True)
        if not text:
            return ""
        text = re.sub(r"\s*/\s*\d+.*$", "", text).strip()
        return self._norm_key(text)

    def _strip_accents(self, value: str) -> str:
        return "".join(ch for ch in unicodedata.normalize("NFD", value) if unicodedata.category(ch) != "Mn")

    def _digits(self, value) -> str:
        return re.sub(r"\D+", "", str(value or ""))

    def _parse_float(self, value) -> float | None:
        if not self._present(value):
            return None
        try:
            return float(str(value).replace(",", "."))
        except ValueError:
            return None

    def _parse_money(self, value) -> float | None:
        if not self._present(value):
            return None
        if isinstance(value, (int, float, Decimal)):
            return float(value)
        text = re.sub(r"[^0-9,.-]+", "", str(value))
        if "," in text and "." in text:
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", ".")
        try:
            return float(text)
        except ValueError:
            return None

    def _parse_bool(self, value) -> bool | None:
        if not self._present(value):
            return None
        text = str(value).strip().upper()
        if text in {"1", "SIM", "TRUE", "S"}:
            return True
        if text in {"0", "NAO", "NÃO", "FALSE", "N"}:
            return False
        return None

    def _parse_date(self, value) -> str | None:
        parsed = self._parse_datetime_object(value)
        return parsed.date().isoformat() if parsed else None

    def _parse_datetime(self, value) -> str | None:
        parsed = self._parse_datetime_object(value)
        return parsed.isoformat() if parsed else None

    def _parse_time(self, value) -> str | None:
        if not self._present(value):
            return None
        if isinstance(value, time):
            return value.isoformat(timespec="minutes")
        if isinstance(value, datetime):
            return value.time().isoformat(timespec="minutes")
        text = str(value).strip()
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                return datetime.strptime(text, fmt).time().isoformat(timespec="minutes")
            except ValueError:
                continue
        return None

    def _parse_datetime_object(self, value) -> datetime | None:
        if not self._present(value):
            return None
        if isinstance(value, datetime):
            return value.replace(tzinfo=value.tzinfo or timezone.utc)
        if isinstance(value, date):
            return datetime.combine(value, time.min, tzinfo=timezone.utc)
        text = str(value).strip()
        if re.fullmatch(r"\d+(\.\d+)?", text):
            try:
                return datetime(1899, 12, 30, tzinfo=timezone.utc) + timedelta(days=float(text))
            except ValueError:
                return None
        for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y", "%Y-%m-%d", "%Y-%m-%dT%H:%M:%S"):
            try:
                parsed = datetime.strptime(text, fmt)
                return parsed.replace(tzinfo=timezone.utc)
            except ValueError:
                continue
        return None

    def _date_from_iso(self, value) -> date | None:
        if not value:
            return None
        return date.fromisoformat(str(value)[:10])

    def _time_from_iso(self, value) -> time | None:
        if not value:
            return None
        return time.fromisoformat(str(value)[:8])

    def _datetime_from_iso(self, value) -> datetime | None:
        if not value:
            return None
        parsed = datetime.fromisoformat(str(value))
        return parsed.replace(tzinfo=parsed.tzinfo or timezone.utc)

    def _map_fine_status(self, value) -> str:
        text = self._normalize_text(value, uppercase=True)
        if text in {"PAGA", "PAGO"}:
            return FineStatus.PAGA.value
        if text == "DEFERIDA":
            return FineStatus.DEFERIDA.value
        if text == "RECURSO":
            return FineStatus.RECURSO.value
        return FineStatus.PENDENTE.value

    def _map_import_vehicle_ownership(self, value) -> str:
        text = self._normalize_text(value, uppercase=True)
        if text and any(token in text for token in ("LOCALIZA", "LOCA", "LOCADO")):
            return VehicleOwnershipType.LOCADO.value
        if text and any(token in text for token in ("SESAB", "CEDID", "CONVENIO", "CONVÊNIO")):
            return VehicleOwnershipType.CEDIDO.value
        return VehicleOwnershipType.PROPRIO.value

    def _normalize_infraction_description(self, value) -> str:
        text = self._strip_accents(str(value or "").upper())
        return re.sub(r"[^A-Z0-9]+", " ", text).strip()

    def _fine_duplicate_key(self, raw: dict) -> str:
        ticket = self._norm_key(self._raw_value(raw, "A. INFRAÇÃO", "AUTO", "AUTO INFRAÇÃO"))
        plate = self._norm_key(self._raw_value(raw, "PLACA"))
        date_value = self._parse_date(self._raw_value(raw, "DATA", "DATA ")) or ""
        return "|".join(value for value in (ticket, plate, date_value) if value)

    def _fine_signature(self, raw: dict) -> str:
        values = [
            self._fine_duplicate_key(raw),
            self._norm_key(self._raw_value(raw, "TIPO DA INFRAÇÃO")),
            str(self._parse_money(self._raw_value(raw, "V.MULTA", "VALOR", "VALOR MULTA")) or ""),
            self._norm_key(self._raw_value(raw, "SITUAÇÃO")),
        ]
        return "|".join(values)

    def _split_import_vehicle_model(self, model_text: str) -> tuple[str, str]:
        text = self._normalize_text(model_text, uppercase=True) or "IMPORTADO"
        if "/" in text:
            brand, model = text.split("/", 1)
            return brand.strip() or "IMPORTADO", model.strip() or text
        parts = text.split(maxsplit=1)
        if len(parts) == 2:
            return parts[0], parts[1]
        return "IMPORTADO", text

    def _importable_fields(self, entity_type: DataImportEntityType) -> list[str]:
        if entity_type == DataImportEntityType.VEHICLE:
            return VEHICLE_IMPORTABLE_FIELDS
        if entity_type == DataImportEntityType.DRIVER:
            return DRIVER_IMPORTABLE_FIELDS
        return FINE_IMPORTABLE_FIELDS

    def _official_extra_fields(self, entity_type: DataImportEntityType) -> list[str]:
        if entity_type == DataImportEntityType.VEHICLE:
            return VEHICLE_OFFICIAL_EXTRA_FIELDS
        if entity_type == DataImportEntityType.DRIVER:
            return DRIVER_OFFICIAL_EXTRA_FIELDS
        return FINE_OFFICIAL_EXTRA_FIELDS

    def _triage_extra_fields(self, entity_type: DataImportEntityType) -> list[str]:
        if entity_type == DataImportEntityType.VEHICLE:
            return VEHICLE_TRIAGE_EXTRA_COLUMNS
        if entity_type == DataImportEntityType.DRIVER:
            return DRIVER_TRIAGE_EXTRA_COLUMNS
        return FINE_TRIAGE_EXTRA_COLUMNS

    def _map_vehicle_type(self, value, conflicts: list[str]) -> str:
        text = self._normalize_text(value, uppercase=True)
        mapping = {
            "SEDAN": VehicleType.SEDAN.value,
            "HATCH": VehicleType.HATCH.value,
            "LEVE": VehicleType.SEDAN.value,
            "PESADO": VehicleType.CAMINHAO.value,
            "CAMINHÃO": VehicleType.CAMINHAO.value,
            "CAMINHAO": VehicleType.CAMINHAO.value,
            "MOTO": VehicleType.MOTOCICLETA.value,
            "MOTOCICLETA": VehicleType.MOTOCICLETA.value,
            "MÁQUINA": VehicleType.MAQUINA.value,
            "MAQUINA": VehicleType.MAQUINA.value,
            "TRATOR": VehicleType.MAQUINA.value,
            "ROÇADEIRA": VehicleType.MAQUINA.value,
            "MOTONIVELADORA": VehicleType.MAQUINA.value,
            "PERFURADOR DE SOLO": VehicleType.MAQUINA.value,
            "MICRO ÔNIBUS": VehicleType.MICRO_ONIBUS.value,
            "MICRO ONIBUS": VehicleType.MICRO_ONIBUS.value,
            "ÔNIBUS": VehicleType.ONIBUS.value,
            "ONIBUS": VehicleType.ONIBUS.value,
            "VAN": VehicleType.VAN.value,
            "AMBULÂNCIA": VehicleType.VAN.value,
            "AMBULANCIA": VehicleType.VAN.value,
            "VIATURA": VehicleType.SUV.value,
        }
        if not text:
            return VehicleType.SEDAN.value
        result = mapping.get(text)
        if not result:
            conflicts.append(f"Tipo de veiculo externo sem mapeamento exato: {text}")
            return VehicleType.SEDAN.value
        return result

    def _map_ownership(self, value, conflicts: list[str]) -> str:
        text = self._normalize_text(value, uppercase=True)
        if text in {"LOCADO"}:
            return VehicleOwnershipType.LOCADO.value
        if text in {"CEDIDA", "CEDIDO", "CONVENIO", "CONVÊNIO", "ESTADUAL"}:
            return VehicleOwnershipType.CEDIDO.value
        if text in {"PROPRIO", "PRÓPRIO"}:
            return VehicleOwnershipType.PROPRIO.value
        if text:
            conflicts.append(f"Tipo de frota externo tratado como proprio: {text}")
        return VehicleOwnershipType.PROPRIO.value

    def _map_vehicle_status(self, value) -> str:
        text = self._normalize_text(value, uppercase=True)
        if text == "INATIVO":
            return VehicleStatus.INATIVO.value
        if text in {"MANUTENCAO", "MANUTENÇÃO"}:
            return VehicleStatus.MANUTENCAO.value
        return VehicleStatus.ATIVO.value

    def _map_driver_category(self, value) -> str | None:
        text = self._normalize_text(value, uppercase=True)
        if text in {category.value for category in DriverLicenseCategory}:
            return text
        return None

    def _map_active(self, value) -> bool:
        text = self._normalize_text(value, uppercase=True)
        return text != "INATIVO"
